import pandas as pd
import re
import io

# ==============================
# TAMAÑO POR DEFECTO DE IMÁGENES
# ==============================
DEFAULT_IMAGE_WIDTH_PT  = 300
DEFAULT_IMAGE_HEIGHT_PT = 200

# ==============================
# MAPEO DE TIPOS DESDE EL EXCEL
# ==============================
TIPO_MAP = {
    "texto":          "text",
    "número":         "text",
    "numero":         "text",
    "condicional":    "text",
    "imagen":         "image",
    "tabla":          "table",
    "loop":           "loop",
    "imagen / tabla": "image",  # se trata como imagen por ahora
}

# ==============================
# VERBOSE / LOGGING
# ==============================
VERBOSE = False

def _log(*args, **kwargs):
    """Imprime solo si VERBOSE está activado."""
    if VERBOSE:
        print(*args, **kwargs)


# ==============================
# CARGA DEL DICCIONARIO DESDE EXCEL
# ==============================
def cargar_diccionario(archivo, project_filter=None):
    """
    Lee el Excel del diccionario y lo convierte al esquema unificado:

    {
        "alias": {
            "placeholder": "{{ Nombre del dato }}",
            "type": "text" | "image" | "table",
            "value": <valor según tipo>
        }
    }

    Para type="text"  → value es string o None
    Para type="image" → value es {"file_id": ..., "width_pt": ..., "height_pt": ...}
    Para type="table" → value es None (Fase 3)

    Columnas opcionales en el Excel:
        "width_pt" y "height_pt" para imágenes — si no existen, se usan los defaults.
    """

    if isinstance(archivo, str):
        if archivo.endswith(".csv"):
            df = pd.read_csv(archivo)
        else:
            df = pd.read_excel(archivo, header=None)
    else:
        df = pd.read_excel(archivo, header=None)

    # Detectar fila de encabezado: buscar la primera fila con 'Nombre del dato'
    header_idx = 0
    for i, row in df.iterrows():
        if any(str(v).strip().lower() == "nombre del dato" for v in row.values):
            header_idx = i
            break

    df.columns = df.iloc[header_idx]
    df = df[header_idx + 1:].reset_index(drop=True)

    # Normalizar nombres de columnas
    df.columns = [str(col).strip().lower() for col in df.columns]

    # DEBUG: detectar columnas duplicadas
    from collections import Counter
    conteo = Counter(df.columns)
    duplicadas = [col for col, n in conteo.items() if n > 1]
    if duplicadas:
        print(f"⚠️  Columnas duplicadas encontradas: {duplicadas}")
    else:
        print("✅ No hay columnas duplicadas")
    print(f"   Todas las columnas: {list(df.columns)}")

    df = df.rename(columns={
        "nombre del dato": "placeholder",
        "valor":           "value",
        "tipo":            "type",
        "alias":           "alias",
    })

    tiene_width  = "width_pt"  in df.columns
    tiene_height = "height_pt" in df.columns
    tiene_sheet      = "sheet"      in df.columns
    tiene_header_row = "header_row" in df.columns

    data = {}

    for _, row in df.iterrows():
        placeholder = str(row.get("placeholder", "")).strip()
        raw_value   = row.get("value")
        raw_type    = str(row.get("type", "")).strip().lower()
        alias       = str(row.get("alias", "")).strip()

        if not placeholder and not alias:
            continue

        key_limpia = limpiar_placeholder(placeholder)
        final_key  = alias if alias else key_limpia

        tipo_normalizado = TIPO_MAP.get(raw_type, "text")

        if tipo_normalizado == "text":
            value = _convertir_a_texto(raw_value)

        elif tipo_normalizado == "image":
            file_id = str(raw_value).strip() if not _es_nulo(raw_value) else None
            width_pt  = _leer_dim(row, "width_pt")  if tiene_width  else DEFAULT_IMAGE_WIDTH_PT
            height_pt = _leer_dim(row, "height_pt") if tiene_height else DEFAULT_IMAGE_HEIGHT_PT
            value = {
                "file_id":   file_id,
                "width_pt":  width_pt,
                "height_pt": height_pt,
            }

        elif tipo_normalizado == "table":
            raw_url = str(raw_value).strip() if not _es_nulo(raw_value) else None
            sheet      = str(row.get("sheet", "")).strip() if tiene_sheet else None
            sheet      = sheet if sheet and sheet.lower() != "nan" else None
            header_row = None
            if tiene_header_row:
                hr = row.get("header_row")
                if not _es_nulo(hr):
                    try:
                        header_row = int(float(hr))
                    except Exception:
                        header_row = None
            value = {"file_id": raw_url, "sheet": sheet, "header_row": header_row} if raw_url else None

        elif tipo_normalizado == "loop":
            # file_id: URL al Excel con múltiples hojas
            # sheet:   prefijo que deben tener los nombres de las hojas a incluir
            # header_row: fila de encabezados (igual que en tabla)
            raw_url = str(raw_value).strip() if not _es_nulo(raw_value) else None
            prefijo = str(row.get("sheet", "")).strip() if tiene_sheet else None
            prefijo = prefijo if prefijo and prefijo.lower() != "nan" else None
            header_row = None
            if tiene_header_row:
                hr = row.get("header_row")
                if not _es_nulo(hr):
                    try:
                        header_row = int(float(hr))
                    except Exception:
                        header_row = None
            value = {"file_id": raw_url, "prefijo": prefijo, "header_row": header_row} if raw_url else None

        else:
            value = _convertir_a_texto(raw_value)

        data[final_key] = {
            "placeholder": placeholder,
            "type":        tipo_normalizado,
            "value":       value,
        }

    return data


# ==============================
# HELPERS INTERNOS
# ==============================

def limpiar_placeholder(texto):
    """Convierte '{{ Voltaje }}' → 'Voltaje'"""
    return re.sub(r"[{}]", "", texto).strip()


def _es_nulo(valor):
    """Devuelve True si el valor es NaN, None o string vacío."""
    try:
        return pd.isna(valor)
    except Exception:
        return valor is None or str(valor).strip() == ""


def _convertir_a_texto(valor):
    """Convierte el valor de celda a string, respetando None."""
    if _es_nulo(valor):
        return None
    if isinstance(valor, float) and valor == int(valor):
        return str(int(valor))
    return str(valor)


def _leer_dim(row, col_name):
    """Lee una dimensión numérica de una fila; devuelve el default si está vacía."""
    val = row.get(col_name)
    if _es_nulo(val):
        return DEFAULT_IMAGE_WIDTH_PT if col_name == "width_pt" else DEFAULT_IMAGE_HEIGHT_PT
    try:
        return int(float(val))
    except Exception:
        return DEFAULT_IMAGE_WIDTH_PT if col_name == "width_pt" else DEFAULT_IMAGE_HEIGHT_PT


# ==============================
# REEMPLAZO DE TEXTO EN GOOGLE DOCS
# ==============================

def reemplazar_textos(doc_id, diccionario, docs_service):
    """
    Reemplaza todos los placeholders de tipo 'text' en un Google Doc,
    preservando el formato existente del documento.

    Usa batchUpdate con replaceAllText, que respeta el formato del párrafo
    donde vive el placeholder (fuente, tamaño, color, negrita, etc.).

    Args:
        doc_id       (str):  ID del documento de Google Docs.
        diccionario  (dict): Diccionario unificado generado por cargar_diccionario().
        docs_service:        Servicio autenticado de Google Docs API.

    Returns:
        dict: {"reemplazados": [...], "omitidos": [...]}
    """
    reemplazados = []
    omitidos     = []
    requests     = []

    for alias, entrada in diccionario.items():

        if entrada.get("type") != "text":
            continue

        placeholder = entrada.get("placeholder", "").strip()
        value       = entrada.get("value")

        if not placeholder:
            omitidos.append({"alias": alias, "razon": "placeholder vacío"})
            continue

        texto_reemplazo = "" if value is None else str(value)

        requests.append({
            "replaceAllText": {
                "containsText": {
                    "text":      placeholder,
                    "matchCase": True
                },
                "replaceText": texto_reemplazo
            }
        })

        reemplazados.append({
            "alias":       alias,
            "placeholder": placeholder,
            "value":       texto_reemplazo
        })

    if not requests:
        _log("⚠️  No se encontraron entradas de tipo 'text' para reemplazar.")
        return {"reemplazados": [], "omitidos": omitidos}

    resultado = docs_service.documents().batchUpdate(
        documentId=doc_id,
        body={"requests": requests}
    ).execute()

    # Asociar ocurrencias reales (las replies vienen en el mismo orden que los requests)
    for i, respuesta in enumerate(resultado.get("replies", [])):
        ocurrencias = respuesta.get("replaceAllText", {}).get("occurrencesChanged", 0)
        reemplazados[i]["ocurrencias"] = ocurrencias
        if ocurrencias == 0:
            omitidos.append({
                "alias": reemplazados[i]["alias"],
                "razon": "placeholder no encontrado en el documento"
            })

    reemplazados_encontrados = [r for r in reemplazados if r.get("ocurrencias", 0) > 0]

    _log(f"✅ Texto reemplazado: {len(reemplazados_encontrados)} placeholders")
    for r in reemplazados_encontrados:
        _log(f"   {r['placeholder']} → '{r['value']}' ({r['ocurrencias']} ocurrencia/s)")

    return {
        "reemplazados": reemplazados_encontrados,
        "omitidos":     omitidos
    }


# ==============================
# CONDICIONALES EN GOOGLE DOCS
# ==============================

def procesar_condicionales(doc_id, diccionario, docs_service):
    """
    Procesa bloques condicionales {% if variable %}...{% endif %} en un Google Doc.

    - Si la variable es True  → elimina solo las etiquetas, conserva el contenido
    - Si la variable es False → elimina el bloque completo incluyendo etiquetas
    - Soporta {% else %} opcional
    - Opera de atrás hacia adelante para no desplazar índices

    Args:
        doc_id       (str):  ID del documento de Google Docs.
        diccionario  (dict): Diccionario unificado generado por cargar_diccionario().
        docs_service:        Servicio autenticado de Google Docs API.

    Returns:
        dict: {"procesados": [...], "omitidos": [...]}
    """
    import re

    # -------------------------------------------------------
    # 1. Obtener el documento y reconstruir texto plano
    #    con mapa de índices reales de la Docs API
    # -------------------------------------------------------
    documento = docs_service.documents().get(documentId=doc_id).execute()
    contenido = documento.get("body", {}).get("content", [])

    texto_plano, mapa_indices = _extraer_texto_con_indices(contenido)

    # -------------------------------------------------------
    # 2. Encontrar todos los bloques {% if %}...{% endif %}
    # -------------------------------------------------------
    patron = re.compile(
        r"\{%-?\s*if\s+(\w+)\s*-?%\}"  # {% if variable %}
        r"(.*?)"                         # contenido interno
        r"\{%-?\s*endif\s*-?%\}",       # {% endif %}
        re.DOTALL
    )

    bloques = list(patron.finditer(texto_plano))

    if not bloques:
        _log("ℹ️  No se encontraron bloques condicionales en el documento.")
        return {"procesados": [], "omitidos": []}

    procesados        = []
    omitidos          = []
    rangos_a_eliminar = []  # lista de (start_doc, end_doc) en índices de Docs API

    for bloque in bloques:
        nombre_variable  = bloque.group(1).strip()
        contenido_bloque = bloque.group(2)

        # Buscar la variable en el diccionario
        entrada = diccionario.get(nombre_variable)
        if entrada is None:
            omitidos.append({
                "variable": nombre_variable,
                "razon":    "variable no encontrada en el diccionario"
            })
            continue

        valor        = entrada.get("value")
        es_verdadero = _evaluar_booleano(valor)

        # Posiciones en el texto plano
        inicio_bloque = bloque.start()    # inicio de {% if %}
        fin_bloque    = bloque.end()      # fin de {% endif %}
        fin_if        = bloque.start(2)   # fin de la etiqueta {% if variable %}
        inicio_endif  = bloque.end(2)     # inicio de {% endif %}

        # Detectar si hay {% else %}
        patron_else = re.compile(r"\{%-?\s*else\s*-?%\}")
        match_else  = patron_else.search(contenido_bloque)

        if match_else:
            inicio_else_abs = bloque.start(2) + match_else.start()
            fin_else_abs    = bloque.start(2) + match_else.end()

            if es_verdadero:
                # Conservar rama if → eliminar etiqueta {% if %} y desde {% else %} hasta fin
                rangos_a_eliminar.append((_idx(inicio_bloque, mapa_indices), _idx(fin_if, mapa_indices)))
                rangos_a_eliminar.append((_idx(inicio_else_abs, mapa_indices), _idx(fin_bloque, mapa_indices)))
            else:
                # Conservar rama else → eliminar desde {% if %} hasta fin de {% else %}, y {% endif %}
                rangos_a_eliminar.append((_idx(inicio_bloque, mapa_indices), _idx(fin_else_abs, mapa_indices)))
                rangos_a_eliminar.append((_idx(inicio_endif, mapa_indices), _idx(fin_bloque, mapa_indices)))

        else:
            if es_verdadero:
                # Conservar contenido → eliminar solo las dos etiquetas
                rangos_a_eliminar.append((_idx(inicio_bloque, mapa_indices), _idx(fin_if, mapa_indices)))
                rangos_a_eliminar.append((_idx(inicio_endif, mapa_indices), _idx(fin_bloque, mapa_indices)))
            else:
                # Eliminar bloque completo
                rangos_a_eliminar.append((_idx(inicio_bloque, mapa_indices), _idx(fin_bloque, mapa_indices)))

        procesados.append({
            "variable":     nombre_variable,
            "valor":        valor,
            "es_verdadero": es_verdadero,
            "tiene_else":   match_else is not None
        })

    # -------------------------------------------------------
    # 3. Ejecutar eliminaciones de atrás hacia adelante
    # -------------------------------------------------------
    if rangos_a_eliminar:
        rangos_validos = sorted(
            [(s, e) for s, e in rangos_a_eliminar if s is not None and e is not None and s < e],
            key=lambda r: r[0],
            reverse=True
        )

        requests = [
            {"deleteContentRange": {"range": {"startIndex": s, "endIndex": e}}}
            for s, e in rangos_validos
        ]

        if requests:
            docs_service.documents().batchUpdate(
                documentId=doc_id,
                body={"requests": requests}
            ).execute()

    # -------------------------------------------------------
    # 4. Log de resultados
    # -------------------------------------------------------
    _log(f"✅ Condicionales procesadas: {len(procesados)}")
    for p in procesados:
        icono    = "✔" if p["es_verdadero"] else "✘"
        else_str = " (con else)" if p["tiene_else"] else ""
        _log(f"   {icono} {{% if {p['variable']} %}} → {p['valor']}{else_str}")

    return {"procesados": procesados, "omitidos": omitidos}


# ==============================
# HELPERS INTERNOS
# ==============================

def _extraer_texto_con_indices(contenido):
    """
    Recorre el body del documento y construye:
    - texto_plano  (str):  todo el texto concatenado
    - mapa_indices (list): posición N en texto_plano → índice real en Docs API
    """
    texto_plano  = ""
    mapa_indices = []

    def recorrer(elementos):
        nonlocal texto_plano, mapa_indices
        for elemento in elementos:
            if "paragraph" in elemento:
                for run in elemento["paragraph"].get("elements", []):
                    contenido_run = run.get("textRun", {}).get("content", "")
                    start_index   = run.get("startIndex", 0)
                    for i, char in enumerate(contenido_run):
                        mapa_indices.append(start_index + i)
                        texto_plano += char
            elif "table" in elemento:
                for fila in elemento["table"].get("tableRows", []):
                    for celda in fila.get("tableCells", []):
                        recorrer(celda.get("content", []))

    recorrer(contenido)
    return texto_plano, mapa_indices


def _idx(pos, mapa_indices):
    """Convierte posición en texto plano a índice real de Docs API."""
    if pos < 0 or pos >= len(mapa_indices):
        return None
    return mapa_indices[pos]


def _evaluar_booleano(valor):
    """Convierte el valor de la variable condicional a booleano."""
    if valor is None:
        return False
    if isinstance(valor, bool):
        return valor
    return str(valor).strip().lower() in ("true", "1", "si", "yes", "sí")


# ==============================
# IMÁGENES EN GOOGLE DOCS
# ==============================

def reemplazar_imagenes(doc_id, diccionario, docs_service, drive_service):
    """
    Reemplaza placeholders de tipo 'image' en un Google Doc.

    Flujo por cada imagen:
        1. Hace el archivo de Drive temporalmente público
        2. Localiza el placeholder en el documento y obtiene su índice
        3. Borra el texto del placeholder
        4. Inserta la imagen en esa posición usando la URL pública
        5. Revoca el acceso público inmediatamente

    Opera de atrás hacia adelante para no desplazar índices.

    Args:
        doc_id        (str):  ID del documento de Google Docs.
        diccionario   (dict): Diccionario unificado generado por cargar_diccionario().
        docs_service:         Servicio autenticado de Google Docs API.
        drive_service:        Servicio autenticado de Google Drive API.

    Returns:
        dict: {"reemplazados": [...], "omitidos": [...]}
    """
    reemplazados = []
    omitidos     = []

    # Filtrar solo entradas de tipo image
    entradas_imagen = [
        (alias, entrada)
        for alias, entrada in diccionario.items()
        if entrada.get("type") == "image"
    ]

    if not entradas_imagen:
        _log("ℹ️  No se encontraron entradas de tipo 'image' para reemplazar.")
        return {"reemplazados": [], "omitidos": omitidos}

    for alias, entrada in entradas_imagen:
        placeholder = entrada.get("placeholder", "").strip()
        valor       = entrada["value"]
        raw_file_id = valor.get("file_id")
        if not raw_file_id:
            omitidos.append({"alias": alias, "razon": "file_id vacío o no definido"})
            continue
        file_id = extraer_id_gdoc(raw_file_id)
        width_pt    = valor.get("width_pt",  DEFAULT_IMAGE_WIDTH_PT)
        height_pt   = valor.get("height_pt", DEFAULT_IMAGE_HEIGHT_PT)

        if not placeholder:
            omitidos.append({"alias": alias, "razon": "placeholder vacío"})
            continue

        permission_id = None

        try:
            # -------------------------------------------------
            # 1. Hacer el archivo público temporalmente
            # -------------------------------------------------
            permiso = drive_service.permissions().create(
                fileId=file_id,
                body={"type": "anyone", "role": "reader"},
                fields="id"
            ).execute()
            permission_id = permiso.get("id")

            url_publica = f"https://drive.google.com/uc?export=download&id={file_id}"

            # -------------------------------------------------
            # 2. Leer el documento y localizar el placeholder
            # -------------------------------------------------
            documento = docs_service.documents().get(documentId=doc_id).execute()
            contenido = documento.get("body", {}).get("content", [])
            texto_plano, mapa_indices = _extraer_texto_con_indices(contenido)

            pos = texto_plano.find(placeholder)
            if pos == -1:
                omitidos.append({"alias": alias, "razon": "placeholder no encontrado en el documento"})
                continue

            # Índices reales en la Docs API
            start_idx = _idx(pos, mapa_indices)
            end_idx   = _idx(pos + len(placeholder), mapa_indices)

            if start_idx is None or end_idx is None:
                omitidos.append({"alias": alias, "razon": "no se pudo mapear el índice del placeholder"})
                continue

            # -------------------------------------------------
            # 3. Borrar el placeholder e insertar la imagen
            #    en un solo batchUpdate
            # -------------------------------------------------
            docs_service.documents().batchUpdate(
                documentId=doc_id,
                body={"requests": [
                    # Primero borrar el texto del placeholder
                    {
                        "deleteContentRange": {
                            "range": {
                                "startIndex": start_idx,
                                "endIndex":   end_idx
                            }
                        }
                    },
                    # Luego insertar la imagen en la misma posición
                    {
                        "insertInlineImage": {
                            "location":  {"index": start_idx},
                            "uri":       url_publica,
                            "objectSize": {
                                "width":  {"magnitude": width_pt,  "unit": "PT"},
                                "height": {"magnitude": height_pt, "unit": "PT"}
                            }
                        }
                    }
                ]}
            ).execute()

            reemplazados.append({
                "alias":       alias,
                "placeholder": placeholder,
                "file_id":     file_id,
                "width_pt":    width_pt,
                "height_pt":   height_pt
            })

        except Exception as e:
            omitidos.append({
                "alias": alias,
                "razon": f"error al procesar: {str(e)}"
            })

        finally:
            # -------------------------------------------------
            # 4. Revocar permiso público siempre, incluso si hubo error
            # -------------------------------------------------
            if permission_id:
                try:
                    drive_service.permissions().delete(
                        fileId=file_id,
                        permissionId=permission_id
                    ).execute()
                except Exception:
                    pass  # Si falla la revocación no interrumpimos el flujo

    # Log de resultados
    _log(f"✅ Imágenes reemplazadas: {len(reemplazados)}")
    for r in reemplazados:
        _log(f"   {r['placeholder']} → {r['file_id']} ({r['width_pt']}x{r['height_pt']} pt)")

    omitidos_reales = [o for o in omitidos if o.get("razon") != "placeholder no encontrado en el documento"]
    if omitidos_reales:
        _log(f"⚠️  Imágenes omitidas: {len(omitidos_reales)}")
        for o in omitidos_reales:
            _log(f"   {o['alias']}: {o['razon']}")

    return {"reemplazados": reemplazados, "omitidos": omitidos}


# ==============================
# TABLAS EN GOOGLE DOCS
# ==============================







# ==============================
# TABLAS EN GOOGLE DOCS
# ==============================

def reemplazar_tablas(doc_id, diccionario, docs_service, drive_service):
    """
    Reemplaza placeholders de tipo 'table' en un Google Doc.

    Diseño de requests optimizado — por cada tabla:
        GET 1  : leer documento y localizar placeholder
        batch 1: deleteContentRange + insertTable  (1 request)
        GET 2  : localizar tabla recién insertada
        batch 2: todos los mergeTableCells en un solo batch (ordenados abajo-derecha → arriba-izquierda)
        GET 3  : leer estructura post-merge para obtener índices reales de celdas
        batch 3: todos los insertText en un solo batch
        batch 4: todos los updateTableCellStyle + updateTextStyle en un solo batch
                 (reusar GET 3 con offset conocido, sin GET extra)

    Total: 3 GETs + 4 batchUpdates por tabla (independiente del nº de celdas/merges).
    Antes: 3 + 3·N GETs + 2 + N batchUpdates donde N = nº de merges.

    Args:
        doc_id        (str):  ID del documento de Google Docs.
        diccionario   (dict): Diccionario unificado generado por cargar_diccionario().
        docs_service:         Servicio autenticado de Google Docs API.
        drive_service:        Servicio autenticado de Google Drive API.

    Returns:
        dict: {"reemplazados": [...], "omitidos": [...]}
    """
    reemplazados = []
    omitidos     = []

    entradas_tabla = [
        (alias, entrada)
        for alias, entrada in diccionario.items()
        if entrada.get("type") == "table"
        and entrada.get("value") is not None
    ]

    if not entradas_tabla:
        _log("ℹ️  No se encontraron entradas de tipo 'table' para reemplazar.")
        return {"reemplazados": [], "omitidos": omitidos}

    for alias, entrada in entradas_tabla:
        placeholder = entrada.get("placeholder", "").strip()
        valor       = entrada["value"]
        url_tabla   = valor.get("file_id")
        sheet_name  = valor.get("sheet")
        header_row  = valor.get("header_row")

        if not url_tabla:
            omitidos.append({"alias": alias, "razon": "URL de tabla vacía"})
            continue

        if not placeholder:
            omitidos.append({"alias": alias, "razon": "placeholder vacío"})
            continue

        try:
            # -------------------------------------------------
            # 1. Descargar datos y formato desde Excel
            # -------------------------------------------------
            df, cell_formats, merges = _cargar_datos_tabla(
                url_tabla, sheet_name, drive_service, header_row,
                alias=alias, placeholder=placeholder,
            )
            if df is None or df.empty:
                omitidos.append({"alias": alias, "razon": "no se pudieron cargar los datos"})
                continue

            n_filas = len(df) + 1   # +1 por fila de encabezados
            n_cols  = len(df.columns)

            # -------------------------------------------------
            # GET 1: localizar el placeholder
            # -------------------------------------------------
            documento   = docs_service.documents().get(documentId=doc_id).execute()
            contenido   = documento.get("body", {}).get("content", [])
            texto_plano, mapa_indices = _extraer_texto_con_indices(contenido)

            pos = texto_plano.find(placeholder)
            if pos == -1:
                omitidos.append({"alias": alias, "razon": f"placeholder '{placeholder}' no encontrado en el documento"})
                continue

            start_idx = _idx(pos, mapa_indices)
            end_idx   = _idx(pos + len(placeholder), mapa_indices)

            if start_idx is None or end_idx is None:
                omitidos.append({"alias": alias, "razon": "no se pudo mapear el índice del placeholder"})
                continue

            inicio_bloque = start_idx

            # -------------------------------------------------
            # batch 1: borrar placeholder + insertar tabla vacía
            # -------------------------------------------------
            docs_service.documents().batchUpdate(
                documentId=doc_id,
                body={"requests": [
                    {"deleteContentRange": {"range": {"startIndex": start_idx, "endIndex": end_idx}}},
                    {"insertTable": {"rows": n_filas, "columns": n_cols, "location": {"index": start_idx}}}
                ]}
            ).execute()

            # -------------------------------------------------
            # GET 2: localizar la tabla recién insertada
            # -------------------------------------------------
            doc_tmp       = docs_service.documents().get(documentId=doc_id).execute()
            contenido_tmp = doc_tmp.get("body", {}).get("content", [])
            tabla_tmp     = None
            for elem in contenido_tmp:
                if "table" in elem and elem.get("startIndex") == inicio_bloque:
                    t = elem["table"]
                    t["startIndex"] = elem["startIndex"]
                    t["endIndex"]   = elem.get("endIndex", 0)
                    tabla_tmp = t
                    break
            if tabla_tmp is None:
                tabla_tmp = _encontrar_tabla_cerca(contenido_tmp, inicio_bloque)
            if not tabla_tmp:
                omitidos.append({"alias": alias, "razon": "no se pudo localizar la tabla recién insertada"})
                continue

            tabla_start_real = tabla_tmp["startIndex"]

            # -------------------------------------------------
            # batch 2: todos los merges en UN solo batchUpdate
            #
            # La API de Google Docs acepta múltiples mergeTableCells en un
            # mismo batchUpdate siempre que estén ordenados de abajo-derecha
            # hacia arriba-izquierda (sin solapamiento de rangos de destino).
            # Esto elimina el loop de N GETs + N batchUpdates del diseño anterior.
            # -------------------------------------------------
            if merges:
                merges_ordenados = sorted(merges, key=lambda m: (-m["min_row"], -m["min_col"]))
                merge_requests   = [
                    {
                        "mergeTableCells": {
                            "tableRange": {
                                "tableCellLocation": {
                                    "tableStartLocation": {"index": tabla_start_real},
                                    "rowIndex":    m["min_row"],
                                    "columnIndex": m["min_col"]
                                },
                                "rowSpan":    m["row_span"],
                                "columnSpan": m["col_span"]
                            }
                        }
                    }
                    for m in merges_ordenados
                ]
                try:
                    docs_service.documents().batchUpdate(
                        documentId=doc_id,
                        body={"requests": merge_requests}
                    ).execute()
                except Exception as e_merge:
                    # Si el batch de merges falla (raro, p.ej. merges solapados en el Excel)
                    # caemos al modo secuencial como fallback para no perder la tabla.
                    _log(f"   ⚠️  Batch de merges falló ({e_merge}), reintentando de forma secuencial…")
                    for m in merges_ordenados:
                        try:
                            docs_service.documents().batchUpdate(
                                documentId=doc_id,
                                body={"requests": [{
                                    "mergeTableCells": {
                                        "tableRange": {
                                            "tableCellLocation": {
                                                "tableStartLocation": {"index": tabla_start_real},
                                                "rowIndex":    m["min_row"],
                                                "columnIndex": m["min_col"]
                                            },
                                            "rowSpan":    m["row_span"],
                                            "columnSpan": m["col_span"]
                                        }
                                    }
                                }]}
                            ).execute()
                        except Exception:
                            pass

            # -------------------------------------------------
            # GET 3: leer estructura post-merge
            # Necesario para obtener los startIndex reales de cada celda
            # después de que los merges hayan desplazado los índices internos.
            # Este es el único GET que no se puede eliminar.
            # -------------------------------------------------
            doc_final  = docs_service.documents().get(documentId=doc_id).execute()
            contenido2 = doc_final.get("body", {}).get("content", [])

            tabla_nueva = None
            for elemento in contenido2:
                if "table" in elemento and elemento.get("startIndex") == tabla_start_real:
                    tabla_nueva = elemento["table"]
                    tabla_nueva["startIndex"] = tabla_start_real
                    tabla_nueva["endIndex"]   = elemento.get("endIndex", 0)
                    break
            if tabla_nueva is None:
                tabla_nueva = _encontrar_tabla_cerca(contenido2, tabla_start_real)
            if tabla_nueva is None:
                omitidos.append({"alias": alias, "razon": "no se pudo localizar la tabla post-merge"})
                continue

            # -------------------------------------------------
            # batch 3: todos los insertText en un solo batchUpdate
            # (de atrás hacia adelante para no desplazar índices)
            # -------------------------------------------------
            todas_las_filas = [list(df.columns)] + df.values.tolist()
            requests_texto  = []

            for fi in range(len(todas_las_filas) - 1, -1, -1):
                fila_doc   = tabla_nueva.get("tableRows", [])[fi]
                fila_datos = todas_las_filas[fi]
                celdas_doc = fila_doc.get("tableCells", [])

                pares               = []
                col_logica          = 0
                col_consumida_hasta = -1
                for celda in celdas_doc:
                    span = celda.get("tableCellStyle", {}).get("columnSpan", 1)
                    if col_logica > col_consumida_hasta:
                        pares.append((celda, col_logica))
                        col_consumida_hasta = col_logica + span - 1
                    col_logica += 1

                for celda, col_excel in reversed(pares):
                    texto = str(fila_datos[col_excel]) if col_excel < len(fila_datos) else ""
                    if not texto or texto in ("nan", "None"):
                        continue
                    parrafos  = celda.get("content", [])
                    elementos = parrafos[0].get("paragraph", {}).get("elements", []) if parrafos else []
                    if not elementos:
                        continue
                    requests_texto.append({
                        "insertText": {
                            "location": {"index": elementos[0].get("startIndex", 0)},
                            "text":     texto
                        }
                    })

            if requests_texto:
                docs_service.documents().batchUpdate(
                    documentId=doc_id,
                    body={"requests": requests_texto}
                ).execute()

            # -------------------------------------------------
            # batch 4: todo el formato en un solo batchUpdate
            # Reutilizamos tabla_nueva (GET 3) más el offset conocido que
            # produce insertText — el formato usa índices de celdas, no de
            # caracteres, así que tabla_nueva sigue siendo válida.
            # -------------------------------------------------
            if cell_formats:
                fmt_requests = _generar_requests_formato_excel(tabla_nueva, cell_formats)
                if fmt_requests:
                    docs_service.documents().batchUpdate(
                        documentId=doc_id,
                        body={"requests": fmt_requests}
                    ).execute()

            reemplazados.append({"alias": alias, "n_filas": n_filas, "n_cols": n_cols})
            _log(f"   ✅ '{alias}' → {n_filas} filas × {n_cols} cols  "
                 f"({len(merges)} merges, {len(requests_texto)} celdas, "
                 f"{len(cell_formats)} formatos)")

        except Exception as e:
            import traceback
            traceback.print_exc()
            omitidos.append({"alias": alias, "razon": f"error: {str(e)}"})

    _log(f"✅ Tablas reemplazadas: {len(reemplazados)}")
    if omitidos:
        _log(f"⚠️  Tablas omitidas: {len(omitidos)}")
        for o in omitidos:
            _log(f"   {o['alias']}: {o['razon']}")

    return {"reemplazados": reemplazados, "omitidos": omitidos}


# ==============================
# HELPERS DE TABLAS
# ==============================

def _cargar_datos_tabla(url, sheet_name, drive_service, header_row=None, alias=None, placeholder=None):
    """
    Descarga datos, formato y merges desde un .xlsx en Drive o Google Sheets nativo.
    Devuelve (df, cell_formats, merges):
        df           : DataFrame con los datos (sin columnas fantasma)
        cell_formats : { (fila_doc, col_excel): {bold, bg_color, font_color, font_size} }
        merges       : [ {min_row, min_col, max_row, max_col} ] en coordenadas 0-based
                       relativas al header_idx (fila_doc 0 = primera fila de la tabla)

    IMPORTANTE: si el archivo Excel tiene varias hojas y no se proporciona
    ``sheet_name`` (o el nombre dado no coincide con ninguna hoja), la función
    lanza ValueError en vez de cargar silenciosamente la hoja activa. Esto
    evita el bug de que todas las tablas sin sheet acabaran reemplazadas con
    los datos de la misma hoja. Si se quiere hacer fallback por nombre del
    placeholder, se puede pasar el ``placeholder`` o el ``alias``.
    """
    import io
    import pandas as pd
    import openpyxl
    from googleapiclient.http import MediaIoBaseDownload
    from googleapiclient.errors import HttpError

    file_id = extraer_id_gdoc(url)

    def _descargar(request):
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        fh.seek(0)
        return fh

    try:
        fh = _descargar(drive_service.files().get_media(fileId=file_id))
    except HttpError as e:
        if e.resp.status in (400, 403):
            fh = _descargar(drive_service.files().export_media(
                fileId=file_id,
                mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ))
        else:
            raise

    fh.seek(0)
    wb = openpyxl.load_workbook(fh, data_only=True)

    # ── Selección estricta de la hoja ───────────────────────────────────────
    # Antes: si sheet_name era None o no existía, se usaba wb.active.
    # Eso causaba que TODAS las tablas que no especificaban hoja terminaran
    # con datos de la primera hoja del archivo (p.ej. EOLOVANOS).
    # Ahora: si hay más de una hoja y no se especifica una válida, error.
    n_hojas = len(wb.sheetnames)

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        hoja_usada = sheet_name
    elif n_hojas == 1:
        # Archivo con una sola hoja: no hay ambigüedad, se usa esa.
        ws = wb.active
        hoja_usada = ws.title
    else:
        # Intentar fallback por nombre del placeholder/alias
        candidato = _resolver_hoja_por_nombre(wb.sheetnames, placeholder, alias)
        if candidato:
            ws = wb[candidato]
            hoja_usada = candidato
            _log(
                f"⚠️  No se especificó 'sheet' para '{placeholder or alias}'. "
                f"Se usará la hoja '{candidato}' detectada por similitud de nombre."
            )
        else:
            hojas_disponibles = ", ".join(f"'{h}'" for h in wb.sheetnames)
            etiqueta = placeholder or alias or "(sin etiqueta)"
            raise ValueError(
                f"El archivo Excel para la tabla {etiqueta} tiene {n_hojas} hojas "
                f"pero no se especificó cuál usar. "
                f"sheet_name recibido: {sheet_name!r}. "
                f"Hojas disponibles: {hojas_disponibles}. "
                f"Agrega una columna 'sheet' en el diccionario con el nombre exacto de la hoja."
            )

    _log(f"   📄 Hoja seleccionada: '{hoja_usada}' ({ws.max_row} filas × {ws.max_column} cols)")

    # ── Detectar última columna, última fila y primera fila con datos reales ──
    max_col_real        = 0
    max_row_real        = 0
    first_row_con_datos = None
    for fila in ws.iter_rows():
        for celda in fila:
            val = celda.value
            if val is not None and str(val).strip() not in ("", "'"):
                max_col_real = max(max_col_real, celda.column)
                max_row_real = max(max_row_real, celda.row)
                if first_row_con_datos is None:
                    first_row_con_datos = celda.row
    if max_col_real == 0:
        max_col_real = ws.max_column
    if max_row_real == 0:
        max_row_real = ws.max_row

    # Si header_row viene definido úsalo; si no, auto-detectar primera fila con datos
    if header_row and header_row > 0:
        header_idx = header_row - 1
        # Validar que esa fila tenga contenido real; si no, avanzar a la siguiente con datos
        fila_candidata = list(ws.iter_rows(min_row=header_idx + 1, max_row=header_idx + 1, max_col=max_col_real, values_only=False))
        if fila_candidata:
            tiene_contenido = any(
                c.value is not None and str(c.value).strip() not in ("", "'")
                for c in fila_candidata[0]
            )
            if not tiene_contenido and first_row_con_datos is not None:
                header_idx = first_row_con_datos - 1
    else:
        header_idx = (first_row_con_datos - 1) if first_row_con_datos else 0

    # ── Extraer datos (solo hasta max_col_real y max_row_real) ───────────────
    filas_excel = list(ws.iter_rows(
        min_row=header_idx + 1,
        max_row=max_row_real,
        max_col=max_col_real,
        values_only=False,
    ))

    def _limpiar_valor_encabezado(val):
        """Elimina el apóstrofe inicial que Excel añade a celdas de texto puro."""
        if val is None:
            return ""
        s = str(val)
        return s.lstrip("'") if s.startswith("'") else s

    encabezados = [_limpiar_valor_encabezado(c.value) for c in filas_excel[0]]
    datos_filas = []
    for fila in filas_excel[1:]:
        datos_filas.append([str(c.value) if c.value is not None else "" for c in fila])

    df = pd.DataFrame(datos_filas, columns=encabezados)
    df = df.fillna("").astype(str)

    # ── Extraer formato (solo hasta max_col_real) ─────────────────────────────
    cell_formats = {}
    for doc_ri, fila in enumerate(filas_excel):
        for ci, celda in enumerate(fila):
            if ci >= max_col_real:
                break
            fmt = {}

            # Color de fondo
            fill = celda.fill
            if fill and fill.fill_type not in (None, "none"):
                color = fill.fgColor
                if color and color.type == "rgb" and color.rgb not in ("00000000", "FFFFFFFF", "FF000000"):
                    fmt["bg_color"] = _hex_argb_a_rgb_float(color.rgb)

            # Negrita
            if celda.font and celda.font.bold:
                fmt["bold"] = True

            # Tamaño de fuente (solo si está definido explícitamente)
            if celda.font and celda.font.size:
                fmt["font_size"] = float(celda.font.size)

            # Color de fuente (solo si no es negro por defecto)
            if celda.font and celda.font.color:
                fc = celda.font.color
                if fc.type == "rgb" and fc.rgb not in ("00000000", "FF000000"):
                    fmt["font_color"] = _hex_argb_a_rgb_float(fc.rgb)

            if fmt:
                cell_formats[(doc_ri, ci)] = fmt

    # ── Extraer merges (ajustados a header_idx y recortados a max_col_real) ───
    merges = []
    for rng in ws.merged_cells.ranges:
        # Convertir a 0-based y relativos a header_idx
        doc_min_row = rng.min_row - 1 - header_idx
        doc_max_row = rng.max_row - 1 - header_idx
        min_col     = rng.min_col - 1
        max_col     = min(rng.max_col - 1, max_col_real - 1)
        # Solo incluir merges dentro del rango visible
        if doc_min_row >= 0 and min_col < max_col_real:
            merges.append({
                "min_row":  doc_min_row,
                "max_row":  doc_max_row,
                "min_col":  min_col,
                "max_col":  max_col,
                "row_span": doc_max_row - doc_min_row + 1,
                "col_span": max_col - min_col + 1,
            })

    return df, cell_formats, merges


def _hex_argb_a_rgb_float(hex_argb):
    """Convierte 'FF00FF00' → {'red': 0.0, 'green': 1.0, 'blue': 0.0}"""
    hex_rgb = hex_argb[-6:]
    r = int(hex_rgb[0:2], 16) / 255
    g = int(hex_rgb[2:4], 16) / 255
    b = int(hex_rgb[4:6], 16) / 255
    return {"red": round(r, 4), "green": round(g, 4), "blue": round(b, 4)}


def _resolver_hoja_por_nombre(hojas_disponibles, placeholder, alias):
    """
    Intenta detectar a qué hoja del Excel corresponde un placeholder, comparando
    palabras significativas. Devuelve el nombre exacto de la hoja si la
    coincidencia es lo bastante fuerte, o None en caso contrario.

    Esto es solo un fallback de conveniencia. La forma correcta es declarar la
    hoja explícitamente en una columna 'sheet' del diccionario.
    """
    import unicodedata, re as _re

    def _normalizar(s):
        if not s:
            return ""
        s = str(s)
        # quitar acentos
        s = "".join(
            c for c in unicodedata.normalize("NFD", s)
            if unicodedata.category(c) != "Mn"
        )
        # solo letras/números, en minúscula
        s = _re.sub(r"[^a-zA-Z0-9]+", " ", s).lower().strip()
        return s

    STOPWORDS = {
        "de", "del", "la", "el", "los", "las", "y", "en", "con", "sobre",
        "por", "para", "a", "al", "un", "una", "mt", "ac", "dc",
    }

    def _tokens(s):
        return {t for t in _normalizar(s).split() if t and t not in STOPWORDS}

    objetivo = _tokens(placeholder) | _tokens(alias)
    if not objetivo:
        return None

    mejor_hoja = None
    mejor_score = 0
    for hoja in hojas_disponibles:
        toks_hoja = _tokens(hoja)
        if not toks_hoja:
            continue
        comunes = objetivo & toks_hoja
        if not comunes:
            continue
        # Cobertura sobre la hoja: qué tan bien cubre el placeholder los
        # tokens de la hoja (más estricto que solo contar palabras comunes).
        score = len(comunes) / max(len(toks_hoja), 1)
        if score > mejor_score:
            mejor_score = score
            mejor_hoja = hoja

    # Umbral conservador: al menos 50% de los tokens de la hoja deben aparecer
    # en el placeholder. Esto evita que "Tablas de regulación MT" se empareje
    # con "VANOS IDEALES DE REGULACIÓN" solo por compartir "regulacion".
    if mejor_score >= 0.5:
        return mejor_hoja
    return None


def _generar_requests_formato_excel(tabla_nueva, cell_formats):
    """
    Genera requests de updateTableCellStyle y updateTextStyle.
    Mapea (fila_doc, col_doc_ajustado) → celda del Doc, teniendo en cuenta
    que las filas con merges tienen menos celdas físicas.
    """
    requests  = []
    tabla_id  = tabla_nueva.get("startIndex", 0)
    filas_doc = tabla_nueva.get("tableRows", [])

    for (fi, ci_excel), fmt in cell_formats.items():
        if fi >= len(filas_doc):
            continue

        # Resolver el índice de celda física en el Doc teniendo en cuenta merges
        # Cada celda puede tener colSpan > 1, lo que desplaza el ci real
        celdas_doc = filas_doc[fi].get("tableCells", [])
        celda_doc  = None
        ci_doc     = None
        col_offset = 0
        ultima_celda_principal_end = -1
        for idx, c in enumerate(celdas_doc):
            span = c.get("tableCellStyle", {}).get("columnSpan", 1)
            paras = c.get("content", [])
            celda_start = paras[0].get("paragraph", {}).get("elements", [{}])[0].get("startIndex", -1) if paras else -1
            es_absorbida = (celda_start != -1 and celda_start < ultima_celda_principal_end)
            if not es_absorbida:
                ultima_celda_principal_end = c.get("endIndex", celda_start + 1)
            if col_offset == ci_excel:
                if not es_absorbida:
                    celda_doc = c
                    ci_doc    = idx
                break
            col_offset += span

        if celda_doc is None:
            continue

        # ── Color de fondo ───────────────────────────────────────────────────
        if "bg_color" in fmt:
            col_span = celda_doc.get("tableCellStyle", {}).get("columnSpan", 1)
            requests.append({
                "updateTableCellStyle": {
                    "tableRange": {
                        "tableCellLocation": {
                            "tableStartLocation": {"index": tabla_id},
                            "rowIndex":    fi,
                            "columnIndex": ci_excel
                        },
                        "rowSpan":    1,
                        "columnSpan": col_span
                    },
                    "tableCellStyle": {
                        "backgroundColor": {
                            "color": {"rgbColor": fmt["bg_color"]}
                        }
                    },
                    "fields": "backgroundColor"
                }
            })

        # ── Texto: negrita, tamaño, color ────────────────────────────────────
        text_style  = {}
        text_fields = []

        if fmt.get("bold"):
            text_style["bold"] = True
            text_fields.append("bold")

        if "font_size" in fmt:
            text_style["fontSize"] = {"magnitude": fmt["font_size"], "unit": "PT"}
            text_fields.append("fontSize")

        if "font_color" in fmt:
            text_style["foregroundColor"] = {
                "color": {"rgbColor": fmt["font_color"]}
            }
            text_fields.append("foregroundColor")

        if text_style:
            parrafos  = celda_doc.get("content", [])
            elementos = parrafos[0].get("paragraph", {}).get("elements", []) if parrafos else []
            if elementos:
                ts_start = elementos[0].get("startIndex", 0)
                ts_end   = celda_doc.get("endIndex", 0) - 1
                if ts_start < ts_end:
                    requests.append({
                        "updateTextStyle": {
                            "range": {
                                "startIndex": ts_start,
                                "endIndex":   ts_end
                            },
                            "textStyle": text_style,
                            "fields":    ",".join(text_fields)
                        }
                    })

    return requests


def _encontrar_tabla_cerca(contenido, indice_referencia):
    """
    Encuentra la tabla cuyo startIndex esté más cercano (y no antes) al índice dado.
    """
    mejor     = None
    menor_dist = float("inf")
    for elemento in contenido:
        if "table" not in elemento:
            continue
        start = elemento.get("startIndex", 0)
        dist  = abs(start - indice_referencia)
        if dist < menor_dist:
            menor_dist = dist
            tabla      = elemento["table"]
            tabla["startIndex"] = start
            tabla["endIndex"]   = elemento.get("endIndex", 0)
            mejor = tabla
    return mejor


# ==============================
# LOOPS (tablas dinámicas por hojas)
# ==============================

def _listar_hojas_con_prefijo(url, prefijo, drive_service):
    """
    Descarga el Excel desde Drive y devuelve los nombres de las hojas
    cuyo nombre empiece con ``prefijo`` (case-insensitive), en el orden
    en que aparecen en el libro.

    Args:
        url          (str): URL o file_id del Excel en Drive.
        prefijo      (str): Prefijo que deben tener los nombres de hoja.
                            Si es None o vacío, devuelve TODAS las hojas.
        drive_service:      Servicio autenticado de Google Drive API.

    Returns:
        list[str]: Nombres de hojas que cumplen el filtro.
    """
    import io
    import openpyxl
    from googleapiclient.http import MediaIoBaseDownload
    from googleapiclient.errors import HttpError

    file_id = extraer_id_gdoc(url)

    def _descargar(request):
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        fh.seek(0)
        return fh

    try:
        fh = _descargar(drive_service.files().get_media(fileId=file_id))
    except HttpError as e:
        if e.resp.status in (400, 403):
            fh = _descargar(drive_service.files().export_media(
                fileId=file_id,
                mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ))
        else:
            raise

    wb = openpyxl.load_workbook(fh, read_only=True, data_only=True)
    hojas = wb.sheetnames

    if not prefijo:
        return hojas

    prefijo_lower = prefijo.lower()
    return [h for h in hojas if h.lower().startswith(prefijo_lower)]


def reemplazar_loops(doc_id, diccionario, docs_service, drive_service):
    """
    Reemplaza placeholders de tipo 'loop' en un Google Doc.

    Un placeholder de tipo loop tiene la forma ``{% loop alias %}`` en el
    documento. El código:

        1. Lista todas las hojas del Excel cuyo nombre empiece con el prefijo
           definido en la columna ``sheet`` del diccionario.
        2. Localiza el placeholder ``{% loop alias %}`` en el documento.
        3. Borra el placeholder.
        4. Por cada hoja (en orden), inserta un salto de página seguido de la
           tabla con los datos y formato de esa hoja.
           - Las tablas se insertan de ATRÁS HACIA ADELANTE para no desplazar
             los índices de las inserciones siguientes.

    El formato de cada tabla (negrita, colores, merges) se preserva igual que
    en ``reemplazar_tablas()``.

    Sintaxis en el diccionario Excel:
        Nombre del dato : (cualquier etiqueta descriptiva)
        Alias           : tabla_tendido        ← nombre del loop
        Tipo            : Loop
        Valor           : https://drive.google.com/...  ← URL del Excel con hojas por cantón
        sheet           : Canton_              ← prefijo de las hojas a incluir
        header_row      : 1                    ← (opcional) fila de encabezados

    Sintaxis en Google Docs:
        {% loop tabla_tendido %}

    Args:
        doc_id        (str):  ID del documento de Google Docs.
        diccionario   (dict): Diccionario unificado generado por cargar_diccionario().
        docs_service:         Servicio autenticado de Google Docs API.
        drive_service:        Servicio autenticado de Google Drive API.

    Returns:
        dict: {"reemplazados": [...], "omitidos": [...]}
    """
    reemplazados = []
    omitidos     = []

    entradas_loop = [
        (alias, entrada)
        for alias, entrada in diccionario.items()
        if entrada.get("type") == "loop"
        and entrada.get("value") is not None
    ]

    if not entradas_loop:
        _log("ℹ️  No se encontraron entradas de tipo 'loop' para reemplazar.")
        return {"reemplazados": [], "omitidos": omitidos}

    for alias, entrada in entradas_loop:
        valor      = entrada["value"]
        url_excel  = valor.get("file_id")
        prefijo    = valor.get("prefijo")
        header_row = valor.get("header_row")

        # El placeholder en el doc es {% loop alias %}
        placeholder = f"{{% loop {alias} %}}"

        if not url_excel:
            omitidos.append({"alias": alias, "razon": "URL de Excel vacía"})
            continue

        try:
            # -------------------------------------------------
            # 1. Listar hojas que cumplan el prefijo
            # -------------------------------------------------
            hojas = _listar_hojas_con_prefijo(url_excel, prefijo, drive_service)

            if not hojas:
                omitidos.append({
                    "alias": alias,
                    "razon": f"No se encontraron hojas con prefijo '{prefijo}' en el Excel"
                })
                continue

            _log(f"   🗂️  Loop '{alias}': {len(hojas)} hojas encontradas → {hojas}")

            # -------------------------------------------------
            # 2. Localizar el placeholder en el documento
            # -------------------------------------------------
            documento   = docs_service.documents().get(documentId=doc_id).execute()
            contenido   = documento.get("body", {}).get("content", [])
            texto_plano, mapa_indices = _extraer_texto_con_indices(contenido)

            pos = texto_plano.find(placeholder)
            if pos == -1:
                omitidos.append({
                    "alias": alias,
                    "razon": f"placeholder '{placeholder}' no encontrado en el documento"
                })
                continue

            start_idx = _idx(pos, mapa_indices)
            end_idx   = _idx(pos + len(placeholder), mapa_indices)

            if start_idx is None or end_idx is None:
                omitidos.append({"alias": alias, "razon": "no se pudo mapear el índice del placeholder"})
                continue

            # -------------------------------------------------
            # 3. Borrar el placeholder
            # -------------------------------------------------
            docs_service.documents().batchUpdate(
                documentId=doc_id,
                body={"requests": [{
                    "deleteContentRange": {
                        "range": {"startIndex": start_idx, "endIndex": end_idx}
                    }
                }]}
            ).execute()

            # -------------------------------------------------
            # 4. Insertar tablas de ATRÁS HACIA ADELANTE
            #    Cada tabla va precedida de un salto de página,
            #    excepto la primera (que queda en el lugar del placeholder).
            #    Insertamos en orden inverso para no desplazar start_idx.
            # -------------------------------------------------
            tablas_insertadas = 0

            for hoja in reversed(hojas):
                try:
                    df, cell_formats, merges = _cargar_datos_tabla(
                        url_excel, hoja, drive_service, header_row,
                        alias=alias, placeholder=hoja,
                    )
                    if df is None or df.empty:
                        _log(f"   ⚠️  Hoja '{hoja}' vacía, se omite.")
                        continue

                    n_filas = len(df) + 1   # +1 fila de encabezados
                    n_cols  = len(df.columns)

                    # ── 4a. Insertar salto de página + tabla vacía ──────────
                    # El salto de página va ANTES de la tabla (excepto la primera).
                    # Como insertamos en reversa, el salto va después en el índice
                    # pero termina antes en el documento final.
                    requests_insercion = []

                    if tablas_insertadas > 0:
                        # Salto de página al inicio de la posición actual
                        requests_insercion.append({
                            "insertPageBreak": {
                                "location": {"index": start_idx}
                            }
                        })

                    requests_insercion.append({
                        "insertTable": {
                            "rows":     n_filas,
                            "columns":  n_cols,
                            "location": {"index": start_idx}
                        }
                    })

                    docs_service.documents().batchUpdate(
                        documentId=doc_id,
                        body={"requests": requests_insercion}
                    ).execute()

                    # ── 4b. Localizar la tabla recién insertada ─────────────
                    doc_tmp       = docs_service.documents().get(documentId=doc_id).execute()
                    contenido_tmp = doc_tmp.get("body", {}).get("content", [])
                    tabla_tmp     = None
                    for elem in contenido_tmp:
                        if "table" in elem and elem.get("startIndex") == start_idx:
                            t = elem["table"]
                            t["startIndex"] = elem["startIndex"]
                            t["endIndex"]   = elem.get("endIndex", 0)
                            tabla_tmp = t
                            break
                    if tabla_tmp is None:
                        tabla_tmp = _encontrar_tabla_cerca(contenido_tmp, start_idx)
                    if not tabla_tmp:
                        _log(f"   ⚠️  No se pudo localizar la tabla insertada para '{hoja}'.")
                        continue

                    tabla_start_real = tabla_tmp["startIndex"]

                    # ── 4c. batch 2: todos los merges en UN solo batchUpdate ──
                    if merges:
                        merges_ordenados = sorted(merges, key=lambda m: (-m["min_row"], -m["min_col"]))
                        merge_requests   = [
                            {
                                "mergeTableCells": {
                                    "tableRange": {
                                        "tableCellLocation": {
                                            "tableStartLocation": {"index": tabla_start_real},
                                            "rowIndex":    m["min_row"],
                                            "columnIndex": m["min_col"]
                                        },
                                        "rowSpan":    m["row_span"],
                                        "columnSpan": m["col_span"]
                                    }
                                }
                            }
                            for m in merges_ordenados
                        ]
                        try:
                            docs_service.documents().batchUpdate(
                                documentId=doc_id,
                                body={"requests": merge_requests}
                            ).execute()
                        except Exception as e_merge:
                            _log(f"   ⚠️  Batch de merges falló ({e_merge}), reintentando secuencialmente…")
                            for m in merges_ordenados:
                                try:
                                    docs_service.documents().batchUpdate(
                                        documentId=doc_id,
                                        body={"requests": [{
                                            "mergeTableCells": {
                                                "tableRange": {
                                                    "tableCellLocation": {
                                                        "tableStartLocation": {"index": tabla_start_real},
                                                        "rowIndex":    m["min_row"],
                                                        "columnIndex": m["min_col"]
                                                    },
                                                    "rowSpan":    m["row_span"],
                                                    "columnSpan": m["col_span"]
                                                }
                                            }
                                        }]}
                                    ).execute()
                                except Exception:
                                    pass

                    # ── 4d. GET 3: leer estructura post-merge ───────────────
                    doc_final  = docs_service.documents().get(documentId=doc_id).execute()
                    contenido2 = doc_final.get("body", {}).get("content", [])
                    tabla_nueva = None
                    for elemento in contenido2:
                        if "table" in elemento and elemento.get("startIndex") == tabla_start_real:
                            tabla_nueva = elemento["table"]
                            tabla_nueva["startIndex"] = tabla_start_real
                            tabla_nueva["endIndex"]   = elemento.get("endIndex", 0)
                            break
                    if tabla_nueva is None:
                        tabla_nueva = _encontrar_tabla_cerca(contenido2, tabla_start_real)
                    if tabla_nueva is None:
                        _log(f"   ⚠️  No se pudo releer la tabla para '{hoja}'.")
                        continue

                    # ── 4e. batch 3: todos los insertText en un solo batch ──
                    todas_las_filas = [list(df.columns)] + df.values.tolist()
                    requests_texto  = []

                    for fi in range(len(todas_las_filas) - 1, -1, -1):
                        fila_doc   = tabla_nueva.get("tableRows", [])[fi]
                        fila_datos = todas_las_filas[fi]
                        celdas_doc = fila_doc.get("tableCells", [])

                        pares = []
                        col_logica           = 0
                        col_consumida_hasta  = -1
                        for celda in celdas_doc:
                            span = celda.get("tableCellStyle", {}).get("columnSpan", 1)
                            if col_logica > col_consumida_hasta:
                                pares.append((celda, col_logica))
                                col_consumida_hasta = col_logica + span - 1
                            col_logica += 1

                        for celda, col_excel in reversed(pares):
                            texto = str(fila_datos[col_excel]) if col_excel < len(fila_datos) else ""
                            if not texto or texto in ("nan", "None"):
                                continue
                            parrafos  = celda.get("content", [])
                            elementos = parrafos[0].get("paragraph", {}).get("elements", []) if parrafos else []
                            if not elementos:
                                continue
                            idx_insercion = elementos[0].get("startIndex", 0)
                            requests_texto.append({
                                "insertText": {
                                    "location": {"index": idx_insercion},
                                    "text":     texto
                                }
                            })

                    if requests_texto:
                        docs_service.documents().batchUpdate(
                            documentId=doc_id,
                            body={"requests": requests_texto}
                        ).execute()

                    # ── 4f. batch 4: todo el formato en un solo batch ───────
                    if cell_formats:
                        fmt_requests = _generar_requests_formato_excel(tabla_nueva, cell_formats)
                        if fmt_requests:
                            docs_service.documents().batchUpdate(
                                documentId=doc_id,
                                body={"requests": fmt_requests}
                            ).execute()

                    tablas_insertadas += 1
                    _log(f"   ✅ Tabla insertada para hoja '{hoja}' ({n_filas} filas × {n_cols} cols)")

                except Exception as e_hoja:
                    import traceback
                    traceback.print_exc()
                    _log(f"   ❌ Error procesando hoja '{hoja}': {e_hoja}")
                    omitidos.append({"alias": alias, "razon": f"error en hoja '{hoja}': {str(e_hoja)}"})

            reemplazados.append({
                "alias":             alias,
                "hojas":             hojas,
                "tablas_insertadas": tablas_insertadas,
            })

        except Exception as e:
            import traceback
            traceback.print_exc()
            omitidos.append({"alias": alias, "razon": f"error: {str(e)}"})

    _log(f"✅ Loops reemplazados: {len(reemplazados)}")
    for r in reemplazados:
        _log(f"   {r['alias']} → {r['tablas_insertadas']} tablas ({', '.join(r['hojas'])})")

    if omitidos:
        _log(f"⚠️  Loops omitidos: {len(omitidos)}")
        for o in omitidos:
            _log(f"   {o['alias']}: {o['razon']}")

    return {"reemplazados": reemplazados, "omitidos": omitidos}


# ==============================
# GOOGLE AUTH
# ==============================
import os
from google.oauth2 import service_account, credentials as oauth2_credentials
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

SCOPES = [
    'https://www.googleapis.com/auth/documents',
    'https://www.googleapis.com/auth/drive',
    'https://www.googleapis.com/auth/spreadsheets'
]


def autenticar_oauth(ruta_client_secret=None, ruta_token=None):
    """
    Autenticación OAuth2 con cuenta personal de Google.

    Las rutas se pueden pasar como argumento o desde variables de entorno:
        GOOGLE_OAUTH_CLIENT → ruta al client_secret.json
        GOOGLE_OAUTH_TOKEN  → ruta donde guardar/leer el token (opcional)

    El token se renueva automáticamente si expiró.
    Si no existe, abre el navegador para autorizar por primera vez.
    """
    if ruta_client_secret is None:
        ruta_client_secret = os.getenv("GOOGLE_OAUTH_CLIENT")
        if not ruta_client_secret:
            raise Exception("Define la variable de entorno GOOGLE_OAUTH_CLIENT con la ruta al client_secret.json")

    if not os.path.exists(ruta_client_secret):
        raise FileNotFoundError(f"No existe el archivo: {ruta_client_secret}")

    if ruta_token is None:
        ruta_token = os.getenv(
            "GOOGLE_OAUTH_TOKEN",
            os.path.join(os.path.dirname(ruta_client_secret), "token.json")
        )

    creds = None

    if os.path.exists(ruta_token):
        creds = oauth2_credentials.Credentials.from_authorized_user_file(ruta_token, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(ruta_client_secret, SCOPES)
            creds = flow.run_local_server(port=0)

        with open(ruta_token, "w") as f:
            f.write(creds.to_json())

    docs_service   = build('docs',   'v1', credentials=creds)
    drive_service  = build('drive',  'v3', credentials=creds)
    sheets_service = build('sheets', 'v4', credentials=creds)

    return docs_service, drive_service, sheets_service


def autenticar_servicio(ruta_credenciales):
    """
    Autenticación con service account (para cuando se configure la Unidad Compartida).
    """
    creds = service_account.Credentials.from_service_account_file(
        ruta_credenciales,
        scopes=SCOPES
    )
    docs_service   = build('docs',   'v1', credentials=creds)
    drive_service  = build('drive',  'v3', credentials=creds)
    sheets_service = build('sheets', 'v4', credentials=creds)

    return docs_service, drive_service, sheets_service


# ==============================
# OPERACIONES DE DRIVE
# ==============================

def descargar_excel_drive(file_id, drive_service):
    """
    Descarga un archivo de Google Drive por file_id y lo retorna como BytesIO.
    """
    request = drive_service.files().get_media(fileId=file_id)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)

    done = False
    while not done:
        status, done = downloader.next_chunk()

    fh.seek(0)
    return fh


def extraer_id_gdoc(url_o_id):
    """
    Acepta una URL de Google Docs/Drive o directamente un ID y devuelve el file ID.

    Ejemplos aceptados:
        https://docs.google.com/document/d/FILE_ID/edit
        https://drive.google.com/drive/folders/FILE_ID
        FILE_ID (directamente)
    """
    match = re.search(r"/(?:d|folders)/([a-zA-Z0-9_-]+)", url_o_id)
    if match:
        return match.group(1)
    return url_o_id.strip()


def copiar_documento(doc_id, nombre_nuevo, drive_service, carpeta_destino_id=None):
    """
    Crea una copia del documento en Drive y devuelve el ID del nuevo archivo.

    Args:
        doc_id             (str): ID del documento original.
        nombre_nuevo       (str): Nombre para la copia.
        drive_service:           Servicio autenticado de Google Drive API.
        carpeta_destino_id (str): ID de la carpeta destino. Si se omite,
                                  la copia queda en la raíz del Drive del usuario.
    Returns:
        str: ID del documento copiado.
    """
    body = {"name": nombre_nuevo}
    if carpeta_destino_id:
        body["parents"] = [carpeta_destino_id]

    copia = drive_service.files().copy(
        fileId=doc_id,
        body=body,
        supportsAllDrives=True
    ).execute()

    return copia["id"]


def listar_archivos_drive(drive_service, page_size=10):
    """Lista archivos visibles por la cuenta autenticada."""
    results = drive_service.files().list(
        pageSize=page_size,
        fields="files(id, name)"
    ).execute()
    return results.get('files', [])


# ==============================
# UTILIDADES
# ==============================

def obtener_ruta_credenciales():
    """Obtiene la ruta del JSON de service account desde variable de entorno GOOGLE_CREDS."""
    ruta = os.getenv("GOOGLE_CREDS")
    if not ruta:
        raise Exception("Debes definir la variable de entorno GOOGLE_CREDS")
    if not os.path.exists(ruta):
        raise FileNotFoundError(f"No existe el archivo: {ruta}")
    return ruta
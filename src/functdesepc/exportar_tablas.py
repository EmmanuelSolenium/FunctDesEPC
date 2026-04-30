"""
exportar_tablas.py
==================
Funciones de exportación a Excel con formato y previsualización en Colab
para las TABLAS DE FLECHADO del proyecto FunctDesEPC.

Uso típico (en crear_tablas_mec.ipynb):
----------------------------------------
    from exportar_tablas import exportar_tablas_flechado, previsualizar_tabla_flechado

    # Exportar todas las tablas a Excel (n hojas normales + m hojas secundarias)
    exportar_tablas_flechado(
        tab_fle=tab_fle,          # lista de DataFrames con header de cada cantón normal
        tablas_p=tablas_p,        # lista de DataFrames con datos de flechado cantón normal
        tab_fle_s=tab_fle_s,      # lista de DataFrames con header de cada cantón secundario
        tablas_s=tablas_s,        # lista de DataFrames con datos de flechado cantón secundario
        filepath="Tablas_Flechado.xlsx"
    )

    # Previsualizar una tabla individual en Colab
    previsualizar_tabla_flechado(tab_fle[0], tablas_p[0], titulo="Cantón 1")
    previsualizar_tabla_flechado(tab_fle_s[0], tablas_s[0], titulo="Cantón 1S")
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, numbers
)
from openpyxl.utils import get_column_letter



# ---------------------------------------------------------------------------
# Estilos reutilizables
# ---------------------------------------------------------------------------

def _thin():
    return Side(style="thin", color="FF000000")

def _medium():
    return Side(style="medium", color="FF000000")

def _border_thin_all():
    t = _thin()
    return Border(left=t, right=t, top=t, bottom=t)

def _border_medium_all():
    m = _medium()
    return Border(left=m, right=m, top=m, bottom=m)

def _border_medium_outer_thin_inner(left=True, right=True, top=True, bottom=True):
    """Borde medium en los lados indicados, thin en el resto."""
    def s(is_medium):
        return _medium() if is_medium else _thin()
    return Border(
        left=s(left), right=s(right), top=s(top), bottom=s(bottom)
    )

FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")   # azul claro para etiquetas
FILL_DATA   = PatternFill("solid", fgColor="FFFFFF")    # blanco para datos
FONT_BOLD   = Font(bold=True, size=9)
FONT_NORMAL = Font(bold=False, size=9)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
NUM_FMT_2DEC = '0.00'
NUM_FMT_4DEC = '0.0000'


def _round4(val):
    """Redondea a 4 decimales los valores float con más de 4 cifras decimales."""
    if isinstance(val, float):
        return round(val, 4)
    return val


# ---------------------------------------------------------------------------
# Helpers internos
# ---------------------------------------------------------------------------

def _write_cell(ws, row, col, value, font=None, alignment=None,
                border=None, fill=None, number_format=None):
    """Escribe un valor en una celda con formato opcional."""
    cell = ws.cell(row=row, column=col, value=_round4(value))
    if font is not None:
        cell.font = font
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
    if fill is not None:
        cell.fill = fill
    if number_format is not None:
        cell.number_format = number_format
    return cell


def _merge_and_write(ws, row, col_start, col_end, value,
                     font=None, alignment=None, border=None, fill=None):
    """Fusiona celdas y escribe el valor en la primera."""
    if col_start < col_end:
        ws.merge_cells(
            start_row=row, start_column=col_start,
            end_row=row,   end_column=col_end
        )
    _write_cell(ws, row, col_start, value, font=font,
                alignment=alignment, border=border, fill=fill)


def _auto_col_width(ws, min_width=8, max_width=30):
    """Ajusta el ancho de columnas automáticamente según contenido."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


# ---------------------------------------------------------------------------
# Escritura del bloque HEADER de una tabla de flechado
# ---------------------------------------------------------------------------

def _escribir_header_tabla(ws, tab_fle_df, start_row=1, start_col=1):
    """
    Escribe la parte superior de la tabla (header) a partir de tab_fle_df.

    tab_fle_df se asume un DataFrame cuyas filas son:
        0: Conductor
        1: Cantón, Poste Inicial
        2: Vano de Regulación, Poste Final
        3: Vano (números de vano)
        4: Longitud (m)
        5: Poste inicial
        6: Poste final
        7: Desnivel

    La primera columna contiene las etiquetas (ej. "Conductor", "Cantón", etc.)
    y las demás columnas contienen los valores.
    Las celdas con valor NaN se dejan completamente sin formato ni contenido.

    Retorna la siguiente fila disponible (start_row + número de filas escritas).
    """
    b_med = _border_medium_all()
    b_thin = _border_thin_all()
    label_fill = FILL_HEADER

    n_rows, n_cols = tab_fle_df.shape
    label_col = start_col
    data_col_start = start_col + 1

    for rel_row, (idx, row_data) in enumerate(tab_fle_df.iterrows()):
        abs_row = start_row + rel_row
        row_vals = row_data.tolist()

        label = row_vals[0] if len(row_vals) > 0 else ""
        data_vals = row_vals[1:]

        is_vano_row = str(label).strip().lower() in {
            "vano", "longitud (m)", "poste inicial", "poste final", "desnivel"
        }
        border = b_med if is_vano_row else b_thin

        # Etiqueta — solo escribir si no es NaN
        if pd.notna(label) and str(label).strip() != "":
            _merge_and_write(
                ws, abs_row, label_col, label_col + 1,
                value=label,
                font=FONT_BOLD,
                alignment=ALIGN_LEFT,
                border=border,
                fill=label_fill
            )

        # Valores de datos — omitir celdas NaN completamente (sin valor ni formato)
        for rel_col, val in enumerate(data_vals):
            c = data_col_start + 1 + rel_col
            try:
                es_nan = pd.isna(val)
            except (TypeError, ValueError):
                es_nan = False
            if es_nan:
                continue
            num_fmt = None
            if isinstance(val, float):
                num_fmt = NUM_FMT_4DEC if abs(val) < 10 else NUM_FMT_2DEC
            _write_cell(
                ws, abs_row, c, val,
                font=FONT_NORMAL,
                alignment=ALIGN_CENTER,
                border=border,
                fill=FILL_DATA,
                number_format=num_fmt
            )

    return start_row + n_rows


# ---------------------------------------------------------------------------
# Escritura del bloque DATOS (tablas_p / tablas_s)
# ---------------------------------------------------------------------------

def _escribir_datos_tabla(ws, tablas_df, start_row=1, start_col=1):
    """
    Escribe la parte inferior de la tabla (datos de temperatura/tense/flechas).

    tablas_df se asume un DataFrame con columnas:
        Temperatura, Tense, f_vano1, f_vano2, ...

    Escribe una fila de encabezado con los nombres de columna y luego
    los datos fila a fila. Las celdas con NaN se omiten (sin formato).
    No se escribe fila de unidades (eliminada por ser redundante).

    Retorna la siguiente fila disponible.
    """
    b_med = _border_medium_all()
    b_thin = _border_thin_all()

    # Encabezado de columnas
    cols = list(tablas_df.columns)
    for rel_col, col_name in enumerate(cols):
        c = start_col + rel_col
        _write_cell(
            ws, start_row, c, col_name,
            font=FONT_BOLD,
            alignment=ALIGN_CENTER,
            border=b_med,
            fill=FILL_HEADER
        )

    # Datos — directamente en la fila siguiente (sin fila de unidades)
    data_start_row = start_row + 1
    for rel_row, (_, row_data) in enumerate(tablas_df.iterrows()):
        abs_row = data_start_row + rel_row
        for rel_col, val in enumerate(row_data):
            c = start_col + rel_col
            try:
                es_nan = pd.isna(val)
            except (TypeError, ValueError):
                es_nan = False
            if es_nan:
                continue  # celda NaN: sin valor ni formato
            num_fmt = None
            if isinstance(val, float):
                num_fmt = NUM_FMT_4DEC
            _write_cell(
                ws, abs_row, c, val,
                font=FONT_NORMAL,
                alignment=ALIGN_CENTER,
                border=b_thin,
                fill=FILL_DATA,
                number_format=num_fmt
            )

    return data_start_row + len(tablas_df)


# ---------------------------------------------------------------------------
# Función principal de exportación
# ---------------------------------------------------------------------------

def exportar_tablas_flechado(
    tab_fle,
    tablas_p,
    tab_fle_s,
    tablas_s,
    filepath="Tablas_Flechado.xlsx"
):
    """
    Exporta las tablas de flechado a un archivo Excel con formato.

    Genera n + m hojas:
    - n hojas para cantones normales  → nombres: "Canton_1", "Canton_2", ...
    - m hojas para cantones secundarios → nombres: "Canton_1S", "Canton_2S", ...

    Cada hoja contiene:
    - Parte superior: header del cantón (tab_fle[i] o tab_fle_s[i])
    - Parte inferior: datos de temperatura/flechas (tablas_p[i] o tablas_s[i])

    Parámetros
    ----------
    tab_fle : list[pd.DataFrame]
        Lista de DataFrames con el header de cada cantón normal.
    tablas_p : list[pd.DataFrame]
        Lista de DataFrames con los datos de flechado de cada cantón normal.
        Debe tener la misma longitud que tab_fle.
    tab_fle_s : list[pd.DataFrame]
        Lista de DataFrames con el header de cada cantón secundario.
    tablas_s : list[pd.DataFrame]
        Lista de DataFrames con los datos de flechado de cada cantón secundario.
        Debe tener la misma longitud que tab_fle_s.
    filepath : str
        Ruta del archivo Excel de salida.

    Retorna
    -------
    str
        Ruta del archivo generado.
    """
    if len(tab_fle) != len(tablas_p):
        raise ValueError(
            f"tab_fle tiene {len(tab_fle)} elementos pero tablas_p tiene {len(tablas_p)}. "
            "Deben tener la misma longitud."
        )
    if len(tab_fle_s) != len(tablas_s):
        raise ValueError(
            f"tab_fle_s tiene {len(tab_fle_s)} elementos pero tablas_s tiene {len(tablas_s)}. "
            "Deben tener la misma longitud."
        )

    wb = Workbook()
    # Eliminar hoja vacía por defecto
    wb.remove(wb.active)

    # -- Cantones normales --
    for i, (header_df, datos_df) in enumerate(zip(tab_fle, tablas_p)):
        sheet_name = f"Canton_{i + 1}"
        ws = wb.create_sheet(title=sheet_name)
        next_row = _escribir_header_tabla(ws, header_df, start_row=1, start_col=1)
        # Sin fila en blanco: datos pegados directamente al header
        _escribir_datos_tabla(ws, datos_df, start_row=next_row, start_col=1)
        _auto_col_width(ws)

    # -- Cantones secundarios --
    for i, (header_df, datos_df) in enumerate(zip(tab_fle_s, tablas_s)):
        sheet_name = f"Canton_{i + 1}S"
        ws = wb.create_sheet(title=sheet_name)
        next_row = _escribir_header_tabla(ws, header_df, start_row=1, start_col=1)
        _escribir_datos_tabla(ws, datos_df, start_row=next_row, start_col=1)
        _auto_col_width(ws)

    wb.save(filepath)
    print(f"✅ Archivo guardado: {filepath}  ({len(tab_fle)} cantones + {len(tab_fle_s)} secundarios)")
    return filepath

def exportar_calculos(ruta_template, ruta_salida, mec, ret, eovanos, carac_postes, van_reg):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Border, Side, Alignment

    thick = Side(style='medium')
    wrap = Alignment(wrap_text=True)
    border = Border(left=thick, right=thick, top=thick, bottom=thick)

    def _clean(value):
        try:
            if pd.isna(value):
                return None
        except (TypeError, ValueError):
            pass
        return value

    config = [
        ("MEC",                           mec,          8),
        ("RET",                           ret,          5),
        ("EOLOVANOS",                     eovanos,      3),
        ("Caracteristicas de los postes", carac_postes, 3),
        ("VANOS IDEALES DE REGULACIÓN",   van_reg,      7),
    ]

    wb = load_workbook(ruta_template)

    for sheet_name, df, start_row in config:
        ws = wb[sheet_name]
        ws.delete_rows(start_row, ws.max_row - start_row + 1)

        # Encabezado
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=col_name)
            cell.border = border
            cell.alignment = wrap

        # Datos
        for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=_round4(_clean(value)))
                cell.border = border
                cell.alignment = wrap

    wb.save(ruta_salida)
    print(f"Archivo guardado en: {ruta_salida}")


# ── Ejemplo de uso ────────────────────────────────────────────────────────────
# exportar_calculos(
#     ruta_template = "calculos_mecanicos_raw.xlsx",
#     ruta_salida   = "calculos_mecanicos_output.xlsx",
#     mec           = mec,
#     ret           = ret,
#     eovanos       = eovanos,
#     carac_postes  = carac_postes,
#     van_reg       = van_reg,
# )

def exportar_todo(
    ruta_template,
    ruta_salida,
    mec, ret, eovanos, carac_postes, van_reg,
    tab_fle, tablas_p, tab_fle_s, tablas_s
):
    """
    Exporta en un único archivo Excel las tablas de flechado y los dataframes
    de cálculos mecánicos, preservando el formato de cada hoja.

    Parámetros
    ----------
    ruta_template : str            - Archivo base con encabezados de MEC, RET, etc.
    ruta_salida   : str            - Ruta del archivo Excel resultante
    mec, ret, eovanos, carac_postes, van_reg : pd.DataFrame
    tab_fle, tablas_p              : listas de DataFrames cantones normales
    tab_fle_s, tablas_s            : listas de DataFrames cantones secundarios
    """
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    # ── Estilos ──────────────────────────────────────────────────────────────
    def _thin():
        return Side(style="thin", color="FF000000")

    def _medium_side():
        return Side(style="medium", color="FF000000")

    def _border_thin_all():
        t = _thin()
        return Border(left=t, right=t, top=t, bottom=t)

    def _border_medium_all():
        m = _medium_side()
        return Border(left=m, right=m, top=m, bottom=m)

    FILL_HEADER  = PatternFill("solid", fgColor="D9E1F2")
    FILL_DATA    = PatternFill("solid", fgColor="FFFFFF")
    FONT_BOLD    = Font(bold=True, size=9)
    FONT_NORMAL  = Font(bold=False, size=9)
    ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    NUM_FMT_2DEC = '0.00'
    NUM_FMT_4DEC = '0.0000'

    medium   = Side(style='medium')
    border   = Border(left=medium, right=medium, top=medium, bottom=medium)
    wrap     = Alignment(wrap_text=True)

    # Colores de encabezado por hoja (del template original)
    header_fills = {
        "MEC":                           PatternFill("solid", fgColor="12501A"),
        "RET":                           PatternFill("solid", fgColor="12501A"),
        "EOLOVANOS":                     PatternFill("solid", fgColor="FE6A0C"),
        "Caracteristicas de los postes": PatternFill("solid", fgColor="C2D59A"),
        "VANOS IDEALES DE REGULACIÓN":   PatternFill("solid", fgColor="D6E3BB"),
    }
    white_font_sheets = {"MEC", "RET", "EOLOVANOS"}

    # ── Helpers flechado ─────────────────────────────────────────────────────
    def _write_cell(ws, row, col, value, font=None, alignment=None,
                    border=None, fill=None, number_format=None):
        cell = ws.cell(row=row, column=col, value=_round4(value))
        if font:          cell.font = font
        if alignment:     cell.alignment = alignment
        if border:        cell.border = border
        if fill:          cell.fill = fill
        if number_format: cell.number_format = number_format
        return cell

    def _merge_and_write(ws, row, col_start, col_end, value,
                         font=None, alignment=None, border=None, fill=None):
        if col_start < col_end:
            ws.merge_cells(start_row=row, start_column=col_start,
                           end_row=row,   end_column=col_end)
        _write_cell(ws, row, col_start, value, font=font,
                    alignment=alignment, border=border, fill=fill)

    def _auto_col_width(ws, min_width=8, max_width=30):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))

    def _escribir_header_tabla(ws, tab_fle_df, start_row=1, start_col=1):
        b_med  = _border_medium_all()
        b_thin = _border_thin_all()
        n_rows, _ = tab_fle_df.shape
        label_col = start_col
        data_col_start = start_col + 1
        for rel_row, (_, row_data) in enumerate(tab_fle_df.iterrows()):
            abs_row  = start_row + rel_row
            row_vals = row_data.tolist()
            label    = row_vals[0] if row_vals else ""
            data_vals = row_vals[1:]
            is_vano_row = str(label).strip().lower() in {
                "vano", "longitud (m)", "poste inicial", "poste final", "desnivel"
            }
            b = b_med if is_vano_row else b_thin
            if pd.notna(label) and str(label).strip():
                _merge_and_write(ws, abs_row, label_col, label_col + 1,
                                 value=label, font=FONT_BOLD, alignment=ALIGN_LEFT,
                                 border=b, fill=FILL_HEADER)
            for rel_col, val in enumerate(data_vals):
                c = data_col_start + 1 + rel_col
                try:
                    es_nan = pd.isna(val)
                except (TypeError, ValueError):
                    es_nan = False
                if es_nan:
                    continue
                num_fmt = None
                if isinstance(val, float):
                    num_fmt = NUM_FMT_4DEC if abs(val) < 10 else NUM_FMT_2DEC
                _write_cell(ws, abs_row, c, val, font=FONT_NORMAL,
                            alignment=ALIGN_CENTER, border=b,
                            fill=FILL_DATA, number_format=num_fmt)
        return start_row + n_rows

    def _escribir_datos_tabla(ws, tablas_df, start_row=1, start_col=1):
        b_med  = _border_medium_all()
        b_thin = _border_thin_all()
        for rel_col, col_name in enumerate(tablas_df.columns):
            _write_cell(ws, start_row, start_col + rel_col, col_name,
                        font=FONT_BOLD, alignment=ALIGN_CENTER,
                        border=b_med, fill=FILL_HEADER)
        data_start_row = start_row + 1
        for rel_row, (_, row_data) in enumerate(tablas_df.iterrows()):
            abs_row = data_start_row + rel_row
            for rel_col, val in enumerate(row_data):
                c = start_col + rel_col
                try:
                    es_nan = pd.isna(val)
                except (TypeError, ValueError):
                    es_nan = False
                if es_nan:
                    continue
                num_fmt = NUM_FMT_4DEC if isinstance(val, float) else None
                _write_cell(ws, abs_row, c, val, font=FONT_NORMAL,
                            alignment=ALIGN_CENTER, border=b_thin,
                            fill=FILL_DATA, number_format=num_fmt)
        return data_start_row + len(tablas_df)

    def _clean(value):
        try:
            if pd.isna(value):
                return None
        except (TypeError, ValueError):
            pass
        return value

    # ── 1. Cargar template (contiene hojas MEC, RET, etc. con encabezados) ───
    wb = load_workbook(ruta_template)

    # ── 2. Escribir dataframes de cálculos mecánicos ────────────────────────
    config = [
        ("MEC",                           mec,          8),
        ("RET",                           ret,          5),
        ("EOLOVANOS",                     eovanos,      3),
        ("Caracteristicas de los postes", carac_postes, 3),
        ("VANOS IDEALES DE REGULACIÓN",   van_reg,      7),
    ]

    for sheet_name, df, start_row in config:
        ws = wb[sheet_name]
        ws.delete_rows(start_row, ws.max_row - start_row + 1)

        fill = header_fills.get(sheet_name)

        # Encabezado: negrita + color de la hoja
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=col_name)
            cell.border = border
            cell.alignment = wrap
            font_color = "FFFFFF" if sheet_name in white_font_sheets else "000000"
            cell.font = Font(bold=True, color=font_color)
            if fill:
                cell.fill = fill

        # Datos
        for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=_round4(_clean(value)))
                cell.border = border
                cell.alignment = wrap

    # ── 3. Agregar hojas de flechado ─────────────────────────────────────────
    for i, (header_df, datos_df) in enumerate(zip(tab_fle, tablas_p)):
        ws = wb.create_sheet(title=f"Canton_{i + 1}")
        next_row = _escribir_header_tabla(ws, header_df)
        _escribir_datos_tabla(ws, datos_df, start_row=next_row)
        _auto_col_width(ws)

    for i, (header_df, datos_df) in enumerate(zip(tab_fle_s, tablas_s)):
        ws = wb.create_sheet(title=f"Canton_{i + 1}S")
        next_row = _escribir_header_tabla(ws, header_df)
        _escribir_datos_tabla(ws, datos_df, start_row=next_row)
        _auto_col_width(ws)

    wb.save(ruta_salida)
    print(f"✅ Archivo guardado: {ruta_salida}")


# ── Ejemplo de uso ────────────────────────────────────────────────────────────
# exportar_todo(
#     ruta_template = DATA + "calculos_mecanicos_raw.xlsx",
#     ruta_salida   = DATA + "calculos_mecanicos.xlsx",
#     mec=mec, ret=ret, eovanos=eovanos, carac_postes=carac_postes, van_reg=van_reg,
#     tab_fle=tab_fle, tablas_p=tablas_p, tab_fle_s=tab_fle_s, tablas_s=tablas_s,
# )

"""
exportar_tablas_afinia.py
=========================
Función exportar_todo_afinia: similar a exportar_todo pero:
  - Excluye los dataframes: mec, ret, carac_postes
  - Incluye eovanos, van_reg (presentes en exportar_todo)
  - Incluye además todos los dataframes del bloque "Tablas formato nuevo AFINIA":
        datos_iniciales_red_mt, informacion_del_apoyo, calculo_esfuerzos_apoyo,
        analisis_hipotesis_normales, analisis_hipotesis_anormales,
        calculo_poste_retenidas, validacion_poste_retenidas,
        tipo_retenidas_ancla, dimension_ancla
  - Las hojas de flechado (cantones normales y secundarios) se mantienen igual.

Uso típico:
-----------
    from exportar_tablas_afinia import exportar_todo_afinia

    exportar_todo_afinia(
        ruta_template = DATA + "plantilla_calculos.xlsx",
        ruta_salida   = DATA + "calculos_mecanicos_afinia.xlsx",
        eovanos                    = eovanos,
        van_reg                    = van_reg,
        tab_fle                    = tab_fle,
        tablas_p                   = tablas_p,
        tab_fle_s                  = tab_fle_s,
        tablas_s                   = tablas_s,
        datos_iniciales_red_mt     = datos_iniciales_red_mt,
        informacion_del_apoyo      = informacion_del_apoyo,
        calculo_esfuerzos_apoyo    = calculo_esfuerzos_apoyo,
        analisis_hipotesis_normales  = analisis_hipotesis_normales,
        analisis_hipotesis_anormales = analisis_hipotesis_anormales,
        calculo_poste_retenidas    = calculo_poste_retenidas,
        validacion_poste_retenidas = validacion_poste_retenidas,
        tipo_retenidas_ancla       = tipo_retenidas_ancla,
        dimension_ancla            = dimension_ancla,
    )
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Border, Side, Alignment, Font, PatternFill
)
from openpyxl.utils import get_column_letter


def exportar_todo_afinia(
    ruta_template,
    ruta_salida,
    # ── Hojas heredadas de exportar_todo (sin mec, ret ni carac_postes) ──
    eovanos,
    van_reg,
    # ── Hojas de flechado ─────────────────────────────────────────────────
    tab_fle,
    tablas_p,
    tab_fle_s,
    tablas_s,
    # ── Tablas formato nuevo AFINIA ───────────────────────────────────────
    datos_iniciales_red_mt,
    informacion_del_apoyo,
    calculo_esfuerzos_apoyo,
    analisis_hipotesis_normales,
    analisis_hipotesis_anormales,
    calculo_poste_retenidas,
    validacion_poste_retenidas,
    tipo_retenidas_ancla,
    dimension_ancla,
):
    """
    Exporta en un único archivo Excel:

    Hojas del template (preservando encabezados originales):
      - EOLOVANOS            ← eovanos
      - VANOS IDEALES DE REGULACIÓN ← van_reg

    Hojas nuevas AFINIA (creadas desde cero con formato estándar):
      - Datos Iniciales Red MT
      - Información del Apoyo
      - Cálculo Esfuerzos Apoyo
      - Análisis Hipótesis Normales
      - Análisis Hipótesis Anormales
      - Cálculo Poste Retenidas
      - Validación Poste Retenidas
      - Tipo Retenidas y Ancla
      - Dimensión Ancla

    Hojas de flechado (cantones normales y secundarios):
      - Canton_1, Canton_2, ...
      - Canton_1S, Canton_2S, ...

    Parámetros
    ----------
    ruta_template : str   – Archivo base con las hojas EOLOVANOS, VANOS IDEALES…
    ruta_salida   : str   – Ruta del archivo Excel resultante
    eovanos, van_reg : pd.DataFrame
    tab_fle, tablas_p : list[pd.DataFrame]   – Cantones normales
    tab_fle_s, tablas_s : list[pd.DataFrame] – Cantones secundarios
    datos_iniciales_red_mt, informacion_del_apoyo, calculo_esfuerzos_apoyo,
    analisis_hipotesis_normales, analisis_hipotesis_anormales,
    calculo_poste_retenidas, validacion_poste_retenidas,
    tipo_retenidas_ancla, dimension_ancla : pd.DataFrame – Tablas formato AFINIA
    """

    # ── Estilos ──────────────────────────────────────────────────────────────
    def _thin():
        return Side(style="thin", color="FF000000")

    def _medium_side():
        return Side(style="medium", color="FF000000")

    def _border_thin_all():
        t = _thin()
        return Border(left=t, right=t, top=t, bottom=t)

    def _border_medium_all():
        m = _medium_side()
        return Border(left=m, right=m, top=m, bottom=m)

    FILL_HEADER  = PatternFill("solid", fgColor="D9E1F2")   # azul claro
    FILL_AFINIA  = PatternFill("solid", fgColor="C2D59A")   # verde AFINIA
    FILL_DATA    = PatternFill("solid", fgColor="FFFFFF")
    FONT_BOLD    = Font(bold=True, size=9)
    FONT_NORMAL  = Font(bold=False, size=9)
    ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    NUM_FMT_2DEC = '0.00'
    NUM_FMT_4DEC = '0.0000'

    medium = Side(style='medium')
    border = Border(left=medium, right=medium, top=medium, bottom=medium)
    wrap   = Alignment(wrap_text=True)

    # Colores de encabezado para hojas del template
    header_fills = {
        "EOLOVANOS":                   PatternFill("solid", fgColor="FE6A0C"),
        "VANOS IDEALES DE REGULACIÓN": PatternFill("solid", fgColor="D6E3BB"),
    }
    white_font_sheets = {"EOLOVANOS"}

    # ── Helpers flechado ─────────────────────────────────────────────────────
    def _write_cell(ws, row, col, value, font=None, alignment=None,
                    border=None, fill=None, number_format=None):
        cell = ws.cell(row=row, column=col, value=_round4(value))
        if font:          cell.font = font
        if alignment:     cell.alignment = alignment
        if border:        cell.border = border
        if fill:          cell.fill = fill
        if number_format: cell.number_format = number_format
        return cell

    def _merge_and_write(ws, row, col_start, col_end, value,
                         font=None, alignment=None, border=None, fill=None):
        if col_start < col_end:
            ws.merge_cells(start_row=row, start_column=col_start,
                           end_row=row,   end_column=col_end)
        _write_cell(ws, row, col_start, value, font=font,
                    alignment=alignment, border=border, fill=fill)

    def _auto_col_width(ws, min_width=8, max_width=30):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))

    def _escribir_header_tabla(ws, tab_fle_df, start_row=1, start_col=1):
        b_med  = _border_medium_all()
        b_thin = _border_thin_all()
        n_rows, _ = tab_fle_df.shape
        label_col = start_col
        data_col_start = start_col + 1
        for rel_row, (_, row_data) in enumerate(tab_fle_df.iterrows()):
            abs_row  = start_row + rel_row
            row_vals = row_data.tolist()
            label    = row_vals[0] if row_vals else ""
            data_vals = row_vals[1:]
            is_vano_row = str(label).strip().lower() in {
                "vano", "longitud (m)", "poste inicial", "poste final", "desnivel"
            }
            b = b_med if is_vano_row else b_thin
            if pd.notna(label) and str(label).strip():
                _merge_and_write(ws, abs_row, label_col, label_col + 1,
                                 value=label, font=FONT_BOLD, alignment=ALIGN_LEFT,
                                 border=b, fill=FILL_HEADER)
            for rel_col, val in enumerate(data_vals):
                c = data_col_start + 1 + rel_col
                try:
                    es_nan = pd.isna(val)
                except (TypeError, ValueError):
                    es_nan = False
                if es_nan:
                    continue
                num_fmt = None
                if isinstance(val, float):
                    num_fmt = NUM_FMT_4DEC if abs(val) < 10 else NUM_FMT_2DEC
                _write_cell(ws, abs_row, c, val, font=FONT_NORMAL,
                            alignment=ALIGN_CENTER, border=b,
                            fill=FILL_DATA, number_format=num_fmt)
        return start_row + n_rows

    def _escribir_datos_tabla(ws, tablas_df, start_row=1, start_col=1):
        b_med  = _border_medium_all()
        b_thin = _border_thin_all()
        for rel_col, col_name in enumerate(tablas_df.columns):
            _write_cell(ws, start_row, start_col + rel_col, col_name,
                        font=FONT_BOLD, alignment=ALIGN_CENTER,
                        border=b_med, fill=FILL_HEADER)
        data_start_row = start_row + 1
        for rel_row, (_, row_data) in enumerate(tablas_df.iterrows()):
            abs_row = data_start_row + rel_row
            for rel_col, val in enumerate(row_data):
                c = start_col + rel_col
                try:
                    es_nan = pd.isna(val)
                except (TypeError, ValueError):
                    es_nan = False
                if es_nan:
                    continue
                num_fmt = NUM_FMT_4DEC if isinstance(val, float) else None
                _write_cell(ws, abs_row, c, val, font=FONT_NORMAL,
                            alignment=ALIGN_CENTER, border=b_thin,
                            fill=FILL_DATA, number_format=num_fmt)
        return data_start_row + len(tablas_df)

    def _clean(value):
        try:
            if pd.isna(value):
                return None
        except (TypeError, ValueError):
            pass
        return value

    def _escribir_df_afinia(ws, df, fill_header=None):
        """
        Escribe un DataFrame completo en una hoja desde la fila 1.
        Fila 1 = encabezados con formato; filas siguientes = datos.
        """
        if fill_header is None:
            fill_header = FILL_AFINIA

        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.border = border
            cell.alignment = wrap
            cell.font = Font(bold=True, size=9)
            cell.fill = fill_header

        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=_round4(_clean(value)))
                cell.border = border
                cell.alignment = wrap
                cell.font = FONT_NORMAL

        _auto_col_width(ws)

    # ── 1. Cargar template ────────────────────────────────────────────────────
    wb = load_workbook(ruta_template)

    # Eliminar hojas del template que esta función no exporta
    hojas_excluidas = ["MEC", "RET", "Caracteristicas de los postes"]
    for nombre in hojas_excluidas:
        if nombre in wb.sheetnames:
            del wb[nombre]

    # ── 2. Escribir hojas heredadas del template (EOLOVANOS y VAN_REG) ───────
    config_template = [
        ("EOLOVANOS",                   eovanos,  3),
        ("VANOS IDEALES DE REGULACIÓN", van_reg,  7),
    ]

    for sheet_name, df, start_row in config_template:
        ws = wb[sheet_name]
        ws.delete_rows(start_row, ws.max_row - start_row + 1)

        fill = header_fills.get(sheet_name)
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=col_name)
            cell.border = border
            cell.alignment = wrap
            font_color = "FFFFFF" if sheet_name in white_font_sheets else "000000"
            cell.font = Font(bold=True, color=font_color)
            if fill:
                cell.fill = fill

        for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=_round4(_clean(value)))
                cell.border = border
                cell.alignment = wrap

    # ── 3. Agregar hojas de tablas formato nuevo AFINIA ──────────────────────
    tablas_afinia = [
        ("Datos Iniciales Red MT",        datos_iniciales_red_mt),
        ("Información del Apoyo",         informacion_del_apoyo),
        ("Cálculo Esfuerzos Apoyo",       calculo_esfuerzos_apoyo),
        ("Análisis Hip. Normales",        analisis_hipotesis_normales),
        ("Análisis Hip. Anormales",       analisis_hipotesis_anormales),
        ("Cálculo Poste Retenidas",       calculo_poste_retenidas),
        ("Validación Poste Retenidas",    validacion_poste_retenidas),
        ("Tipo Retenidas y Ancla",        tipo_retenidas_ancla),
        ("Dimensión Ancla",               dimension_ancla),
    ]

    hojas_escritas = []
    hojas_omitidas = []
    for sheet_name, df in tablas_afinia:
        # Omitir hojas cuyo DataFrame no tenga filas (ej. proyectos sin retenidas)
        if df is None or len(df) == 0:
            hojas_omitidas.append(sheet_name)
            continue
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(title=sheet_name)
        _escribir_df_afinia(ws, df)
        hojas_escritas.append(sheet_name)

    # ── 4. Agregar hojas de flechado (cantones normales y secundarios) ────────
    for i, (header_df, datos_df) in enumerate(zip(tab_fle, tablas_p)):
        ws = wb.create_sheet(title=f"Canton_{i + 1}")
        next_row = _escribir_header_tabla(ws, header_df)
        _escribir_datos_tabla(ws, datos_df, start_row=next_row)
        _auto_col_width(ws)

    for i, (header_df, datos_df) in enumerate(zip(tab_fle_s, tablas_s)):
        ws = wb.create_sheet(title=f"Canton_{i + 1}S")
        next_row = _escribir_header_tabla(ws, header_df)
        _escribir_datos_tabla(ws, datos_df, start_row=next_row)
        _auto_col_width(ws)

    wb.save(ruta_salida)
    omitidas_str = (
        f"\n   ⚠️  Hojas omitidas (sin datos): {', '.join(hojas_omitidas)}"
        if hojas_omitidas else ""
    )
    print(
        f"✅ Archivo guardado: {ruta_salida}\n"
        f"   • Hojas template    : EOLOVANOS, VANOS IDEALES DE REGULACIÓN\n"
        f"   • Tablas AFINIA     : {len(hojas_escritas)} hojas{omitidas_str}\n"
        f"   • Cantones normales : {len(tab_fle)}\n"
        f"   • Cantones secund.  : {len(tab_fle_s)}"
    )


# ── Ejemplo de uso ────────────────────────────────────────────────────────────
# exportar_todo_afinia(
#     ruta_template              = DATA + "plantilla_calculos.xlsx",
#     ruta_salida                = DATA + "calculos_mecanicos_afinia.xlsx",
#     eovanos                    = eovanos,
#     van_reg                    = van_reg,
#     tab_fle                    = tab_fle,
#     tablas_p                   = tablas_p,
#     tab_fle_s                  = tab_fle_s,
#     tablas_s                   = tablas_s,
#     datos_iniciales_red_mt     = datos_iniciales_red_mt,
#     informacion_del_apoyo      = informacion_del_apoyo,
#     calculo_esfuerzos_apoyo    = calculo_esfuerzos_apoyo,
#     analisis_hipotesis_normales  = analisis_hipotesis_normales,
#     analisis_hipotesis_anormales = analisis_hipotesis_anormales,
#     calculo_poste_retenidas    = calculo_poste_retenidas,
#     validacion_poste_retenidas = validacion_poste_retenidas,
#     tipo_retenidas_ancla       = tipo_retenidas_ancla,
#     dimension_ancla            = dimension_ancla,
# )



# ─────────────────────────────────────────────────────────────────────────────
# exportar_nuevo_formato
# Escribe los DataFrames sobre la plantilla calculos_mecanicos_afinia.xlsx
# conservando íntegramente todos los encabezados (celdas combinadas, colores,
# fuentes). Configuración de filas extraída del Diccionario_Template.
# ─────────────────────────────────────────────────────────────────────────────
def exportar_nuevo_formato(
    ruta_template,
    ruta_salida,
    eovanos,
    van_reg,
    tab_fle,
    tablas_p,
    tab_fle_s,
    tablas_s,
    datos_iniciales_red_mt,
    informacion_del_apoyo,
    calculo_esfuerzos_apoyo,
    analisis_hipotesis_normales,
    analisis_hipotesis_anormales,
    calculo_poste_retenidas=None,
    validacion_poste_retenidas=None,
    tipo_retenidas_ancla=None,
    dimension_ancla=None,
):
    """
    Vuelca los DataFrames sobre la plantilla nueva (calculos_mecanicos_afinia.xlsx)
    conservando todo el formato de encabezados original.

    Configuración por hoja (del Diccionario_Template):
      header_row = fila que contiene los nombres de columna (para mapeo DF→Excel)
      data_start = primera fila donde se escriben los datos

      Hoja                       | header_row | data_start
      Datos Iniciales Red MT     |     2      |     3      (dict dice 1, col names reales en fila 2)
      VANOS IDEALES DE REGULACIÓN|     6      |     7
      EOLOVANOS                  |     3      |     4
      Información del Apoyo      |     4      |     5
      Cálculo Esfuerzos Apoyo    |     2      |     3
      Análisis Hip. Normales     |     3      |     4
      Análisis Hip. Anormales    |     3      |     4
      Cálculo Poste Retenidas    |     2      |     3
      Validación Poste Retenidas |     2      |     3
      Tipo Retenidas y Ancla     |     2      |     3
      Dimensión Ancla            |     1      |     2
    """
    import pandas as pd
    import unicodedata
    import copy
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side

    # ── Configuración: (header_row, data_start) ───────────────────────────────
    SHEET_CFG = {
        "Datos Iniciales Red MT":      (2, 3),
        "VANOS IDEALES DE REGULACIÓN": (6, 7),
        "EOLOVANOS":                   (3, 4),
        "Información del Apoyo":       (4, 5),
        "Cálculo Esfuerzos Apoyo":     (2, 3),
        "Análisis Hip. Normales":      (3, 4),
        "Análisis Hip. Anormales":     (3, 4),
        "Cálculo Poste Retenidas":     (2, 3),
        "Validación Poste Retenidas":  (2, 3),
        "Tipo Retenidas y Ancla":      (2, 3),
        "Dimensión Ancla":             (1, 2),
    }

    # ── Estilo fijo de celdas de datos (igual en todas las hojas del template) ─
    _medium = Side(style="medium", color="FF000000")
    DATA_BORDER = Border(left=_medium, right=_medium, top=_medium, bottom=_medium)
    DATA_FONT   = Font(name="Arial", size=7, bold=False)
    DATA_ALIGN  = Alignment(wrap_text=True)

    # ── Helpers ────────────────────────────────────────────────────────────────
    def _clean(val):
        try:
            if pd.isna(val):
                return None
        except (TypeError, ValueError):
            pass
        return val

    def _val(val):
        v = _clean(val)
        return round(v, 4) if isinstance(v, float) else v

    def _norm(text):
        if text is None:
            return ""
        s = str(text).strip().lower()
        s = unicodedata.normalize("NFKD", s)
        return "".join(c for c in s if not unicodedata.combining(c))

    def _build_col_map(ws, header_row):
        """Devuelve {nombre_normalizado: col_idx} leyendo la fila header_row."""
        col_map = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            if v is not None:
                col_map[_norm(v)] = c
        return col_map

    def _map_columns(df_cols, col_map):
        """Empareja columnas del DF con columnas de Excel por nombre normalizado."""
        result = []
        for col in df_cols:
            key = _norm(col)
            idx = col_map.get(key)
            if idx is None:                          # búsqueda parcial como fallback
                for mk, mc in col_map.items():
                    if key in mk or mk in key:
                        idx = mc
                        break
            result.append((col, idx))
        return result

    def _write_df(ws, df, header_row, data_start):
        """
        Preserva las filas 1…data_start-1 del template, borra desde data_start
        en adelante y escribe el DataFrame con el estilo de datos del template.
        El mapeo DF→columnas Excel se hace por nombre contra header_row.
        """
        if df is None or len(df) == 0:
            return

        # 1. Leer mapa de columnas ANTES de borrar
        col_map = _build_col_map(ws, header_row)
        mapping = _map_columns(list(df.columns), col_map)

        # 2. Borrar filas de datos existentes
        n_delete = ws.max_row - data_start + 1
        if n_delete > 0:
            ws.delete_rows(data_start, n_delete)

        # 3. Escribir nuevos datos
        for rel, (_, row) in enumerate(df.iterrows()):
            abs_row = data_start + rel
            for col_name, xcol in mapping:
                if xcol is None:
                    continue
                cell = ws.cell(abs_row, xcol, value=_val(row[col_name]))
                cell.font      = DATA_FONT
                cell.border    = DATA_BORDER
                cell.alignment = DATA_ALIGN

    # ── 1. Cargar plantilla ────────────────────────────────────────────────────
    wb = load_workbook(ruta_template)

    # ── 2. Hojas de cálculo (EOLOVANOS y VANOS IDEALES) ───────────────────────
    for sname, df in [("EOLOVANOS", eovanos), ("VANOS IDEALES DE REGULACIÓN", van_reg)]:
        hr, ds = SHEET_CFG[sname]
        _write_df(wb[sname], df, hr, ds)

    # ── 3. Tablas AFINIA ───────────────────────────────────────────────────────
    tablas = [
        ("Datos Iniciales Red MT",      datos_iniciales_red_mt),
        ("Información del Apoyo",       informacion_del_apoyo),
        ("Cálculo Esfuerzos Apoyo",     calculo_esfuerzos_apoyo),
        ("Análisis Hip. Normales",      analisis_hipotesis_normales),
        ("Análisis Hip. Anormales",     analisis_hipotesis_anormales),
        ("Cálculo Poste Retenidas",     calculo_poste_retenidas),
        ("Validación Poste Retenidas",  validacion_poste_retenidas),
        ("Tipo Retenidas y Ancla",      tipo_retenidas_ancla),
        ("Dimensión Ancla",             dimension_ancla),
    ]
    escritas, omitidas = [], []
    for sname, df in tablas:
        if df is None or len(df) == 0:
            omitidas.append(sname)
            continue
        if sname not in wb.sheetnames:
            omitidas.append(f"{sname} (no existe en template)")
            continue
        hr, ds = SHEET_CFG[sname]
        _write_df(wb[sname], df, hr, ds)
        escritas.append(sname)

    # ── 4. Hojas de flechado (cantones) ───────────────────────────────────────
    for s in [s for s in wb.sheetnames if s.startswith("Canton_")]:
        del wb[s]

    for i, (hdr, dat) in enumerate(zip(tab_fle, tablas_p)):
        ws = wb.create_sheet(title=f"Canton_{i + 1}")
        nxt = _escribir_header_tabla(ws, hdr, start_row=1, start_col=1)
        _escribir_datos_tabla(ws, dat, start_row=nxt, start_col=1)
        _auto_col_width(ws)

    for i, (hdr, dat) in enumerate(zip(tab_fle_s, tablas_s)):
        ws = wb.create_sheet(title=f"Canton_{i + 1}S")
        nxt = _escribir_header_tabla(ws, hdr, start_row=1, start_col=1)
        _escribir_datos_tabla(ws, dat, start_row=nxt, start_col=1)
        _auto_col_width(ws)

    # ── 5. Guardar ────────────────────────────────────────────────────────────
    wb.save(ruta_salida)

    om_str = f"\n   ⚠️  Omitidas: {', '.join(omitidas)}" if omitidas else ""
    print(
        f"✅ Guardado: {ruta_salida}\n"
        f"   • Tablas AFINIA     : {len(escritas)} hojas{om_str}\n"
        f"   • Cantones normales : {len(tab_fle)}\n"
        f"   • Cantones secund.  : {len(tab_fle_s)}"
    )


# ── Ejemplo de uso ─────────────────────────────────────────────────────────────
# exportar_nuevo_formato(ruta_template=DATA+"calculos_mecanicos_afinia.xlsx", ruta_salida=DATA+"calculos_mecanicos_nuevo.xlsx", eovanos=eovanos, van_reg=van_reg, tab_fle=tab_fle, tablas_p=tablas_p, tab_fle_s=tab_fle_s, tablas_s=tablas_s, datos_iniciales_red_mt=datos_iniciales_red_mt, informacion_del_apoyo=informacion_del_apoyo, calculo_esfuerzos_apoyo=calculo_esfuerzos_apoyo, analisis_hipotesis_normales=analisis_hipotesis_normales, analisis_hipotesis_anormales=analisis_hipotesis_anormales, calculo_poste_retenidas=calculo_poste_retenidas, validacion_poste_retenidas=validacion_poste_retenidas, tipo_retenidas_ancla=tipo_retenidas_ancla, dimension_ancla=dimension_ancla)

import numpy as np
import pandas as pd
import re 
import math 
def kgf_a_daN(F_kgf, g=9.8066500):
    """
    Convierte kilogramo-fuerza (kgf) a decanewton (daN).

    Conversión:
        1 kgf = g N
        1 daN = 10 N
        daN = (kgf * g) / 10
    """
    return F_kgf * g / 10.0

def N_a_daN(F_N, g=9.8066500):
    """
    Convierte N a daN.
    La constante g se mantiene solo para estandarización,
    aunque no interviene en la conversión directa.

    Fórmula:
        1 daN = 10 N
        daN = N / 10
    """
    return F_N / 10.0

def suma_vectores(magnitudes, angulos_relativos_deg):
    """
    Calcula la magnitud de la suma de N vectores dados por:
    - magnitudes[i]: magnitud del vector i
    - angulos_relativos_deg[i]: ángulo del vector i respecto al primer vector (en grados)

    El primer vector se asume con ángulo absoluto 0°.
    El resto se posiciona sumándole el ángulo relativo.

    Parámetros:
        magnitudes (list or array): magnitudes de los vectores [m1, m2, ..., mn]
        angulos_relativos_deg (list or array): ángulos relativos respecto al primer vector (en grados)
        Debe tener longitud n, donde el primer ángulo debe ser 0.

    Retorna:
        tuple: (magnitud_total, suma_x, suma_y)
    """

    n = len(magnitudes)

    if len(angulos_relativos_deg) != n:
        raise ValueError("La lista de ángulos debe tener la misma longitud que la de magnitudes.")

    # Convertir grados a radianes
    ang_rad = np.radians(angulos_relativos_deg)

    # Componentes X y Y
    comp_x = np.sum(magnitudes * np.cos(ang_rad))
    comp_y = np.sum(magnitudes * np.sin(ang_rad))

    # Magnitud total del vector resultante
    magnitud_total = np.sqrt(comp_x**2 + comp_y**2)

    return magnitud_total

""" print(suma_vectores([1,1,1],[0,90,-90])) """

def vano_regulacion(vanos_m, desniveles_m, usar_k_truxa=True):
    """
    Calcula el vano ideal de regulación de un cantón según el método de Truxá.

    Parámetros
    ----------
    vanos_m : array-like
        Longitud horizontal de cada vano (a_i), en metros.
    desniveles_m : array-like o None
        Desnivel de cada vano (b_i), en metros.
        Si es None, se supone bi = 0 para todos los vanos (cantón nivelado).
    usar_k_truxa : bool, opcional (default = True)
        Si True, aplica el factor de Truxá k.
        Si False, asume k = 1 (equivalente a ignorar el desnivel en el vano ideal).

    Returns
    -------
    ar : float
        Longitud del vano ideal de regulación (m).
    k : float
        Factor de Truxá utilizado (k = 1 para vanos nivelados).
    """

    # Convertir a arrays
    a = np.asarray(vanos_m, dtype=float)

    if a.ndim != 1:
        raise ValueError("vanos_m debe ser un vector 1D (lista o array de vanos).")

    if desniveles_m is None:
        b = np.zeros_like(a)
    else:
        b = np.asarray(desniveles_m, dtype=float)
        if b.shape != a.shape:
            raise ValueError("vanos_m y desniveles_m deben tener la misma longitud.")

    # Longitud real de cada vano (á_i)
    a_real = np.sqrt(a**2 + b**2)

    # Vano equivalente base (caso nivelado)
    suma_a3 = np.sum(a**3)
    suma_a = np.sum(a)

    if suma_a <= 0:
        raise ValueError("La suma de los vanos debe ser mayor que cero.")

    ar_base = np.sqrt(suma_a3 / suma_a)

    if usar_k_truxa:
        # Factor de Truxá k (forma reconstruida a partir de formularios típicos)
        # Propiedades:
        #  - adimensional
        #  - si a_real == a  => k = 1
        num = np.sum(a_real**3) * np.sum(a**2)
        den = np.sum(a**3) * np.sum(a * a_real)

        if den <= 0:
            raise ValueError("Datos de vanos/desniveles inválidos (denominador de k <= 0).")

        k = np.sqrt(num / den)
    else:
        k = 1.0

    # Vano ideal de regulación
    ar = k * ar_base
    return ar

""" print(vano_regulacion([21,22,24,45,23],[1,0,2,3,0.5])) """

def identificar_poste(codigo: str, detallado: bool = False):
    """
    Identifica el tipo de poste según el código de armado de AFINIA.

    Si detallado=False → retorna solo las siglas del tipo de poste: FL, AL, ANG, ANC.
    Si detallado=True  → retorna un diccionario con información completa.
    Si el código no es válido → retorna np.nan
    """

    try:
        # --- Validación básica ---
        if not isinstance(codigo, str) or "-" not in codigo:
            return np.nan

        parte_armado, parte_tension = codigo.split("-")

        # Letras y números
        letras = ''.join(c for c in parte_armado if c.isalpha())
        numeros = ''.join(c for c in parte_armado if c.isdigit())

        if len(numeros) != 3 or len(letras) < 2:
            return np.nan

        # --- Interpretación de letras ---
        nivel_tension = letras[:2]
        if nivel_tension == "BT":
            nivel = "Baja Tensión"
        elif nivel_tension == "MT":
            nivel = "Media Tensión"
        else:
            return np.nan  # nivel no reconocido

        tipo_cable = "Forrado" if (len(letras) == 3 and letras[2] == "F") else "Desnudo"

        # --- Interpretación de dígitos ---
        d1, d2, d3 = map(int, numeros)

        # Armado general
        if d1 == 6:
            armado_general = "Autosoportado (1 circuito)"
        elif d1 == 7:
            armado_general = "Autosoportado (2 circuitos)"
        else:
            return np.nan

        # Fases
        if d2 == 3:
            fases = "Trifásico"
        elif d2 == 2:
            fases = "Bifásico"
        else:
            return np.nan

        # Tipo de poste
        if d3 == 1:
            sigla_poste = "FL"
            tipo_poste = "Fin de Línea"
        elif d3 == 2:
            sigla_poste = "AL"
            tipo_poste = "Alineación"
        elif d3 == 3:
            sigla_poste = "ANG"
            tipo_poste = "Ángulo"
        elif d3 in (4, 5):
            sigla_poste = "ANC"
            tipo_poste = "Anclaje"
        else:
            return np.nan

        # Tensión del circuito
        if parte_tension == "1":
            tension = "13.2 kV"
        elif parte_tension == "2":
            tension = "34.5 kV"
        else:
            return np.nan

        # --- Salida ---
        if not detallado:
            return sigla_poste

        return {
            "Código": codigo,
            "Sigla": sigla_poste,
            "Tipo de Poste": tipo_poste,
            "Nivel de Tensión": nivel,
            "Tipo de Cable": tipo_cable,
            "Armado General": armado_general,
            "Fases": fases,
            "Tensión del Circuito": tension
        }

    except Exception:
        return np.nan


def calcular_cantones(armados, rutas, postes, vanos_adelante, detallado=False):
    """
    Calcula la longitud de los cantones de una línea de MT a partir de:
    - armados: lista de códigos de armado
    - rutas: lista que indica la ruta/derivación a la que pertenece cada poste
    - postes: nombres/identificadores de cada poste
    - vanos_adelante: distancia al siguiente poste de la misma ruta
    """

    # Agrupar índices de postes por ruta
    rutas_dict = {}
    for i, ruta in enumerate(rutas):
        rutas_dict.setdefault(ruta, []).append(i)

    # Asegurar que cada ruta queda en el orden en que aparecen
    for r in rutas_dict:
        rutas_dict[r].sort()

    cantones = []
    num_canton = 0

    for nombre_ruta, indices in rutas_dict.items():
        if len(indices) < 2:
            continue  # una ruta con 1 poste no genera canton

        # inicio del primer cantón
        inicio = indices[0]
        longitud = 0.0

        for j in range(len(indices) - 1):
            actual = indices[j]
            siguiente = indices[j + 1]

            # sumar el vano desde actual → siguiente
            longitud += vanos_adelante[actual]

            # identificar tipo del siguiente poste
            tipo_sig = identificar_poste(armados[siguiente])

            # condiciones para cerrar el cantón
            fin_por_tipo = tipo_sig in ("FL", "ANC")
            fin_por_ruta = (j + 1 == len(indices) - 1)

            if fin_por_tipo or fin_por_ruta:
                num_canton += 1

                if not detallado:
                    cantones.append(longitud)
                else:
                    cantones.append({
                        "canton": num_canton,
                        "ruta": nombre_ruta,
                        "poste_inicio": postes[inicio],
                        "poste_fin": postes[siguiente],
                        "longitud": longitud
                    })

                # reiniciar para el siguiente cantón
                inicio = siguiente
                longitud = 0.0

    return cantones

""" armados = ["MTF331-2", "MTF332-1","MTF334-1", "MTF332-1","MTF334-1", "MTF331-1"]
rutas   = ["ruta1",     "ruta1",   "ruta1",  "ruta1",     "ruta2",     "ruta2"]
postes  = ["EPP01",     "EPP02",  "EPP03",      "EPP04",     "EPP05", "EPP06"]
vanos_adelante = [25, 30,45, 0,24,0]

print(calcular_cantones(armados, rutas, postes, vanos_adelante, detallado=True)) """

import re

def extraer_datos_poste(cadena):
    """
    Extrae la altura del poste y la carga de rotura
    desde un string con formato: "PH ##/#### kg-f".
    
    Parámetros
    ----------
    cadena : str
        Texto con el formato del poste: "PH ##/#### kg-f"

    
    Retorna
    -------
    (altura, carga, altura libre, altura del esfuerzo)
        altura : int     → en metros
        carga_daN : float → capacidad en daN
    """

    # Buscar el patrón "PH XX/YYYY"
    patron = r"PH\s*(\d{2})/(\d{3,4})"
    match = re.search(patron, cadena.upper())

    if not match:
        raise ValueError(f"Formato no válido: {cadena}")

    altura = int(match.group(1))
    altura_libre = altura-altura*0.1-0.7
    altura_esfuerzo = altura_libre-0.2
    carga = int(match.group(2))

    

    return altura, carga, altura_libre, altura_esfuerzo

""" print(extraer_datos_poste("PH 12/1050 kg-f")) """


def construir_c2t1(
    tabla1: pd.DataFrame,
    c1t1: pd.Series,
    c2t1: str,
    c1t2: pd.Series,
    c2t2: pd.Series
):
    """
    Construye o actualiza la columna `c2t1` en tabla1 a partir de Series externas.

    Reglas:
    - Para cada valor en c1t1 se buscan coincidencias en c1t2.
    - Se toman los valores correspondientes de c2t2.
    - Se ignoran NaN, "-", cadenas vacías y 0.
    - Si queda un único valor válido, se asigna.
    - Si no hay valores válidos, se asigna NaN.
    - Si hay más de un valor válido distinto, se lanza error.

    La función modifica tabla1 in-place y retorna el DataFrame.
    """

    if len(c1t2) != len(c2t2):
        raise ValueError("c1t2 y c2t2 deben tener la misma longitud")

    ref = pd.DataFrame({
        "key": c1t2,
        "value": c2t2
    })

    resultados = []

    for valor in c1t1:
        valores_validos = (
            ref.loc[ref["key"] == valor, "value"]
            .replace([0, "-", ""], pd.NA)
            .dropna()
            .unique()
        )

        if len(valores_validos) == 0:
            resultados.append(pd.NA)
        elif len(valores_validos) == 1:
            resultados.append(valores_validos[0])
        else:
            raise ValueError(
                f"Conflicto para '{valor}': valores distintos {list(valores_validos)}"
            )

    tabla1[c2t1] = resultados
    return tabla1

def construir_c2t1_vano(
    tabla1: pd.DataFrame,
    c1t1: pd.Series,
    c2t1: str,
    c1t2: pd.Series,
    c2t2: pd.Series
):
    """
    Construye o actualiza la columna `c2t1` en tabla1 tomando el valor máximo válido.

    Reglas:
    - Se ignoran NaN, '-', cadenas vacías y 0.
    - Si no hay valores válidos → NaN.
    - Si hay uno o más valores válidos → se asigna el VALOR MÁXIMO.

    La función modifica tabla1 in-place y retorna el DataFrame.
    """

    if len(c1t2) != len(c2t2):
        raise ValueError("c1t2 y c2t2 deben tener la misma longitud")

    ref = pd.DataFrame({
        "key": c1t2,
        "value": c2t2
    })

    resultados = []

    for valor in c1t1:
        valores_validos = (
            ref.loc[ref["key"] == valor, "value"]
            .replace([0, "-", ""], pd.NA)
            .dropna()
        )

        resultados.append(
            pd.NA if valores_validos.empty else valores_validos.max()
        )

    tabla1[c2t1] = resultados
    return tabla1


def convertir_texto_kgf_a_daN(texto: str) -> str:
    """
    Convierte expresiones del tipo 'PH ##/#### kg-f' a 'PH ##/XXX daN'.

    - Extrae el valor numérico después del slash (/)
    - Convierte de kgf a daN
    - Redondea hacia arriba a la unidad más cercana
    - Reemplaza 'kg-f' por 'daN'
    """

    patron = r"(.*?/)(\d+)(\s*kg-f)"

    match = re.search(patron, texto)
    if not match:
        raise ValueError(f"Formato no reconocido: {texto}")

    prefijo = match.group(1)        # 'PH 12/'
    valor_kgf = float(match.group(2))
    
    valor_daN = kgf_a_daN(valor_kgf)
    valor_daN_red = round(valor_daN)

    return f"{prefijo}{valor_daN_red} daN"

""" texto = "PH 12/1350 kg-f"
resultado = convertir_texto_kgf_a_daN(texto)

print(resultado) """

def limpiar_saltos_linea_columnas(df):
    """
    Reemplaza saltos de línea '\\n' por espacios simples en los nombres
    de columnas, incluyendo MultiIndex.
    """

    if isinstance(df.columns, pd.MultiIndex):
        df.columns = pd.MultiIndex.from_tuples(
            tuple(
                str(level).replace("\n", " ").strip()
                for level in col
            )
            for col in df.columns
        )
    else:
        df.columns = (
            df.columns
            .astype(str)
            .str.replace("\n", " ", regex=False)
            .str.strip()
        )

    return df




def extraer_series_por_indice(
    df: pd.DataFrame,
    nombre: str,
    nivel: int = 1
) -> list[pd.Series]:
    """
    Extrae todas las Series de un DataFrame con columnas MultiIndex
    cuyo nombre en un nivel dado coincide con `nombre`, excluyendo
    aquellas columnas que estén completamente vacías o inválidas.

    Se consideran valores inválidos:
    - NaN
    - 0
    - "-"
    - cadenas vacías

    Parámetros
    ----------
    df : pd.DataFrame
        DataFrame con columnas MultiIndex.
    nombre : str
        Nombre a buscar en el nivel especificado.
    nivel : int, default 1
        Nivel del MultiIndex donde se realizará la búsqueda.

    Retorna
    -------
    list[pd.Series]
        Lista de Series válidas encontradas.
    """

    if not isinstance(df.columns, pd.MultiIndex):
        raise TypeError("El DataFrame no tiene columnas MultiIndex")

    series_validas = []

    for col in df.columns:
        if col[nivel] != nombre:
            continue

        serie = df[col]

        # Normalización de valores inválidos
        serie_limpia = (
            serie
            .replace([0, "-", ""], pd.NA)
            .dropna()
        )

        # Si después de limpiar no queda nada, se ignora la columna
        if serie_limpia.empty:
            continue

        series_validas.append(serie)

    return series_validas

def sumar_lista_series(lista):
    """
    Suma fila a fila una lista de pd.Series.
    Retorna una pd.Series o None si la lista está vacía.
    """
    if not lista:
        return None
    df = pd.concat(lista, axis=1)
    return df.replace("-", np.nan).astype(float).sum(axis=1)



def deflexion_a_angulo(delta, grados=True):
    """
    Convierte un ángulo de deflexión en el ángulo real entre dos vanos.

    Parámetros
    ----------
    delta : float, array-like o pandas Series
        Ángulo de deflexión (entre la prolongación de un vano y el siguiente).
    grados : bool, default=True
        True si delta está en grados.
        False si delta está en radianes.

    Retorna
    -------
    float, array-like o pandas Series
        Ángulo entre los dos vanos.
    """
    if grados:
        return 180.0 - delta
    else:
        return np.pi - delta
    

import numpy as np
import pandas as pd


def calcular_ftvc_flmc(
    tabla,
    o_postes,
    l_postes,
    angulo_b,          # ángulo de DEFLEXIÓN δ (grados)
    f_viento_at,
    f_viento_ad,
    tiro_at,
    tiro_ad,
    col_ftvc="FTVC",
    col_flmc="FLMC"
):
    """
    Calcula simultáneamente:
    - FTVC: esfuerzo transversal por viento
    - FLMC: esfuerzo longitudinal mecánico combinado

    Geometría:
    - Caso 1 (sin derivaciones): eje = bisectriz del ángulo de deflexión
    - Caso 2 (con derivaciones):
        * Se calcula el vector resultante de tensiones
        * Se calcula la fuerza sobre conductores por el viento tomando  
            el viento como perpendicular al vano atrás del poste que tiene 
            fuerzas adelante  y atrás o del poste de referencia (angulo 0) 
        *se define el eje transversal con la misma orientación que el vector de fuerzas del viento
        *se proyecta el vector de tensiones sobre cada eje y se suma con el vector de viento para obtener ambas fuerzas
    """

    fv_at = sumar_lista_series(f_viento_at)
    fv_ad = sumar_lista_series(f_viento_ad)
    ta = sumar_lista_series(tiro_at)
    td = sumar_lista_series(tiro_ad)

    tabla[col_ftvc] = np.nan
    tabla[col_flmc] = np.nan

    for poste in o_postes:

        mask = l_postes == poste
        n_rep = mask.sum()
        if n_rep == 0:
            continue
        
        delta = np.deg2rad(angulo_b[mask].astype(float))

        fv_at_p = fv_at[mask] if fv_at is not None else pd.Series(0, index=delta.index)
        fv_ad_p = fv_ad[mask] if fv_ad is not None else pd.Series(0, index=delta.index)
        ta_p = ta[mask] if ta is not None else pd.Series(0, index=delta.index)
        td_p = td[mask] if td is not None else pd.Series(0, index=delta.index)

        # ============================================================
        # CASO 1: SIN DERIVACIONES
        # ============================================================
        if n_rep == 1:

            d = delta.iloc[0]
            sen_d2 = np.sin(d / 2)
            cos_d2 = np.cos(d / 2)

            # ---------------- FTVC ----------------
            # El viento YA está en la bisectriz → NO se normaliza 
            # (en redlin cuando se tiene un poste con vanos adelante 
            # y atrás el programa calcula la fuerza del viento proyectado
            # sobre transversal al eje de la bisectriz, sin embargo como aquí 
            # se calculan fuerzas longitudinales y transversales se normaliza  
            # para poder obtener la proyección sobre el eje longitudinal )

            ftvc = (
                fv_at_p.iloc[0]
                + fv_ad_p.iloc[0]
                + (ta_p.iloc[0] + td_p.iloc[0]) * sen_d2  #si el poste está en angulo las tensiones tienen una componente transversal en la fuerza
            )

            # ---------------- FLMC ----------------
            # Normalizar SOLO si hay viento adelante y atrás
            if fv_at_p.iloc[0] > 0 and fv_ad_p.iloc[0] > 0:
                fv_at_c = fv_at_p.iloc[0] / cos_d2  
                fv_ad_c = fv_ad_p.iloc[0] / cos_d2
            else:
                fv_at_c = fv_at_p.iloc[0]
                fv_ad_c = fv_ad_p.iloc[0]

            flmc = (
                (td_p.iloc[0] - ta_p.iloc[0]) * cos_d2
                + (fv_at_c - fv_ad_c) * sen_d2  #si el poste está en angulo las fuerzas del viento tienen una componente longitudinal en la fuerza
            )

            tabla.loc[tabla[o_postes.name] == poste, col_ftvc] = ftvc
            tabla.loc[tabla[o_postes.name] == poste, col_flmc] = flmc
            continue

        # ============================================================
        # CASO 2: CON DERIVACIONES
        # ============================================================

        theta = np.pi - delta  # ángulo real entre vanos

        # Normalización del viento SOLO si hay adelante y atrás
        fv_at_c = fv_at_p.copy()
        fv_ad_c = fv_ad_p.copy()

        for idx in delta.index:
            if fv_at_p.loc[idx] > 0 and fv_ad_p.loc[idx] > 0:
                fv_at_c.loc[idx] /= np.cos(delta.loc[idx] / 2)
                fv_ad_c.loc[idx] /= np.cos(delta.loc[idx] / 2)

        # ------------------------------------------------------------
        # 1) Vector resultante de tensiones
        # ------------------------------------------------------------
        T_res = np.array([0.0, 0.0])

        for idx in delta.index:
            th = theta.loc[idx]
            dt = delta.loc[idx]

            if ta_p.loc[idx] > 0 and td_p.loc[idx] > 0:
                # eje alineado con tiro atrás

                tx  = ta_p.loc[idx]*np.cos(0) + td_p.loc[idx]*np.cos(th)
                ty = ta_p.loc[idx]*np.sin(0) + td_p.loc[idx]*np.sin(th)
                
            else:
                T = ta_p.loc[idx] if ta_p.loc[idx] > 0 else td_p.loc[idx]
                tx = T*np.cos(th) if dt != 0 else T*np.cos(dt) #se agrega un condicional para reconocer el poste de referencia, en caso contrario como se convierte el angulo de deflexión a angulo real entonces quedaría con la direccción contraria
                ty = T*np.sin(th) if dt != 0 else T*np.sin(dt)
                

            T_res += np.array([
                tx,
                ty
            ])



        # ------------------------------------------------------------
        # 2) Proyección de fuerzas
        # ------------------------------------------------------------
        V_vec = np.array([0.0, 0.0])

        for idx in delta.index:

            th = theta.loc[idx]

            # ---------- VIENTO ----------
            if fv_at_c.loc[idx] > 0 and fv_ad_c.loc[idx] > 0:
                
                
                # se define la dirección del viento a + 90° del vano atrás
                v1a = np.array([
                    fv_at_c.loc[idx] * np.cos( np.pi / 2),
                    fv_at_c.loc[idx] * np.sin( np.pi / 2)
                ])

                v1d = np.array([
                    fv_ad_c.loc[idx] * np.cos(th + np.pi / 2),
                    fv_ad_c.loc[idx] * np.sin(th + np.pi / 2)
                ])            
                v2d = np.array([
                    fv_ad_c.loc[idx] * np.cos(th - np.pi / 2),
                    fv_ad_c.loc[idx] * np.sin(th - np.pi / 2)
                ])
                
                Vvecd = v1d if np.dot(v1d, v1a) >= 0 else v2d
                V_vect = v1a + Vvecd
                

            
            else:
                Fv = fv_at_c.loc[idx] if fv_at_c.loc[idx] > 0 else fv_ad_c.loc[idx]
                v1 = np.array([
                    Fv * np.cos(th + np.pi / 2),
                    Fv * np.sin(th + np.pi / 2)
                ])
                v2 = np.array([
                    Fv * np.cos(th - np.pi / 2),
                    Fv * np.sin(th - np.pi / 2)
                ])
                # Dirección del viento tomando el vano atrás como base
                V_vect = v1 if np.dot(v1,[0,1]) >= 0 else v2   
            V_vec += V_vect
                    
        if np.linalg.norm(V_vec) == 0:
            e_L = np.array([1.0, 0.0])
        else:
            e_T = V_vec / np.linalg.norm(V_vec)

        e_L = np.array([-e_T[1], e_T[0]])


        flmc = np.dot(V_vec, e_L) +  np.dot(T_res, e_L)
        ftvc = np.dot(V_vec, e_T) +  np.dot(T_res, e_T)
        tabla.loc[tabla[o_postes.name] == poste, col_flmc] = flmc
        
        tabla.loc[tabla[o_postes.name] == poste, col_ftvc] = ftvc

    return tabla

########### Prueba función ####################

""" tabla = pd.DataFrame({
    "Numero de apoyo": ["P01", "P02", "P03", "P04"]
})
o_postes = tabla["Numero de apoyo"]
l_postes = pd.Series([
    "P01",
    "P02",
    "P02",  # derivación
    "P03",
    "P04"
])
angulo_b = pd.Series([
    0,   # P01
    0,   # P02 (vano principal)
    11,   # P02 (derivación)
    45,   # P03
    0     # P04 (alineado)
])
f_viento_at = [
    pd.Series([
        20,  # P01
        1,  # P02 principal
        1,  # P02 derivación
        20,  # P03
        10    # P04
    ])
]
f_viento_ad = [
    pd.Series([
        20,  # P01
        1,  # P02 principal
        0,  # P02 derivación
        30,  # P03
        0    # P04
    ])
]
tiro_at = [
    pd.Series([
        40,  # P01
        50,  # P02 principal
        50,   # P02 derivación
        30,  # P03
        20   # P04
    ])
]
tiro_ad = [
    pd.Series([
        40,  # P01
        0,  # P02 principal
        0,   # P02 derivación
        30,  # P03
        0   # P04
    ])
] 


tabla = calcular_ftvc_flmc(tabla,o_postes,l_postes,angulo_b,f_viento_at,f_viento_ad,tiro_at,tiro_ad)

tabla["F_check"] = np.sqrt(tabla["FTVC"]**2 + tabla["FLMC"]**2)

print(tabla)  """

def calcular_ftve(
    mec,
    zona_viento,              # "A", "B" o "C"
    area,                     # "Rural" o "Urbana"
    tabla_B2_4,               # DataFrame
    Sxe,                      # área frontal del reconectador (m²)
    postes_reco,              # pd.Series con nombres de postes
    altura_reconectador=None, # pd.Series opcional (m)
    col_poste="Numero de apoyo",
    col_salida="FTVE"
):
    """
    Calcula la Fuerza Transversal por Viento sobre Equipo (FTVE)
    y agrega la columna correspondiente al dataframe mec.

    FTVE = q0 * Cxe * Gt * Sxe
    """

    # ------------------------------------------------------------
    # Constante
    # ------------------------------------------------------------
    Cxe = 2.0

    # ------------------------------------------------------------
    # Inicializar columna
    # ------------------------------------------------------------
    mec[col_salida] = 0.0

    # ------------------------------------------------------------
    # Obtener q0 desde Tabla B2.4
    # ------------------------------------------------------------
    fila_q0 = tabla_B2_4[
        (tabla_B2_4["Area"] == area) &
        (tabla_B2_4["Zona"] == zona_viento)
    ]

    if fila_q0.empty:
        raise ValueError(
            f"No se encontró q0 para Zona={zona_viento}, Area={area}"
        )

    q0 = fila_q0.iloc[0]["q0 (daN / (m ^ 2))"]


    # ------------------------------------------------------------
    # Altura del reconectador
    # ------------------------------------------------------------
    if altura_reconectador is None:
        he = pd.Series(
            5.0,
            index=postes_reco.index
        )
    else:
        he = altura_reconectador.astype(float)

    # ------------------------------------------------------------
    # Calcular Gt según zona y área
    # ------------------------------------------------------------
    def calcular_gt(h):
        if area == "Rural" and zona_viento == "A":
            return -0.0002 * h**2 + 0.0232 * h + 1.4661

        if area == "Rural" and zona_viento in ["B", "C"]:
            return -0.0002 * h**2 + 0.0274 * h + 1.6820

        if area == "Urbana":
            return -0.0002 * h**2 + 0.0384 * h + 2.9284

        raise ValueError("Combinación Zona/Área no válida")

    Gt = he.apply(calcular_gt)

    # ------------------------------------------------------------
    # Calcular FTVE solo en postes con reconectador
    # ------------------------------------------------------------
    for idx, poste in postes_reco.items():

        mask = mec[col_poste] == poste

        if not mask.any():
            continue

        ftve = q0 * Cxe * Gt.loc[idx] * Sxe.loc[idx]
        mec.loc[mask, col_salida] = ftve
        

    return mec


def calcular_flee(
    mec,
    postes_reco,
    LE=None,   # distancia centro reconectador – poste (m)
    HE=None,   # altura del reconectador (m)
    PE=None,   # peso del equipo (daN)
    col_poste="Numero de apoyo",
    col_flee="FLEE"
):
    """
    Calcula la Fuerza Longitudinal Equivalente por Excentricidad del peso
    del equipo (FLEE).

    Fórmula:
        FLEE = HE * LE * PE

    Condiciones:
    - Solo se calcula para postes con reconectador
    - Postes sin reconectador → FLEE = 0
    - Valores por defecto (solo para postes con reconectador):
        * LE = 0.75 m
        * HE = 5.0 m
        * PE = 600 kg
    """

    # Inicializar columna
    mec[col_flee] = 0.0

    # Valores por defecto
    LE_def = 0.75
    HE_def = 5.0
    PE_def = 600.0  # daN (solo postes con reconectador)

    for idx, row in mec.iterrows():

        poste = row[col_poste]

        if poste not in postes_reco.values:
            continue

        # LE
        if LE is not None and poste in LE.index:
            le = LE.loc[poste]
        else:
            le = LE_def

        # HE
        if HE is not None and poste in HE.index:
            he = HE.loc[poste]
        else:
            he = HE_def

        # PE
        if PE is not None and poste in PE.index:
            pe = PE.loc[poste]
        else:
            pe = PE_def

        # Cálculo FLEE
        mec.at[idx, col_flee] =  le * pe / he

    return mec



def calcular_fve(
    mec,
    postes_reco,
    PE=None,                 # peso del equipo por poste (daN)
    col_poste="Numero de apoyo",
    col_fve="Fve"
):
    """
    Calcula la Fuerza Vertical por Equipos (FVE).

    Definición:
    - FVE = PE

    Condiciones:
    - Solo se aplica a postes con reconectador
    - Postes sin reconectador → FVE = 0
    - Valor por defecto (solo postes con reconectador):
        * PE = 600 daN
    """

    # Inicializar columna
    mec[col_fve] = 0.0

    PE_def = 600.0  # daN

    for idx, row in mec.iterrows():

        poste = row[col_poste]

        if poste not in postes_reco.values:
            continue

        # Peso del equipo
        if PE is not None and poste in PE.index:
            pe = PE.loc[poste]
        else:
            pe = PE_def

        mec.at[idx, col_fve] = pe

    return mec

def agregar_columna_suma(
    df,
    postes_df,      # Serie del dataframe SIN repetición
    postes_rep,     # Serie CON repetición
    valores,        # Serie de valores a sumar
    nombre_columna
):
    """
    Crea una nueva columna en df con la suma de valores por poste.
    Respeta el orden del dataframe original.
    """

    # 1) Agrupar suma por poste (serie con índice = nombre del poste)
    suma_por_poste = valores.groupby(postes_rep).sum()

    # 2) Mapear al dataframe SIN perder el orden
    df[nombre_columna] = postes_df.map(suma_por_poste).fillna(0)

    return df

import numpy as np
import pandas as pd
import re

def calcular_FTEC(
    mec,
    armados,            # Serie SIN repetición
    Fvc,                # Serie SIN repetición
    altura_postes,      # Serie SIN repetición
    d_cruceta=1.12,     # float o Serie
    nombre_columna="FTEC"
):
    """
    Calcula la Fuerza por Excentricidad del peso de los conductores (FTEC).

    Solo aplica para postes con armado tipo bandera.
    """

    # ---------------------------------------------------------
    # Preparar d_cruceta
    # ---------------------------------------------------------
    if isinstance(d_cruceta, pd.Series):
        d = d_cruceta.reindex(mec.index).astype(float)
    else:
        d = pd.Series(d_cruceta, index=mec.index, dtype=float)

    # ---------------------------------------------------------
    # Identificación de armados tipo bandera
    # patrón: 1##-#
    # ---------------------------------------------------------
    def es_bandera(codigo):
        if pd.isna(codigo):
            return False
        m = re.search(r'(\d{3}-\d)$', str(codigo))
        if not m:
            return False
        return m.group(1)[0] == "1"

    bandera = armados.apply(es_bandera)

    # ---------------------------------------------------------
    # Inicializar columna
    # ---------------------------------------------------------
    mec[nombre_columna] = 0.0

    # ---------------------------------------------------------
    # Cálculo FTEC solo en postes bandera
    # ---------------------------------------------------------
    mask = bandera.values

    mec.loc[mask, nombre_columna] = (
        Fvc.loc[mask].astype(float)
        * d.loc[mask]
        / altura_postes.loc[mask].astype(float)
    )

    return mec



def calcular_EVU(
    mec,
    postes,
    altura_postes,
    carga_rotura_poste,
    tabla_capacidad,
    Hn=None,
    nombre_columna="E.V.U."
):
    """
    Calcula el Esfuerzo Vertical Último (E.V.U) por poste,
    tomando de la tabla el valor MENOR más cercano de altura
    y carga de rotura.
    """

    mec[nombre_columna] = 0.0

    # Columnas de referencia Hn
    hn_offsets = {
        "hN": 0.0,
        "hN_0_4m": 0.4,
        "hN_0_8m": 0.8,
        "hN_3_3m": 3.3
    }

    for idx in postes.index:

        poste = postes.loc[idx]
        h_poste = float(altura_postes.loc[idx])
        carga = float(carga_rotura_poste.loc[idx])

        # ---------------- Hn efectivo ----------------
        if Hn is None:
            hn_val = h_poste - 2.0
        else:
            hn_val = float(Hn.loc[idx])

        # ------------------------------------------------
        # 1) Selección de altura menor más cercana
        # ------------------------------------------------
        alturas_validas = tabla_capacidad[
            tabla_capacidad["altura_m"] <= h_poste
        ]

        if alturas_validas.empty:
            mec.loc[mec[postes.name] == poste, nombre_columna] = np.nan
            continue

        altura_sel = alturas_validas["altura_m"].max()

        tabla_altura = alturas_validas[
            alturas_validas["altura_m"] == altura_sel
        ]

        # ------------------------------------------------
        # 2) Selección de carga menor más cercana
        # ------------------------------------------------
        cargas_validas = tabla_altura[
            tabla_altura["carga_flexion_daN"] <= carga
        ]

        if cargas_validas.empty:
            mec.loc[mec[postes.name] == poste, nombre_columna] = np.nan
            continue

        carga_sel = cargas_validas["carga_flexion_daN"].max()

        fila = cargas_validas[
            cargas_validas["carga_flexion_daN"] == carga_sel
        ].iloc[0]

        # ------------------------------------------------
        # 3) Selección de columna Hn más cercana
        # ------------------------------------------------
        diferencias = {
            col: abs(hn_val - (altura_sel - off))
            for col, off in hn_offsets.items()
        }

        col_sel = min(diferencias, key=diferencias.get)

        EVU = fila[col_sel]

        mec.loc[mec[postes.name] == poste, nombre_columna] = EVU

    return mec



def calcular_FI(
    mec,
    postes_orden,
    postes_export,
    tipo_poste_orden,
    df_tiros,
    nombre_tiro_ad,
    nombre_tiro_at,
    nombre_columna="Fl"
):
    """
    Calcula el Esfuerzo Horizontal Longitudinal por 50% de desequilibrio de tracciones.

    - Aplica solo a postes tipo ANC o FL
    - Busca el máximo absoluto entre una cantidad indefinida de columnas
      identificadas por nombres en un MultiIndex
    - Considera todas las repeticiones del poste
    """

    # Inicialización conservadora
    mec[nombre_columna] = 0.0

    tipos_validos = {"ANC", "FL"}

    # Identificación de columnas relevantes en el MultiIndex
    cols_ad = [c for c in df_tiros.columns if c[1] == nombre_tiro_ad]
    cols_at = [c for c in df_tiros.columns if c[1] == nombre_tiro_at]

    cols_tiros = cols_ad + cols_at

    if len(cols_tiros) == 0:
        return mec

    # Iteración por poste final (sin repeticiones)
    for poste in postes_orden:

        # Tipo de poste (alineado con postes_orden)
        tipo = tipo_poste_orden.loc[
            tipo_poste_orden.index[postes_orden == poste][0]
        ]

        if tipo not in tipos_validos:
            continue

        # Filas exportadas correspondientes a este poste
        mask = postes_export == poste

        if not mask.any():
            continue

        # Subconjunto de tiros:
        # filas → repeticiones
        # columnas → todos los tiros adelante y atrás
        tiros_poste = df_tiros.loc[mask, cols_tiros]

        # Máximo absoluto global
        t_max = tiros_poste.abs().to_numpy().max()

        if pd.isna(t_max):
            continue

        FI = 0.5 * t_max

        # Escritura final por poste
        mec.loc[mec[postes_orden.name] == poste, nombre_columna] = FI

    return mec

def calcular_mr(
    mec,
    postes_orden,
    postes_export,
    tipo_poste_orden,
    df_tiros,
    nombre_tiro_ad,
    nombre_tiro_at,
    brazo,
    nombre_columna="Mr"
):
    """
    Calcula el momento torsor mr a partir del máximo desequilibrio de tensiones.

    - Aplica solo a postes tipo ANC o FL
    - El brazo puede ser:
        * una Series numérica (brazo por poste)
        * una Series de armados (string), de la cual se deduce el brazo
    """

    # Inicialización conservadora
    mec[nombre_columna] = 0.0

    tipos_validos = {"ANC", "FL"}

    # Identificación de columnas de tensiones en el MultiIndex
    cols_ad = [c for c in df_tiros.columns if c[1] == nombre_tiro_ad]
    cols_at = [c for c in df_tiros.columns if c[1] == nombre_tiro_at]
    cols_tiros = cols_ad + cols_at

    if len(cols_tiros) == 0:
        return mec

    # Identificar si el brazo es numérico por poste
    brazo_es_numerico = brazo.map(
        lambda x: isinstance(x, (int, float)) and not pd.isna(x)
    )

    # Iteración por poste final
    for poste in postes_orden:

        # Tipo de poste
        tipo = tipo_poste_orden.loc[
            tipo_poste_orden.index[postes_orden == poste][0]
        ]

        if tipo not in tipos_validos:
            continue

        # Filas exportadas correspondientes a este poste
        mask = postes_export == poste

        if not mask.any():
            continue

        # Subconjunto de tensiones del poste
        tiros_poste = df_tiros.loc[mask, cols_tiros]

        # Máximo absoluto global
        t_max = tiros_poste.abs().to_numpy().max()

        if pd.isna(t_max):
            continue

        # Índice del poste en la serie ordenada
        idx_poste = postes_orden.index[postes_orden == poste][0]

        # Obtención del brazo
        if brazo_es_numerico.loc[idx_poste]:
            brazo_p = brazo.loc[idx_poste]
        else:
            armado = str(brazo.loc[idx_poste]).strip()

            primer_digito = None
            for ch in armado:
                if ch.isdigit():
                    primer_digito = ch
                    break

            if primer_digito in {"6", "7"}:
                brazo_p = 0.52
            else:
                brazo_p = 1.12

        # Momento torsor
        mr = t_max * brazo_p

        # Escritura final por poste
        mec.loc[mec[postes_orden.name] == poste, nombre_columna] = mr

    return mec

def calcular_Mut(
    mec,
    postes,
    altura_postes,
    carga_rotura_poste,
    tabla_capacidad,
    nombre_columna="Mut"
):
    """
    Calcula el Momento Último de Torsión (Mut) por poste.

    Criterio:
    1) Se selecciona la mayor altura de tabla <= altura del poste.
    2) Para esa altura:
       - Si existe alguna carga con diferencia relativa <= 5% respecto
         a la carga del poste, se toma la más cercana (aunque sea mayor).
       - Si no, se toma la mayor carga menor inmediata.
    3) Si la combinación no existe, se itera hacia cargas menores
       hasta encontrar una fila válida.
    """

    mec[nombre_columna] = 0.0

    for idx in postes.index:

        poste = postes.loc[idx]
        h_poste = float(altura_postes.loc[idx])
        carga_poste = float(carga_rotura_poste.loc[idx])

        # ------------------------------------------------
        # 1) Altura menor o igual más cercana
        # ------------------------------------------------
        alturas_validas = tabla_capacidad[
            tabla_capacidad["altura_m"] <= h_poste
        ]

        if alturas_validas.empty:
            mec.loc[mec[postes.name] == poste, nombre_columna] = np.nan
            continue

        altura_sel = alturas_validas["altura_m"].max()

        tabla_altura = alturas_validas[
            alturas_validas["altura_m"] == altura_sel
        ]

        # ------------------------------------------------
        # 2) Selección de carga con criterio 5%
        # ------------------------------------------------
        cargas_tabla = tabla_altura["carga_flexion_daN"].unique()

        if len(cargas_tabla) == 0:
            mec.loc[mec[postes.name] == poste, nombre_columna] = np.nan
            continue

        # Diferencias relativas
        diffs = {
            c: abs(c - carga_poste) / carga_poste
            for c in cargas_tabla
        }

        # Cargas dentro del 5%
        cargas_5 = [c for c, d in diffs.items() if d <= 0.05]

        if cargas_5:
            # Más cercana, aunque sea mayor
            carga_sel = min(cargas_5, key=lambda c: abs(c - carga_poste))
        else:
            # Piso inmediato
            cargas_menores = [c for c in cargas_tabla if c <= carga_poste]
            if not cargas_menores:
                mec.loc[mec[postes.name] == poste, nombre_columna] = np.nan
                continue
            carga_sel = max(cargas_menores)

        # ------------------------------------------------
        # 3) Fallback si la combinación no existe
        # ------------------------------------------------
        cargas_ordenadas = sorted(cargas_tabla, reverse=True)

        fila = None
        for c in cargas_ordenadas:
            if c > carga_sel:
                continue
            candidata = tabla_altura[
                tabla_altura["carga_flexion_daN"] == c
            ]
            if not candidata.empty:
                fila = candidata.iloc[0]
                break

        if fila is None:
            mec.loc[mec[postes.name] == poste, nombre_columna] = np.nan
            continue

        # ------------------------------------------------
        # 4) Momento último de torsión
        # ------------------------------------------------
        Mut = fila["momento_torsion_daN_m"]

        mec.loc[mec[postes.name] == poste, nombre_columna] = Mut

    return mec


def crear_fase_mensajero(
    carac_postes,
    postes_orden,
    postes_export,
    texto_export,
    col_fase="Fase",
    col_mensajero="Mensajero"
):
    """
    Crea las columnas 'Fase' y 'Mensajero' a partir de un texto exportado.

    Reglas:
    - Si el texto cumple exactamente el formato 'FASE / MENSAJERO':
        Fase = FASE
        Mensajero = MENSAJERO
    - En cualquier otro caso:
        Fase = texto completo
        Mensajero = NaN
    - Datos provenientes de exportación (pueden estar desordenados y repetidos)
    - Para postes repetidos, se toma el PRIMER valor válido encontrado
    - Valor válido: no NaN, no "", no "-", no "0"
    - Si el valor seleccionado (n) es inválido, se hereda el valor n-1
      (aplica tanto para Fase como para Mensajero)
    """

    # --------------------------------------------------
    # Inicialización
    # --------------------------------------------------
    carac_postes[col_fase] = np.nan
    carac_postes[col_mensajero] = np.nan

    # --------------------------------------------------
    # Funciones auxiliares
    # --------------------------------------------------
    def es_valido(v):
        if pd.isna(v):
            return False
        v_str = str(v).strip()
        return v_str not in {"", "-", "0"}

    def obtener_fase_mensajero(v):
        """
        Aplica la regla de parsing:
        - Solo si hay exactamente un ' / ' se separa
        - En otro caso, todo es fase
        """
        if not es_valido(v):
            return np.nan, np.nan

        v_str = str(v).strip()

        if v_str.count(" / ") == 1:
            f, m = v_str.split(" / ")
            return f.strip(), m.strip()

        return v_str, np.nan

    def seleccionar_valor_fase_mensajero(idxs):
        """
        Devuelve (fase, mensajero) aplicando:
        - primer valor válido
        - herencia n-1 si corresponde
        """
        for i in idxs:
            f, m = obtener_fase_mensajero(texto_export.iloc[i])

            if es_valido(f) or es_valido(m):
                return f, m

            if i > 0:
                f_ant, m_ant = obtener_fase_mensajero(texto_export.iloc[i - 1])
                if es_valido(f_ant) or es_valido(m_ant):
                    return f_ant, m_ant

        return np.nan, np.nan

    # --------------------------------------------------
    # Iteración por poste final (ordenado)
    # --------------------------------------------------
    for poste in postes_orden:

        idxs = np.where(postes_export.values == poste)[0]

        if len(idxs) == 0:
            continue

        fase_sel, mensajero_sel = seleccionar_valor_fase_mensajero(idxs)

        carac_postes.loc[
            carac_postes[postes_orden.name] == poste, col_fase
        ] = fase_sel

        carac_postes.loc[
            carac_postes[postes_orden.name] == poste, col_mensajero
        ] = mensajero_sel

    return carac_postes



def determinar_tense(
    carac_postes,
    postes_orden,
    postes_export,
    armado_export,
    tiro_adelante_export,
    tiro_atras_export,
    tabla_tiro_rotura,
    cable_export,
    nombre_columna="Tense"
):
    """
    Determina el tipo de tense por poste.

    Reglas:
    - Si el primer dígito numérico del armado es 6 o 7 → "Normal"
    - En otro caso:
        * Se obtiene el tiro máximo (adelante / atrás) considerando todas
          las repeticiones del poste
        * Se obtiene la carga de rotura máxima de los cables asociados al poste
        * Si tiro_max > 0.08 * carga_rotura → "Normal"
          else → "Reducido"
    """

    # ------------------------------------------------------------
    # Inicialización
    # ------------------------------------------------------------
    carac_postes[nombre_columna] = np.nan

    # ------------------------------------------------------------
    # Consolidación de tiros (listas de Series → Series única)
    # ------------------------------------------------------------
    if isinstance(tiro_adelante_export, list):
        tiro_adelante = sumar_lista_series(tiro_adelante_export)
    else:
        tiro_adelante = tiro_adelante_export

    if isinstance(tiro_atras_export, list):
        tiro_atras = sumar_lista_series(tiro_atras_export)
    else:
        tiro_atras = tiro_atras_export

    postes_exp = postes_export.values

    # ------------------------------------------------------------
    # Función auxiliar: primer dígito numérico del armado
    # ------------------------------------------------------------
    def primer_digito_numerico(txt):
        if not isinstance(txt, str):
            return None
        for ch in txt:
            if ch.isdigit():
                return int(ch)
        return None

    # ------------------------------------------------------------
    # Función auxiliar: carga de rotura del cable
    # ------------------------------------------------------------
    def carga_rotura_cable(nombre_cable):
        if not isinstance(nombre_cable, str):
            return None

        for _, fila in tabla_tiro_rotura.iterrows():
            conductor = fila["Conductor"]
            if isinstance(conductor, str) and nombre_cable in conductor:
                return fila["Carga de Rotura (daN)"]

        return None

    # ------------------------------------------------------------
    # Iteración por poste final (ordenado)
    # ------------------------------------------------------------
    for idx in postes_orden.index:

        poste = postes_orden.loc[idx]

        # Repeticiones del poste en exportación
        idxs = np.where(postes_exp == poste)[0]

        if len(idxs) == 0:
            continue

        # --------------------------------------------------------
        # 1) Evaluación directa por armado
        # --------------------------------------------------------
        armado = armado_export.loc[idx]
        dig = primer_digito_numerico(armado)

        if dig in (6, 7):
            carac_postes.loc[
                carac_postes[postes_orden.name] == poste,
                nombre_columna
            ] = "Normal"
            continue

        # --------------------------------------------------------
        # 2) Tiro máximo (adelante / atrás) del poste
        # --------------------------------------------------------
        tiros = []

        for i in idxs:
            if not pd.isna(tiro_adelante.iloc[i]):
                tiros.append(abs(tiro_adelante.iloc[i]))
            if not pd.isna(tiro_atras.iloc[i]):
                tiros.append(abs(tiro_atras.iloc[i]))

        if not tiros:
            continue

        tiro_max = max(tiros)

        # --------------------------------------------------------
        # 3) Carga de rotura máxima de los cables del poste
        # --------------------------------------------------------
        cargas = []

        for i in idxs:
            carga = carga_rotura_cable(cable_export.iloc[i])
            if carga is not None:
                cargas.append(carga)

        if not cargas:
            continue

        carga_rotura_max = max(cargas)

        # --------------------------------------------------------
        # 4) Comparación 8%
        # --------------------------------------------------------
        if tiro_max > 0.08 * carga_rotura_max:
            tense = "Normal"
        else:
            tense = "Reducido"

        # Escritura final por poste
        carac_postes.loc[
            carac_postes[postes_orden.name] == poste,
            nombre_columna
        ] = tense

    return carac_postes



def calcular_vanos_adelante_atras(
    carac_postes,
    postes_orden,
    postes_export,
    numero_estructura_export,
    vano_adelante_export,
    col_vano_post="Vano posterior",
    col_vano_ant="Vano anterior"
):
    """
    Determina el vano anterior y vano posterior por poste, considerando
    rutas, posiciones (inicial / intermedia / final) y repeticiones.
    """

    carac_postes[col_vano_post] = 0.0
    carac_postes[col_vano_ant] = 0.0

    ne = numero_estructura_export.values
    vanos = vano_adelante_export.values
    postes_exp = postes_export.values

    n = len(ne)

    # ------------------------------------------------------------
    # 1) Clasificar cada fila de la exportación
    # ------------------------------------------------------------
    tipo_posicion = []

    for i in range(n):
        if ne[i] == 0:
            tipo_posicion.append("inicio")
        else:
            # final si el siguiente es 0 o si es el último registro
            if i == n - 1 or ne[i + 1] == 0:
                tipo_posicion.append("final")
            else:
                tipo_posicion.append("intermedio")

    tipo_posicion = np.array(tipo_posicion)

    # ------------------------------------------------------------
    # 2) Funciones auxiliares
    # ------------------------------------------------------------
    def primera_posicion(indices, tipo):
        for i in indices:
            if tipo_posicion[i] == tipo:
                return i
        return None

    # ------------------------------------------------------------
    # 3) Iteración por poste final ordenado
    # ------------------------------------------------------------
    for poste in postes_orden:

        idxs = np.where(postes_exp == poste)[0]

        if len(idxs) == 0:
            continue

        # Clasificar repeticiones
        idx_inicio = [i for i in idxs if tipo_posicion[i] == "inicio"]
        idx_inter = [i for i in idxs if tipo_posicion[i] == "intermedio"]
        idx_final  = [i for i in idxs if tipo_posicion[i] == "final"]

        vano_ant = 0.0
        vano_post = 0.0

        # --------------------------------------------------------
        # CASO 1: hay al menos una aparición intermedia
        # --------------------------------------------------------
        if idx_inter:
            i = idx_inter[0]
            vano_post = vanos[i]
            vano_ant = vanos[i - 1] if i > 0 else 0.0

        # --------------------------------------------------------
        # CASO 2: poste inicial
        # --------------------------------------------------------
        elif idx_inicio:
            # 2.1 hay repetición final
            if idx_final:
                i_ini = idx_inicio[0]
                i_fin = idx_final[0]
                vano_post = vanos[i_ini]
                vano_ant = vanos[i_fin - 1] if i_fin > 0 else 0.0

            # 2.2 solo repeticiones iniciales
            elif len(idx_inicio) >= 2:
                vano_post = vanos[idx_inicio[0]]
                vano_ant = vanos[idx_inicio[1]]

            # 2.3 único inicial
            else:
                i = idx_inicio[0]
                vano_post = vanos[i]
                vano_ant = 0.0

        # --------------------------------------------------------
        # CASO 3: poste final
        # --------------------------------------------------------
        elif idx_final:
            # 3.1 solo finales
            if len(idx_final) >= 2:
                i1 = idx_final[0]
                i2 = idx_final[1]
                vano_ant = vanos[i1 - 1] if i1 > 0 else 0.0
                vano_post = vanos[i2 - 1] if i2 > 0 else 0.0

            # 3.2 único final
            else:
                i = idx_final[0]
                vano_ant = vanos[i - 1] if i > 0 else 0.0
                vano_post = 0.0

        # --------------------------------------------------------
        # Escritura en carac_postes
        # --------------------------------------------------------
        carac_postes.loc[
            carac_postes[postes_orden.name] == poste, col_vano_ant
        ] = vano_ant

        carac_postes.loc[
            carac_postes[postes_orden.name] == poste, col_vano_post
        ] = vano_post

    return carac_postes

def identificar_retenida(
    carac_postes,
    postes_orden,
    postes_export,
    retenidas_export,
    armado_orden,
    col_bisectora="Bisectora",
    col_90="Conjunto a 90º"
):
    """
    Determina si un poste tiene retenida bisectora o a 90°.

    Regla correcta del armado:
    - Se ignora el último '-#'
    - Se toma el bloque inmediatamente anterior
    - Se extrae el ÚLTIMO dígito numérico de ese bloque
    - Si ese dígito es 5 → Conjunto a 90°
    - En cualquier otro caso → Bisectora
    """

    # Inicialización
    carac_postes[col_bisectora] = np.nan
    carac_postes[col_90] = np.nan

    def es_valido(v):
        if pd.isna(v):
            return False
        v_str = str(v).strip()
        return v_str not in {"", "-", "0"}

    def clasificar_armado(armado):
        if pd.isna(armado):
            return None

        s = str(armado).strip()

        if "-" not in s:
            return None

        # Quitar el último '-#'
        bloque = s.rsplit("-", 1)[0]

        # Buscar el ÚLTIMO dígito numérico del bloque
        for c in reversed(bloque):
            if c.isdigit():
                return "90" if c == "5" else "B"

        return None

    # --------------------------------------------------
    # Iteración por poste final
    # --------------------------------------------------
    for idx, poste in postes_orden.items():

        mask = postes_export == poste

        if not mask.any():
            continue

        # ¿Existe al menos una retenida válida?
        if not any(es_valido(v) for v in retenidas_export.loc[mask]):
            continue

        armado = armado_orden.loc[idx]
        tipo = clasificar_armado(armado)

        if tipo == "90":
            carac_postes.loc[
                carac_postes[postes_orden.name] == poste, col_90
            ] = "X"
        elif tipo == "B":
            carac_postes.loc[
                carac_postes[postes_orden.name] == poste, col_bisectora
            ] = "X"

    return carac_postes


def calcular_tiro_maximo(
    carac_postes,
    postes_orden,
    postes_export,
    tiro_adelante_export,
    tiro_atras_export,
    nombre_columna="Tiro_max"
):
    """
    Calcula por cada poste el tiro máximo entre tiro adelante y tiro atrás,
    considerando todas las repeticiones provenientes de la exportación.

    - Se toma el máximo absoluto.
    - Si el poste se repite, se evalúan todas sus repeticiones.
    - Escritura final ordenada por postes_orden.
    """

    # Inicialización conservadora
    carac_postes[nombre_columna] = np.nan

    postes_exp = postes_export.values
    ta = tiro_adelante_export.values
    td = tiro_atras_export.values

    # ------------------------------------------------------------
    # Iteración por poste final (ordenado)
    # ------------------------------------------------------------
    for idx in postes_orden.index:

        poste = postes_orden.loc[idx]

        # Repeticiones en exportación
        idxs = np.where(postes_exp == poste)[0]

        if len(idxs) == 0:
            continue

        # Tiros de todas las repeticiones
        tiros = []

        for i in idxs:
            if not pd.isna(ta[i]):
                tiros.append(abs(ta[i]))
            if not pd.isna(td[i]):
                tiros.append(abs(td[i]))

        if not tiros:
            continue

        tiro_max = max(tiros)

        # Escritura final por poste
        carac_postes.loc[
            carac_postes[postes_orden.name] == poste, nombre_columna
        ] = tiro_max

    return carac_postes



def fuerza_residual_retenidas(
    carac_postes,
    postes_orden,
    postes_export,
    retenidas_export,
    flmc,
    altura_postes,
    capacidad_poste,
    tablas_coeficientes,
    tipo_retenida,
    tipo_poste,
    tabla_cables_acero,
    calibre_retenida="3/8",
    altura_retenidas=None,
    col_fres="Fuerza Residual Fres (daN)",
    col_fvert="Fuerza vertical por retenida Fvert (daN)",
    col_trac="Tracción total cable ret.(daN)",
    col_pret="Pretensionado de la Retenida (daN)",
    col_rot="Carga rotura cable ret. (daN)",
):
    """
    Calcula fuerzas asociadas a retenidas:
    - Fuerza residual (Fres)
    - Fuerza vertical (Fvert)
    - Tracción en la retenida (Tracr)
    - Pretensionado de la retenida (Pretr)
    - Carga de rotura del cable de retenida (Rotr)
    """

    # --------------------------------------------------
    # Inicialización columnas
    # --------------------------------------------------
    for c in [col_fres, col_fvert, col_trac, col_pret, col_rot]:
        carac_postes[c] = np.nan

    if altura_retenidas is None:
        altura_retenidas = altura_postes

    # --------------------------------------------------
    # Funciones auxiliares
    # --------------------------------------------------
    def es_valido(v):
        if pd.isna(v):
            return False
        v = str(v).strip()
        return v not in {"", "-", "0"}

    def obtener_ru(calibre):
        fila = tabla_cables_acero[
            tabla_cables_acero["Denominación"].astype(str).str.contains(str(calibre))
        ]
        if fila.empty:
            return np.nan
        return float(fila.iloc[0]["Carga de Rotura (daN)"])

    def seleccionar_tabla(tipo_ret):
        """
        Selección de tabla:
        - Índice impar/par según calibre
        - Con o sin beta según tipo
        """
        if calibre_retenida == "3/8":
            idxs = [1, 3]
        else:
            idxs = [0, 2]

        for i in idxs:
            df = tablas_coeficientes[i]
            if tipo_ret == "90" and "β (°)" in df.index.names:
                return df
            if tipo_ret == "Bisectora" and "β (°)" not in df.index.names:
                return df

        return None

    def seleccionar_col_delta(delta):
        if delta <= 1.2:
            return "≤1.2 m"
        if delta <= 2.2:
            return "1.2 < L ≤ 2.2 m"
        return ">3.0 m"

    def valor_cercano(valor, index_vals):
        index_vals = np.array(index_vals, dtype=float)
        dif = np.abs(index_vals - valor) / valor
        if (dif <= 0.05).any():
            return index_vals[dif.argmin()]
        return index_vals[index_vals <= valor].max()

    # --------------------------------------------------
    # Iteración por poste ordenado
    # --------------------------------------------------
    for i, poste in enumerate(postes_orden):

        idxs = np.where(postes_export.values == poste)[0]

        if len(idxs) == 0:
            continue

        if not any(es_valido(retenidas_export.iloc[j]) and float(retenidas_export.iloc[j]) > 0 for j in idxs):
            continue

        # Tipo de retenida
        fila_tipo = tipo_retenida.loc[
            carac_postes[postes_orden.name] == poste
        ]

        if fila_tipo.empty:
            continue

        conj_90 = fila_tipo["Conjunto a 90º"].iloc[0]
        bisec = fila_tipo["Bisectora"].iloc[0]

        if conj_90 == "X":
            tipo_ret = "90"
        elif bisec == "X":
            tipo_ret = "Bisectora"
        else:
            continue


        # Parámetros geométricos
        Hj = altura_retenidas.iloc[i]
        Hn = altura_postes.iloc[i]
        delta = Hn - Hj
        col_L = seleccionar_col_delta(delta)

        # Tabla coeficientes
        df_coef = seleccionar_tabla(tipo_ret)
        if df_coef is None:
            continue

        carga_poste = capacidad_poste.iloc[i]
        cargas_tabla = df_coef.index.get_level_values(-1).unique()
        carga_sel = valor_cercano(carga_poste, cargas_tabla)

        if "β (°)" in df_coef.index.names:
            beta = valor_cercano(flmc.index[i], df_coef.index.get_level_values(0).unique())
            fila = df_coef.loc[(beta, carga_sel)]
        else:
            fila = df_coef.loc[carga_sel]

        A = fila[(col_L, "A")]
        B = fila[(col_L, "B")]
        C = fila[(col_L, fila.index.get_level_values(1)[-1])]

        fl = flmc.iloc[i]

        # Fuerzas
        Fres = fl * A * Hj / Hn
        Fvert = B * fl
        Tracr = 1.5 * C * fl
        Ru = obtener_ru(calibre_retenida)

        # Pretensionado
        tp = tipo_poste.iloc[i]
        if tp == "ANC":
            pretr = max(2 * abs(A * fl - carga_poste / 1.5), 0.05 * Ru)
        else:
            pretr = max(2 * abs(A * fl - carga_poste / 2.5), 0.05 * Ru)

        # Asignación
        carac_postes.loc[i, col_fres] = Fres
        carac_postes.loc[i, col_fvert] = Fvert
        carac_postes.loc[i, col_trac] = Tracr
        carac_postes.loc[i, col_pret] = pretr
        carac_postes.loc[i, col_rot] = Ru

    return carac_postes


def calcular_fuerza_residual_retenidas(
    carac_postes,
    postes_orden,
    postes_export,
    retenidas,
    flmc,
    altura_postes,
    capacidad_poste,
    tablas_coeficientes,
    tipo_retenida,
    tipo_poste,
    tabla_cables_acero,
    calibre_retenida="3/8",
    altura_retenidas=None,
    col_fres="Fuerza Residual Fres (daN)",
    col_fvert="Fuerza vertical por retenida Fvert (daN)",
    col_trac="Tracción total cable ret.(daN)",
    col_pret="Pretensionado de la Retenida (daN)",
    col_rot="Carga rotura cable ret. (daN)",
):
    """
    Calcula la fuerza residual por retenidas y agrega la columna
    'Fuerza Residual Fres (daN)' al dataframe carac_postes.
    """

    # Inicialización
    for c in [col_fres, col_fvert, col_trac, col_pret, col_rot]:
        carac_postes[c] = np.nan
        
    if altura_retenidas is None:
        altura_retenidas = altura_postes

    # Selección de tablas según calibre
    if calibre_retenida == "3/8":
        tablas_calibre = [tablas_coeficientes[0], tablas_coeficientes[2]]
    else:
        tablas_calibre = [tablas_coeficientes[1], tablas_coeficientes[3]]
    # --------------------------------------------------
    # Funciones auxiliares
    # --------------------------------------------------        
    def obtener_ru(calibre):
        fila = tabla_cables_acero[
            tabla_cables_acero["Denominación"].astype(str).str.contains(str(calibre))
        ]
        if fila.empty:
            return np.nan
        return float(fila.iloc[0]["Carga de Rotura (daN)"])
    
    def seleccionar_longitud(delta_h):
        if delta_h <= 1.2:
            return "≤1.2 m"
        elif delta_h <= 2.2:
            return "1.2 < L ≤ 2.2 m"
        else:
            return ">3.0 m"

    def valor_cercano(valores, objetivo):
        valores = np.array(sorted(valores))
        dif_rel = np.abs(valores - objetivo) / objetivo

        mask_5 = dif_rel <= 0.05
        if mask_5.any():
            return valores[mask_5][
                np.argmin(np.abs(valores[mask_5] - objetivo))
            ]

        menores = valores[valores <= objetivo]
        if len(menores) > 0:
            return menores[-1]

        return valores[0]

    # Iteración por poste (orden final)
    for i, poste in enumerate(postes_orden):

        mask = postes_export == poste

        if not mask.any():
            continue

        # 1. Verificar retenidas
        ret_vals = retenidas.loc[mask]
        ret_vals = ret_vals[ret_vals > 0]

        if ret_vals.empty:
            continue

        # 2. Alturas
        Hn = altura_postes.iloc[i]
        Hj = altura_retenidas.iloc[i]
        delta_h = Hn - Hj

        # 3. Tipo de retenida
        tipo_fila = tipo_retenida.iloc[i]

        if tipo_fila["Bisectora"] == "X":
            es_90 = False
        elif tipo_fila["Conjunto a 90º"] == "X":
            es_90 = True
        else:
            continue

        # 4. Selección de tabla
        if es_90:
            tabla = tablas_calibre[1]  # con beta
            betas = tabla.index.get_level_values("β (°)").unique()
            beta_sel = valor_cercano(betas, 90)
        else:
            tabla = tablas_calibre[0]  # sin beta

        # Carga de rotura
        carga_poste = capacidad_poste.iloc[i]

        if es_90:
            cargas_tabla = tabla.index.get_level_values("Carga (daN)").unique()
        else:
            cargas_tabla = tabla.index

        carga_sel = valor_cercano(cargas_tabla, carga_poste)

        # Longitud
        col_long = seleccionar_longitud(delta_h)

        # Coeficientes A,B,C
        if es_90:
            A = tabla.loc[(beta_sel, carga_sel), (col_long, "A")]
            B = tabla.loc[(beta_sel, carga_sel), (col_long, "B")]
            C = tabla.loc[(beta_sel, carga_sel), (col_long, "C")]
        else:
            A = tabla.loc[carga_sel, (col_long, "A")]
            B = tabla.loc[carga_sel, (col_long, "B")]
            C = tabla.loc[carga_sel, (col_long, "C")]

        # 5. Fuerza residual
        fres = flmc.iloc[i] * A * Hj / Hn
        fver = flmc.iloc[i]*B
        trac = flmc.iloc[i]*C*1.5
        Ru = obtener_ru(calibre_retenida)
        tp = tipo_poste.iloc[i]
        if tp == "ANC":
            pretr = max(2 * abs(A * flmc.iloc[i] - carga_poste / 1.5), 0.05 * Ru)
        else:
            pretr = max(2 * abs(A * flmc.iloc[i] - carga_poste / 2.5), 0.05 * Ru)

        carac_postes.loc[
            carac_postes[postes_orden.name] == poste, col_fres
        ] = fres
        carac_postes.loc[i, col_fvert] = fver
        carac_postes.loc[i, col_trac] = trac
        carac_postes.loc[i, col_pret] = pretr
        carac_postes.loc[i, col_rot] = Ru        
        

    return carac_postes



import numpy as np
import pandas as pd


def capacidad_vertical_ultima_retenida(
    carac_postes,
    postes_orden,
    altura_poste,
    altura_libre,
    carga_rotura_poste,
    tabla_evu_postes,
    altura_retenida=None,
    col_salida='Capacidad vertical del poste (daN)',
):
    """
    Calcula la capacidad vertical última de un poste con retenida
    según tabla AFINIA de capacidad vertical con retenidas.

    Reglas generales:
    - Iteración por postes_orden
    - Uso del principio del 5% para selección de valores tabulados
    - Si no hay coincidencia exacta, se toma el valor más cercano inferior
    """

    # --------------------------------------------------
    # Inicialización
    # --------------------------------------------------
    carac_postes[col_salida] = np.nan

    if altura_retenida is None:
        altura_retenida = altura_poste

    # --------------------------------------------------
    # Funciones auxiliares
    # --------------------------------------------------
    def valor_tabla_5pct(objetivo, valores):
        """
        Aplica regla:
        - ±5% → valor más cercano
        - si no, toma el mayor <= objetivo
        """
        valores = np.array(sorted(valores), dtype=float)

        dif_rel = np.abs(valores - objetivo) / objetivo
        mask_5 = dif_rel <= 0.05

        if mask_5.any():
            return valores[mask_5][np.argmin(np.abs(valores[mask_5] - objetivo))]

        menores = valores[valores <= objetivo]
        if len(menores) > 0:
            return menores[-1]

        return valores[0]

    def seleccionar_columna_hn(delta_h):
        """
        Selección de columna según delta de altura
        """
        if delta_h <= 0.4:
            return "hN"
        elif delta_h <= 0.8:
            return "hN_0_4m"
        elif delta_h <= 3.3:
            return "hN_0_8m"
        else:
            return "hN_3_3m"

    # --------------------------------------------------
    # Iteración por poste (orden final)
    # --------------------------------------------------
    for i, poste in enumerate(postes_orden):

        H_total = altura_poste.iloc[i]
        H_libre = altura_libre.iloc[i]
        H_ret = altura_retenida.iloc[i]
        carga_poste = carga_rotura_poste.iloc[i]

        if pd.isna(H_total) or pd.isna(H_libre) or pd.isna(carga_poste):
            continue

        # --------------------------------------------------
        # 1. Delta de alturas
        # --------------------------------------------------
        delta_h = H_libre - H_ret
        col_hn = seleccionar_columna_hn(delta_h)

        # --------------------------------------------------
        # 2. Selección de altura y carga en tabla
        # --------------------------------------------------
        alturas_tabla = tabla_evu_postes["altura_m"].unique()
        cargas_tabla = tabla_evu_postes["carga_flexion_daN"].unique()

        altura_sel = valor_tabla_5pct(H_total, alturas_tabla)
        carga_sel = valor_tabla_5pct(carga_poste, cargas_tabla)

        # --------------------------------------------------
        # 3. Fila final en tabla
        # --------------------------------------------------
        fila = tabla_evu_postes[
            (tabla_evu_postes["altura_m"] == altura_sel)
            & (tabla_evu_postes["carga_flexion_daN"] == carga_sel)
        ]

        if fila.empty:
            continue

        cap_vertical = float(fila.iloc[0][col_hn])

        # --------------------------------------------------
        # 4. Escritura final
        # --------------------------------------------------
        carac_postes.loc[
            carac_postes[postes_orden.name] == poste, col_salida
        ] = cap_vertical

    return carac_postes



def calcular_cs(
    Ret: pd.DataFrame,
    Mec: pd.DataFrame,
    col_salida: str = 'Fuerza Total Horiz. Resultante (daN)'
):
    """
    Calcula la combinación C.S. (2,5) y agrega la columna al dataframe Ret.

    Para cada fila:
    - Se compara FTVC vs FLMC en Mec
    - Se toma el máximo
    - Se aplica la expresión correspondiente
    """

    # ------------------------------------------------------------
    # Inicializar columna con valor por defecto
    # ------------------------------------------------------------
    Ret[col_salida] = np.nan

    # ------------------------------------------------------------
    # Comparación fila a fila
    # ------------------------------------------------------------
    max_es_FLMC = Mec["FLMC"] >= Mec["FTVC"]
    max_es_FTVC = Mec["FTVC"] > Mec["FLMC"]

    # ------------------------------------------------------------
    # Caso 1: máximo en FLMC
    # ------------------------------------------------------------
    Ret.loc[max_es_FLMC, col_salida] = np.sqrt(
        (
            Mec.loc[max_es_FLMC, "FTVC"]
            + Mec.loc[max_es_FLMC, "FTVP"]
            + Mec.loc[max_es_FLMC, "FTVE"]
            + Mec.loc[max_es_FLMC, "FTEC"]
        ) ** 2
        +
        (
            Mec.loc[max_es_FLMC, "FLEE"]
            + Ret.loc[max_es_FLMC, "Fuerza Residual Fres (daN)"]
        ) ** 2
    )

    # ------------------------------------------------------------
    # Caso 2: máximo en FTVC
    # ------------------------------------------------------------
    Ret.loc[max_es_FTVC, col_salida] = np.sqrt(
        (
            Ret.loc[max_es_FTVC, "Fuerza Residual Fres (daN)"]
            + Mec.loc[max_es_FTVC, "FTVP"]
            + Mec.loc[max_es_FTVC, "FTVE"]
            + Mec.loc[max_es_FTVC, "FTEC"]
        ) ** 2
        +
        (
            Mec.loc[max_es_FTVC, "FLMC"]
            + Mec.loc[max_es_FTVC, "FLEE"]
        ) ** 2
    )

    return Ret



def clasificar_cantones(
    postes_exportacion,     # Series con identificador del poste (solo referencia)
    tipo_poste,             # Series ("ANC", "FL", etc.)
    numero_en_ruta          # Series numérico
):
    """
    Clasifica cada poste en su(s) cantón(es) según reglas definidas.

    Retorna:
        pd.Series con:
        - int → pertenece a un solo cantón
        - list[int, int] → es fin de un cantón e inicio de otro
    """

    n = len(postes_exportacion)

    # ------------------------------------------------------------
    # Identificar inicio / fin de cantón por poste
    # ------------------------------------------------------------
    inicio = [False] * n
    fin = [False] * n

    for i in range(n):

        tipo = tipo_poste.iloc[i]
        nr = numero_en_ruta.iloc[i]
        es_ultimo = (i == n - 1) or (numero_en_ruta.iloc[i+1] == 0)

        # --- Regla 1: tipo ANC o FL ---
        if tipo in ["ANC", "FL"]:
            inicio[i] = True
            fin[i] = True

        # --- Regla 2: cambio de ruta ---
        if nr == 0:
            inicio[i] = True

        if nr != 0:
            if es_ultimo:
                fin[i] = True
            else:
                if numero_en_ruta.iloc[i + 1] == 0:
                    fin[i] = True

        # --------------------------------------------------------
        # AJUSTE CLAVE:
        # El último poste NO puede iniciar un cantón nuevo
        # --------------------------------------------------------
        if es_ultimo:
            inicio[i] = False
        
        #El primero no puede terminar un ruta
        if (numero_en_ruta.iloc[i] == 0) and not (es_ultimo):
            fin[i] = False
            
        

    # ------------------------------------------------------------
    # Asignar cantones en orden de exportación
    # ------------------------------------------------------------
    canton_actual = 0
    resultado = [None] * n
    iniciar_nuevo = True

    for i in range(n):

        if iniciar_nuevo:
            canton_actual += 1
            iniciar_nuevo = False

        if inicio[i] and fin[i]:
            # Fin e inicio simultáneo (NO ocurre en el último poste)
            resultado[i] = [canton_actual, canton_actual + 1]
            canton_actual += 1
            iniciar_nuevo = False

        elif fin[i]:
            resultado[i] = canton_actual
            iniciar_nuevo = True

        else:
            resultado[i] = canton_actual

    return pd.Series(resultado, index=postes_exportacion.index, name="Canton")



def max_canton(
    van_reg,     # DataFrame base
    cantones,    # Series con cantón o lista de cantones por poste
    lista,       # Series con valores en orden de exportación
    col          # Nombre de la columna de salida
):
    """
    Obtiene un único valor máximo por cantón usando la regla n-1
    (se excluye el último poste del cantón).

    - La salida tiene una fila por cantón, en orden creciente.
    - Si van_reg tiene menos filas que la salida, se expande y
      las columnas existentes se rellenan con NaN.
    """

    # ------------------------------------------------------------
    # Expandir relación poste–cantón
    # ------------------------------------------------------------
    pares = []

    for idx, c in cantones.items():
        if isinstance(c, list):
            for ci in c:
                pares.append((ci, idx))
        else:
            pares.append((c, idx))

    df = pd.DataFrame(pares, columns=["canton", "idx"])

    # ------------------------------------------------------------
    # Calcular máximo por cantón (regla n-1)
    # ------------------------------------------------------------
    resultados = []

    cantones_ordenados = sorted(df["canton"].unique())

    for canton in cantones_ordenados:

        idxs = df[df["canton"] == canton]["idx"].tolist()

        if len(idxs) < 2:
            resultados.append(np.nan)
            continue

        idxs_validos = idxs[:-1]  # excluir último poste del cantón

        valores = lista.loc[idxs_validos].dropna()

        if valores.empty:
            resultados.append(np.nan)
        else:
            resultados.append(valores.max())

    # ------------------------------------------------------------
    # Ajustar tamaño del DataFrame
    # ------------------------------------------------------------
    n_out = len(resultados)
    n_df = len(van_reg)

    if n_df < n_out:
        # Expandir DataFrame
        filas_extra = n_out - n_df
        df_extra = pd.DataFrame(
            np.nan,
            index=range(filas_extra),
            columns=van_reg.columns
        )
        van_reg = pd.concat([van_reg, df_extra], ignore_index=True)

    elif n_df > n_out:
        # Recortar DataFrame
        van_reg = van_reg.iloc[:n_out].copy()

    # ------------------------------------------------------------
    # Asignar columna de salida
    # ------------------------------------------------------------
    van_reg[col] = resultados

    return van_reg



def min_canton(
    van_reg,
    cantones,          # Series en orden de exportación (int o list)
    lista,             # Series en orden de exportación
    col                # nombre de la columna a crear / reemplazar
):
    """
    Obtiene el valor mínimo por cantón usando solo los valores hasta n-1.
    Si el mínimo es 0, se toma el siguiente mínimo distinto de 0.

    La salida tiene una fila por cantón.
    Ajusta el tamaño de van_reg si es necesario.
    """

    # ------------------------------------------------------------
    # Expandir relación poste – cantón
    # ------------------------------------------------------------
    registros = []

    for idx, c in cantones.items():
        if isinstance(c, list):
            for ci in c:
                registros.append((ci, idx))
        else:
            registros.append((c, idx))

    df = pd.DataFrame(registros, columns=["canton", "idx"]).sort_values("idx")

    # ------------------------------------------------------------
    # Cantones únicos en orden de aparición
    # ------------------------------------------------------------
    cantones_ordenados = []
    for c in df["canton"]:
        if c not in cantones_ordenados:
            cantones_ordenados.append(c)

    # ------------------------------------------------------------
    # Calcular mínimos por cantón (regla n-1 + exclusión de ceros)
    # ------------------------------------------------------------
    valores = []

    for c in cantones_ordenados:
        idxs = df.loc[df["canton"] == c, "idx"].tolist()

        # n-1
        if len(idxs) <= 1:
            valores.append(np.nan)
            continue

        idxs_validos = idxs[:-1]
        datos = lista.loc[idxs_validos].astype(float)

        # eliminar ceros
        datos_no_cero = datos[datos != 0]

        if datos_no_cero.empty:
            valores.append(np.nan)
        else:
            valores.append(datos_no_cero.min())

    serie_min = pd.Series(valores, name=col)

    # ------------------------------------------------------------
    # Ajustar tamaño del dataframe
    # ------------------------------------------------------------
    n_df = len(van_reg)
    n_col = len(serie_min)

    if n_df > n_col:
        van_reg = van_reg.iloc[:n_col].copy()
    elif n_df < n_col:
        filas_extra = pd.DataFrame(
            np.nan,
            index=range(n_df, n_col),
            columns=van_reg.columns
        )
        van_reg = pd.concat([van_reg, filas_extra], ignore_index=True)

    # ------------------------------------------------------------
    # Asignar columna
    # ------------------------------------------------------------
    van_reg[col] = serie_min.values

    return van_reg


def resumen_cantones(
    reg_van,
    postes,
    cantones,
    col1="Cantón",
    col2="Poste Inicial",
    col3="Poste Final",
):
    """
    Agrega / reemplaza en reg_van las columnas:
    - Cantón
    - Poste Inicial
    - Poste Final

    Cada fila representa un cantón.
    """

    # ------------------------------------------------------------
    # Expandir relación poste–cantón
    # ------------------------------------------------------------
    registros = []

    for idx, c in cantones.items():
        if isinstance(c, list):
            for ci in c:
                registros.append((ci, idx))
        else:
            registros.append((c, idx))

    df = pd.DataFrame(registros, columns=["canton", "idx"]).sort_values("idx")

    # ------------------------------------------------------------
    # Cantones únicos en orden de exportación
    # ------------------------------------------------------------
    cantones_ordenados = []
    for c in df["canton"]:
        if c not in cantones_ordenados:
            cantones_ordenados.append(c)

    # ------------------------------------------------------------
    # Poste inicial y final por cantón
    # ------------------------------------------------------------
    cant_out = []
    poste_ini = []
    poste_fin = []

    for c in cantones_ordenados:
        idxs = df.loc[df["canton"] == c, "idx"].tolist()
        cant_out.append(c)
        poste_ini.append(postes.loc[idxs[0]])
        poste_fin.append(postes.loc[idxs[-1]])

    n = len(cant_out)

    # ------------------------------------------------------------
    # Ajustar tamaño de reg_van
    # ------------------------------------------------------------
    if len(reg_van) > n:
        reg_van = reg_van.iloc[:n].copy()
    elif len(reg_van) < n:
        filas_extra = n - len(reg_van)
        reg_van = pd.concat(
            [reg_van, pd.DataFrame(np.nan, index=range(filas_extra), columns=reg_van.columns)],
            ignore_index=True
        )

    # ------------------------------------------------------------
    # Asignar columnas (sin borrar las demás)
    # ------------------------------------------------------------
    reg_van[col1] = cant_out
    reg_van[col2] = poste_ini
    reg_van[col3] = poste_fin

    return reg_van




def longitud_canton(
    reg_van,
    cantones,                 # Series en orden de exportación (int o list)
    longitudes,               # Series en orden de exportación
    col="Longitud Total del cantón"
):
    """
    Calcula la longitud total de cada cantón (hasta n-1)
    y agrega/reemplaza la columna en reg_van.
    """

    # ------------------------------------------------------------
    # Expandir relación poste – cantón
    # ------------------------------------------------------------
    registros = []

    for idx, c in cantones.items():
        if isinstance(c, list):
            for ci in c:
                registros.append((ci, idx))
        else:
            registros.append((c, idx))

    df = pd.DataFrame(registros, columns=["canton", "idx"]).sort_values("idx")

    # ------------------------------------------------------------
    # Cantones únicos en orden de aparición
    # ------------------------------------------------------------
    cantones_ordenados = []
    for c in df["canton"]:
        if c not in cantones_ordenados:
            cantones_ordenados.append(c)

    # ------------------------------------------------------------
    # Calcular longitud total por cantón (hasta n-1)
    # ------------------------------------------------------------
    valores = []

    for c in cantones_ordenados:
        idxs = df.loc[df["canton"] == c, "idx"].tolist()

        if len(idxs) <= 1:
            valores.append(0)
        else:
            valores.append(longitudes.loc[idxs[:-1]].sum())

    serie_longitud = pd.Series(valores, name=col)

    # ------------------------------------------------------------
    # Ajustar tamaño del dataframe
    # ------------------------------------------------------------
    n_df = len(reg_van)
    n_col = len(serie_longitud)

    if n_df > n_col:
        reg_van = reg_van.iloc[:n_col].copy()
    elif n_df < n_col:
        filas_extra = pd.DataFrame(
            np.nan,
            index=range(n_df, n_col),
            columns=reg_van.columns
        )
        reg_van = pd.concat([reg_van, filas_extra], ignore_index=True)

    # ------------------------------------------------------------
    # Asignar columna
    # ------------------------------------------------------------
    reg_van[col] = serie_longitud.values

    return reg_van



def agregar_vano_regulacion(
    van_reg,
    cantones,                 # Series en orden de exportación (int o list)
    vanos,                    # Series en orden de exportación
    desniveles=None,          # Series en orden de exportación o None
    usar_k_truxa=True,
    col="Vano de Regulación"
):
    """
    Calcula el vano de regulación por cantón (método de Truxá)
    y agrega/reemplaza la columna en van_reg.
    """

    # ------------------------------------------------------------
    # Expandir relación poste – cantón
    # ------------------------------------------------------------
    registros = []

    for idx, c in cantones.items():
        if isinstance(c, list):
            for ci in c:
                registros.append((ci, idx))
        else:
            registros.append((c, idx))

    df = pd.DataFrame(registros, columns=["canton", "idx"]).sort_values("idx")

    # ------------------------------------------------------------
    # Cantones únicos en orden de aparición
    # ------------------------------------------------------------
    cantones_ordenados = []
    for c in df["canton"]:
        if c not in cantones_ordenados:
            cantones_ordenados.append(c)

    # ------------------------------------------------------------
    # Calcular vano de regulación por cantón
    # ------------------------------------------------------------
    valores = []

    for c in cantones_ordenados:
        idxs = df.loc[df["canton"] == c, "idx"].tolist()

        # Solo hasta n-1
        if len(idxs) <= 1:
            valores.append(np.nan)
            continue

        idxs_validos = idxs[:-1]

        a = vanos.loc[idxs_validos].values

        if desniveles is None:
            b = None
        else:
            b = desniveles.loc[idxs_validos].values

        ar = vano_regulacion(
            vanos_m=a,
            desniveles_m=b,
            usar_k_truxa=usar_k_truxa
        )

        valores.append(ar)

    serie_vr = pd.Series(valores, name=col)

    # ------------------------------------------------------------
    # Ajustar tamaño del dataframe
    # ------------------------------------------------------------
    n_df = len(van_reg)
    n_col = len(serie_vr)

    if n_df > n_col:
        van_reg = van_reg.iloc[:n_col].copy()
    elif n_df < n_col:
        filas_extra = pd.DataFrame(
            np.nan,
            index=range(n_df, n_col),
            columns=van_reg.columns
        )
        van_reg = pd.concat([van_reg, filas_extra], ignore_index=True)

    # ------------------------------------------------------------
    # Asignar columna
    # ------------------------------------------------------------
    van_reg[col] = serie_vr.values

    return van_reg



def canton_eovanos(
    df_eovanos,          # DataFrame a modificar
    postes,              # Series de postes en orden (df_eovanos)
    cantones,            # Series de cantones en orden de exportación
    postes_exportacion,  # Series de postes en orden de exportación
    col="No de Cantón"
):
    """
    Agrega al df_eovanos la columna 'No de Cantón'.

    - Si un poste pertenece a varios cantones, retorna string "1,3,4"
    - Si pertenece a uno solo, retorna el entero
    """

    resultado = []

    for poste in postes:
        # índices donde aparece el poste en exportación
        idxs = postes_exportacion[postes_exportacion == poste].index

        cantones_poste = []

        for idx in idxs:
            c = cantones.loc[idx]
            if isinstance(c, list):
                cantones_poste.extend(c)
            else:
                cantones_poste.append(c)

        # eliminar duplicados conservando orden
        cantones_unicos = list(dict.fromkeys(cantones_poste))

        if len(cantones_unicos) == 0:
            resultado.append(None)
        elif len(cantones_unicos) == 1:
            resultado.append(cantones_unicos[0])
        else:
            resultado.append(",".join(str(x) for x in cantones_unicos))

    df_eovanos[col] = resultado
    return df_eovanos




def ajustar_df(df, fila):
    n_cols = max(df.shape[1], len(fila))
    if df.shape[1] < n_cols:
        for i in range(df.shape[1], n_cols):
            df[i] = np.nan
    if len(fila) < n_cols:
        fila = fila + [np.nan] * (n_cols - len(fila))
    return df, fila



def ajustar_df(df, fila):
    n_cols = max(df.shape[1], len(fila))
    if df.shape[1] < n_cols:
        for i in range(df.shape[1], n_cols):
            df[i] = np.nan
    if len(fila) < n_cols:
        fila = fila + [np.nan] * (n_cols - len(fila))
    return df, fila


def extraer_cantones_unicos(cantones):
    unicos = set()
    for c in cantones.dropna():
        if isinstance(c, list):
            unicos.update(c)
        else:
            unicos.add(c)
    return sorted(unicos)


def pertenece_canton(valor, k):
    if isinstance(valor, list):
        return k in valor
    return valor == k


def tablas_por_canton(
    postes_orden,
    postes_exportacion,
    cantones,
    conductor,
    vano_regulacion
):
    # -------------------------
    # Normalización
    # -------------------------
    postes_orden = pd.Series(postes_orden).reset_index(drop=True)
    postes_exportacion = pd.Series(postes_exportacion).reset_index(drop=True)
    cantones = pd.Series(cantones).reset_index(drop=True)
    conductor = pd.Series(conductor).reset_index(drop=True)
    vano_regulacion = pd.Series(vano_regulacion).reset_index(drop=True)

    cantones_unicos = extraer_cantones_unicos(cantones)

    tablas = []

    # -------------------------
    # Iterar por cantón
    # -------------------------
    for k in cantones_unicos:
        df = pd.DataFrame()

        # índices de postes que pertenecen al cantón k
        idx_exp = [
            i for i, c in cantones.items()
            if pertenece_canton(c, k)
        ]

        if len(idx_exp) == 0:
            continue

        postes_canton = postes_exportacion.iloc[idx_exp].values

        poste_ini = postes_canton[0]
        poste_fin = postes_canton[-1]

        # -------------------------
        # Conductor del cantón
        # -------------------------
        conductores = []
        for p in postes_canton:
            idx_ord = postes_orden[postes_orden == p].index
            if len(idx_ord) > 0:
                conductores.append(conductor.iloc[idx_ord[0]])

        cond_canton = conductores[0] if conductores else np.nan

        fila = ["Conductor", cond_canton]
        df, fila = ajustar_df(df, fila)
        df.loc[len(df)] = fila

        # -------------------------
        # Fila Cantón
        # -------------------------
        fila = ["Cantón", k, "Poste Inicial", poste_ini]
        df, fila = ajustar_df(df, fila)
        df.loc[len(df)] = fila

        # -------------------------
        # Vano de Regulación
        # -------------------------
        pos = int(k) - 1 if pd.notna(k) else None
        vano = (
            vano_regulacion.iloc[pos]
            if pos is not None and 0 <= pos < len(vano_regulacion)
            else np.nan
        )

        fila = ["Vano de Regulación", vano, "Poste Final", poste_fin]
        df, fila = ajustar_df(df, fila)
        df.loc[len(df)] = fila

        tablas.append(df)

    return tablas


def ajustar_df(df, fila):
    n_cols = max(df.shape[1], len(fila))
    if df.shape[1] < n_cols:
        for i in range(df.shape[1], n_cols):
            df[i] = np.nan
    if len(fila) < n_cols:
        fila = fila + [np.nan] * (n_cols - len(fila))
    return df, fila


def pertenece_canton(valor, k):
    if isinstance(valor, list):
        return k in valor
    return valor == k


def filas_canton(
    lista_tablas,
    cantones,
    vanos,
    postes_export,
    desnivel
):
    cantones = pd.Series(cantones).reset_index(drop=True)
    vanos = pd.Series(vanos).reset_index(drop=True)
    postes_export = pd.Series(postes_export).reset_index(drop=True)
    desnivel = pd.Series(desnivel).reset_index(drop=True)

    # Cantones únicos (ya sabemos que la lista_tablas está en ese orden)
    cantones_unicos = []
    for c in cantones.dropna():
        if isinstance(c, list):
            for x in c:
                if x not in cantones_unicos:
                    cantones_unicos.append(x)
        else:
            if c not in cantones_unicos:
                cantones_unicos.append(c)

    # Iterar por cada tabla / cantón
    for df, k in zip(lista_tablas, cantones_unicos):

        # índices de postes del cantón k
        idx = [
            i for i, c in cantones.items()
            if pertenece_canton(c, k)
        ]

        if len(idx) < 2:
            continue  # no hay vanos

        # Postes del cantón
        postes_canton = postes_export.iloc[idx].values

        # -------------------------
        # Vano (numeración)
        # -------------------------
        fila = ["Vano"] + list(range(1, len(postes_canton)))
        df, fila = ajustar_df(df, fila)
        df.loc[len(df)] = fila

        # -------------------------
        # Longitud (m)
        # -------------------------
        fila = ["Longitud (m)"] + vanos.iloc[idx[:-1]].tolist()
        df, fila = ajustar_df(df, fila)
        df.loc[len(df)] = fila

        # -------------------------
        # Poste inicial
        # -------------------------
        fila = ["Poste inicial"] + list(postes_canton[:-1])
        df, fila = ajustar_df(df, fila)
        df.loc[len(df)] = fila

        # -------------------------
        # Poste final
        # -------------------------
        fila = ["Poste final"] + list(postes_canton[1:])
        df, fila = ajustar_df(df, fila)
        df.loc[len(df)] = fila

        # -------------------------
        # Desnivel
        # -------------------------
        fila = ["Desnivel"] + desnivel.iloc[idx[:-1]].tolist()
        df, fila = ajustar_df(df, fila)
        df.loc[len(df)] = fila

    return lista_tablas


def tab_fle_canton(
    tabla_fle,          # DataFrame depurado y transpuesto
    cantones,           # Series en orden de exportación
    postes_export       # Series en orden de exportación
):
    """
    Construye una lista de tablas de flechado por cantón.
    Retorna:
        lista_df_normales, lista_df_secundarios
    """

    # ---------------------------------------------------------
    # 1. Determinar vanos globales por cantón
    # ---------------------------------------------------------
    canton_postes = {}
    for poste, c in zip(postes_export, cantones):
        if isinstance(c, list):
            for cc in c:
                canton_postes.setdefault(cc, []).append(poste)
        else:
            canton_postes.setdefault(c, []).append(poste)

    canton_vanos = {}
    vano_actual = 1

    for c in sorted(canton_postes.keys()):
        n_postes = len(canton_postes[c])
        n_vanos = max(n_postes - 1, 0)
        canton_vanos[c] = list(range(vano_actual, vano_actual + n_vanos))
        vano_actual += n_vanos

    # ---------------------------------------------------------
    # 2. Separar vanos normales y con S
    # ---------------------------------------------------------
    cols = tabla_fle.columns
    vanos_normales = {}
    vanos_s = {}

    for v, tipo in cols:
        if tipo != "Flecha (m)":
            continue

        if isinstance(v, str) and v.endswith("S"):
            vanos_s[int(v[:-1])] = v
        else:
            vanos_normales[int(v)] = v

    # ---------------------------------------------------------
    # 3. Función interna para construir tablas
    # ---------------------------------------------------------
    def construir_tablas(vanos_dict, sufijo=""):
        tablas = []

        for c, vanos in canton_vanos.items():
            vanos_validos = [v for v in vanos if v in vanos_dict]
            if not vanos_validos:
                continue

            df = pd.DataFrame(index=tabla_fle.index)
            df["Temperatura (°C)"] = tabla_fle.index

            # Tense: del primer vano
            v0 = vanos_validos[0]
            df["Tense (daN)"] = tabla_fle[(vanos_dict[v0], "Tiro H. (kg)")]

            # Flechas
            for v in vanos_validos:
                nombre = f"f{v}{sufijo}(m)"
                df[nombre] = tabla_fle[(vanos_dict[v], "Flecha (m)")]

            tablas.append(df.reset_index(drop=True))

        return tablas

    # ---------------------------------------------------------
    # 4. Construir tablas
    # ---------------------------------------------------------
    tablas_normales = construir_tablas(vanos_normales)
    tablas_secundarios = construir_tablas(vanos_s, sufijo="S")

    return tablas_normales, tablas_secundarios




def limpiar_flechado(tablas_flechado: pd.DataFrame) -> pd.DataFrame:
    df = tablas_flechado.copy()

    # ============================================================
    # 1. Eliminar filas por valores específicos en Temp (°C)
    # ============================================================
    eliminar_temp = ["Tiro Extremo Ini (kg)", "Tiro Extremo Fin (kg)"]
    df = df[~df["Temp (°C)\\"].isin(eliminar_temp)]

    # ============================================================
    # 2. Eliminar filas repetidas de encabezados
    # ============================================================
    def es_fila_encabezado(row):
        return str(row["N°"]).strip() == "N°"

    df = df[~df.apply(es_fila_encabezado, axis=1)]
    df = df.reset_index(drop=True)

    # ============================================================
    # 3. PROCESAMIENTO CORRECTO DE N° VANO
    #    (primarios y secundarios desacoplados)
    # ============================================================
    col_tipo = "N°"
    col_vano = "N° Vano"

    es_string = df[col_tipo].apply(lambda x: isinstance(x, str))
    es_sec = es_string & df[col_tipo].str.contains("Secundari", case=False, na=False)

    # -------------------------
    # PASADA 1 — PRIMARIOS
    # -------------------------
    ultimo_primario = None
    primarios_idx = []

    i = 0
    while i < len(df):

        if es_sec.iloc[i]:
            i += 1
            while i < len(df) and not es_string.iloc[i]:
                i += 1
            continue

        if es_string.iloc[i]:
            i += 1
            continue

        primarios_idx.append(i)
        i += 1

    for idx in primarios_idx:
        val = df.at[idx, col_vano]

        if not pd.isna(val):
            if ultimo_primario is None:
                ultimo_primario = int(val)
            else:
                ultimo_primario += 1
                df.at[idx, col_vano] = ultimo_primario
        else:
            df.at[idx, col_vano] = ultimo_primario

    base_sec = ultimo_primario

    # -------------------------
    # PASADA 2 — SECUNDARIOS
    # -------------------------
    contador = base_sec
    ultimo_s = None

    i = 0
    while i < len(df):

        if not es_sec.iloc[i]:
            i += 1
            continue

        i += 1
        while i < len(df) and not es_string.iloc[i]:

            val = df.at[i, col_vano]

            if not pd.isna(val):
                contador += 1
                ultimo_s = f"{contador}S"
                df.at[i, col_vano] = ultimo_s
            else:
                df.at[i, col_vano] = ultimo_s

            i += 1

    # -------------------------
    # ELIMINAR FILAS STRING (solo eran marcadores)
    # -------------------------
    df = df[~es_string].reset_index(drop=True)

    # ============================================================
    # 5. Reestructuración a MultiIndex
    # ============================================================
    columnas_datos = df.columns[df.columns.get_loc("Temp (°C)\\") + 1 :]

    registros = []

    for _, row in df.iterrows():
        vano_id = row["N° Vano"]
        tipo = row["Temp (°C)\\"]
        valores = row[columnas_datos].tolist()

        for col, val in zip(columnas_datos, valores):
            registros.append({
                "Vano": vano_id,
                "Tipo": tipo,
                "Col": col,
                "Valor": val
            })

    nuevo = pd.DataFrame(registros)

    tabla_final = (
        nuevo
        .pivot_table(
            index=["Vano", "Tipo"],
            columns="Col",
            values="Valor",
            aggfunc="first"
        )
        .sort_index()
    )

    return tabla_final



def clasificar_cantones_secundarios(tipo_postes: pd.Series) -> pd.Series:
    salida = []
    canton = 1

    def es_1a(v):
        return isinstance(v, str) and ("ANC" in v or "FL" in v)

    def es_1b(v):
        return isinstance(v, str) and not es_1a(v)

    def es_1c(v):
        return not isinstance(v, str)

    n = len(tipo_postes)

    # encontrar primer poste válido
    idx_primer_valido = next(
        (i for i, v in enumerate(tipo_postes) if not es_1c(v)),
        None
    )

    for i, v in enumerate(tipo_postes):

        # --- 1.c ---
        if es_1c(v):
            salida.append(np.nan)
            continue

        prev_v = tipo_postes[i - 1] if i > 0 else None
        next_v = tipo_postes[i + 1] if i < n - 1 else None

        prev_es_1c = (i == 0) or es_1c(prev_v)
        next_es_1c = (i == n - 1) or es_1c(next_v)

        # --- 1.a ---
        if es_1a(v):

            # inicio y final
            if not prev_es_1c and not next_es_1c:
                salida.append([f"{canton}S", f"{canton + 1}S"])
                canton += 1
                continue

            # inicio
            if i == idx_primer_valido or prev_es_1c:
                salida.append(f"{canton}S")
                continue

            # final
            salida.append(f"{canton}S")
            canton += 1
            continue

        # --- 1.b ---
        if es_1b(v):

            # inicio
            if i == idx_primer_valido or prev_es_1c:
                salida.append(f"{canton}S")
            else:
                salida.append(f"{canton}S")

            continue

    return pd.Series(salida, index=tipo_postes.index)

def tab_fle_canton_v2(
    tabla_fle,         # DataFrame con MultiIndex en columnas: (vano, 'Flecha (m)' | 'Tiro H. (kg)')
    cantones,          # pd.Series: pertenencia de cada poste a cantón normal (int o list[int])
    cantones_s         # pd.Series: pertenencia de cada poste a cantón secundario (str '1S' o list)
):
    """
    Construye dos listas de DataFrames de flechado, uno por cantón normal
    y uno por cantón secundario.

    Lógica de asignación de vanos:
    - n postes generan n-1 vanos, en el mismo orden.
    - Si el poste i pertenece a [c1, c2], el vano i pertenece a c1
      (es el último vano de c1, no el primero de c2).

    Retorna:
        tablas_normales   : list[pd.DataFrame]
        tablas_secundarios: list[pd.DataFrame]
    """

    # ----------------------------------------------------------
    # Helpers
    # ----------------------------------------------------------
    def normalizar(serie):
        """Convierte cada elemento a lista de cantones."""
        resultado = []
        for v in serie:
            if isinstance(v, list):
                resultado.append(v)
            elif v is None or (isinstance(v, float) and np.isnan(v)):
                resultado.append([])
            else:
                resultado.append([v])
        return resultado

    def extraer_cantones_unicos(norm):
        vistos, orden = set(), []
        for lista in norm:
            for c in lista:
                if c not in vistos:
                    vistos.add(c)
                    orden.append(c)
        return orden

    def separar_columnas(tabla):
        """
        Separa columnas de vanos normales y secundarios.
        Retorna dos dicts: {vano_id: {'flecha': col, 'tiro': col}}
        """
        normales, secundarios = {}, {}
        for col in tabla.columns:
            vano_id, tipo = col
            # identificar si es secundario
            es_s = isinstance(vano_id, str) and str(vano_id).endswith("S")
            d = secundarios if es_s else normales
            if vano_id not in d:
                d[vano_id] = {}
            if tipo == "Flecha (m)":
                d[vano_id]["flecha"] = col
            elif tipo == "Tiro H. (kg)":
                d[vano_id]["tiro"] = col
        return normales, secundarios

    def asignar_vanos_a_cantones(norm_cantones, cols_dict):
        """
        Agrupa postes válidos por cantón, luego asigna n-1 vanos
        a cada cantón de forma independiente.

        Para cada cantón con k postes válidos → k-1 vanos,
        tomados en orden secuencial del pool global de vanos.
        """
        vanos_ids = sorted(cols_dict.keys(),
                            key=lambda x: int(str(x).replace("S", "")))

        # Agrupar postes válidos por cantón, conservando orden de aparición
        canton_postes = {}
        cantones_orden = []
        for lista_c in norm_cantones:
            if not lista_c:
                continue
            # poste con [c1,c2]: cuenta para c1 (fin) y c2 (inicio)
            for c in lista_c:
                if c not in canton_postes:
                    canton_postes[c] = 0
                    cantones_orden.append(c)
                canton_postes[c] += 1

        # Asignar vanos secuencialmente: cada cantón con k postes toma k-1 vanos
        canton_vanos = {}
        vano_idx = 0
        for c in cantones_orden:
            n_postes = canton_postes[c]
            n_vanos = max(n_postes - 1, 0)
            canton_vanos[c] = vanos_ids[vano_idx: vano_idx + n_vanos]
            vano_idx += n_vanos

        return canton_vanos

    def construir_df(canton, vanos_ids, cols_dict, tabla):
        """Construye el DataFrame de un cantón."""
        df = pd.DataFrame()
        df["Temperatura (°C)"] = tabla.index.tolist()

        # Tiro: primer vano del cantón
        if vanos_ids:
            col_tiro = cols_dict[vanos_ids[0]].get("tiro")
            if col_tiro is not None:
                df["Tiro H. (kg)"] = tabla[col_tiro].values

        # Flechas
        for v in vanos_ids:
            col_f = cols_dict[v].get("flecha")
            if col_f is not None:
                df[f"f{v}(m)"] = tabla[col_f].values

        return df.reset_index(drop=True)

    # ----------------------------------------------------------
    # Proceso principal
    # ----------------------------------------------------------
    norm_cant  = normalizar(cantones)
    norm_cant_s = normalizar(cantones_s)

    cols_normales, cols_secundarios = separar_columnas(tabla_fle)

    canton_vanos_n = asignar_vanos_a_cantones(norm_cant,   cols_normales)
    canton_vanos_s = asignar_vanos_a_cantones(norm_cant_s, cols_secundarios)

    cantones_n_orden = extraer_cantones_unicos(norm_cant)
    cantones_s_orden = extraer_cantones_unicos(norm_cant_s)

    tablas_normales = [
        construir_df(c, canton_vanos_n.get(c, []), cols_normales, tabla_fle)
        for c in cantones_n_orden
        if c in canton_vanos_n
    ]

    tablas_secundarios = [
        construir_df(c, canton_vanos_s.get(c, []), cols_secundarios, tabla_fle)
        for c in cantones_s_orden
        if c in canton_vanos_s
    ]

    return tablas_normales, tablas_secundarios


def filas_canton_s(
    lista_tablas,       # list[DataFrame] — una tabla por cantón secundario
    cantones_s,         # pd.Series en orden de exportación (ej: '1S', NaN, '2S'...)
    vanos,              # pd.Series en orden de exportación
    postes_export,      # pd.Series en orden de exportación
    desnivel            # pd.Series en orden de exportación
):
    cantones_s   = pd.Series(cantones_s).reset_index(drop=True)
    vanos        = pd.Series(vanos).reset_index(drop=True)
    postes_export = pd.Series(postes_export).reset_index(drop=True)
    desnivel     = pd.Series(desnivel).reset_index(drop=True)

    # Máscara de postes válidos (no NaN)
    mascara_valida = cantones_s.notna()

    # Filtrar solo postes válidos
    cantones_s_val  = cantones_s[mascara_valida].reset_index(drop=True)
    vanos_val       = vanos[mascara_valida].reset_index(drop=True)
    postes_val      = postes_export[mascara_valida].reset_index(drop=True)
    desnivel_val    = desnivel[mascara_valida].reset_index(drop=True)

    # Cantones únicos en orden de aparición
    cantones_unicos = []
    for c in cantones_s_val:
        if isinstance(c, list):
            for x in c:
                if x not in cantones_unicos:
                    cantones_unicos.append(x)
        else:
            if c not in cantones_unicos:
                cantones_unicos.append(c)

    def pertenece(valor, k):
        if isinstance(valor, list):
            return k in valor
        return valor == k

    for df, k in zip(lista_tablas, cantones_unicos):

        # Índices válidos que pertenecen al cantón k
        idx = [
            i for i, c in enumerate(cantones_s_val)
            if pertenece(c, k)
        ]

        if len(idx) < 2:
            continue

        postes_canton = postes_val.iloc[idx].values

        # Vano (numeración)
        fila = ["Vano"] + list(range(1, len(postes_canton)))
        df, fila = ajustar_df(df, fila)
        df.loc[len(df)] = fila

        # Longitud (m)
        fila = ["Longitud (m)"] + vanos_val.iloc[idx[:-1]].tolist()
        df, fila = ajustar_df(df, fila)
        df.loc[len(df)] = fila

        # Poste inicial
        fila = ["Poste inicial"] + list(postes_canton[:-1])
        df, fila = ajustar_df(df, fila)
        df.loc[len(df)] = fila

        # Poste final
        fila = ["Poste final"] + list(postes_canton[1:])
        df, fila = ajustar_df(df, fila)
        df.loc[len(df)] = fila

        # Desnivel
        fila = ["Desnivel"] + desnivel_val.iloc[idx[:-1]].tolist()
        df, fila = ajustar_df(df, fila)
        df.loc[len(df)] = fila

    return lista_tablas


def tablas_por_canton_s(
    postes_orden,       # pd.Series sin repetición
    postes_exportacion, # pd.Series en orden de exportación
    cantones_s,         # pd.Series en orden de exportación ('1S', NaN, '2S'...)
    conductor,          # pd.Series alineada con postes_orden
    vano_regulacion     # pd.Series alineada con cantones (una entrada por cantón secundario)
):
    postes_orden       = pd.Series(postes_orden).reset_index(drop=True)
    postes_exportacion = pd.Series(postes_exportacion).reset_index(drop=True)
    cantones_s         = pd.Series(cantones_s).reset_index(drop=True)
    conductor          = pd.Series(conductor).reset_index(drop=True)
    vano_regulacion    = pd.Series(vano_regulacion).reset_index(drop=True)

    # Filtrar solo postes válidos (no NaN en cantones_s)
    mascara_valida     = cantones_s.notna()
    cantones_s_val     = cantones_s[mascara_valida].reset_index(drop=True)
    postes_exp_val     = postes_exportacion[mascara_valida].reset_index(drop=True)

    # Cantones únicos en orden de aparición
    cantones_unicos = []
    for c in cantones_s_val:
        if isinstance(c, list):
            for x in c:
                if x not in cantones_unicos:
                    cantones_unicos.append(x)
        else:
            if c not in cantones_unicos:
                cantones_unicos.append(c)

    def pertenece(valor, k):
        if isinstance(valor, list):
            return k in valor
        return valor == k

    tablas = []

    for i, k in enumerate(cantones_unicos):

        df = pd.DataFrame()

        # Índices de postes válidos que pertenecen al cantón k
        idx_exp = [
            j for j, c in enumerate(cantones_s_val)
            if pertenece(c, k)
        ]

        if not idx_exp:
            continue

        postes_canton = postes_exp_val.iloc[idx_exp].values
        poste_ini = postes_canton[0]
        poste_fin = postes_canton[-1]

        # Conductor del cantón (primer poste válido)
        conductores = []
        for p in postes_canton:
            idx_ord = postes_orden[postes_orden == p].index
            if len(idx_ord) > 0:
                conductores.append(conductor.iloc[idx_ord[0]])
        cond_canton = conductores[0] if conductores else np.nan

        fila = ["Conductor", cond_canton]
        df, fila = ajustar_df(df, fila)
        df.loc[len(df)] = fila

        fila = ["Cantón", k, "Poste Inicial", poste_ini]
        df, fila = ajustar_df(df, fila)
        df.loc[len(df)] = fila

        vano = vano_regulacion.iloc[i] if i < len(vano_regulacion) else np.nan
        fila = ["Vano de Regulación", vano, "Poste Final", poste_fin]
        df, fila = ajustar_df(df, fila)
        df.loc[len(df)] = fila

        tablas.append(df)

    return tablas




"""
exportar_tablas.py
==================
Funciones de exportación a Excel con formato y previsualización en Colab
para las TABLAS DE FLECHADO del proyecto FunctDesEPC.

Uso típico (en crear_tablas_mec.ipynb):
----------------------------------------
    from exportar_tablas import exportar_tablas_flechado, previsualizar_tabla_flechado

    # Exportar todas las tablas a Excel (n hojas normales + m hojas secundarias)
    exportar_tablas_flechado(
        tab_fle=tab_fle,          # lista de DataFrames con header de cada cantón normal
        tablas_p=tablas_p,        # lista de DataFrames con datos de flechado cantón normal
        tab_fle_s=tab_fle_s,      # lista de DataFrames con header de cada cantón secundario
        tablas_s=tablas_s,        # lista de DataFrames con datos de flechado cantón secundario
        filepath="Tablas_Flechado.xlsx"
    )

    # Previsualizar una tabla individual en Colab
    previsualizar_tabla_flechado(tab_fle[0], tablas_p[0], titulo="Cantón 1")
    previsualizar_tabla_flechado(tab_fle_s[0], tablas_s[0], titulo="Cantón 1S")
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, numbers
)
from openpyxl.utils import get_column_letter
from IPython.display import display, HTML


# ---------------------------------------------------------------------------
# Estilos reutilizables
# ---------------------------------------------------------------------------

def _thin():
    return Side(style="thin", color="FF000000")

def _medium():
    return Side(style="medium", color="FF000000")

def _border_thin_all():
    t = _thin()
    return Border(left=t, right=t, top=t, bottom=t)

def _border_medium_all():
    m = _medium()
    return Border(left=m, right=m, top=m, bottom=m)

def _border_medium_outer_thin_inner(left=True, right=True, top=True, bottom=True):
    """Borde medium en los lados indicados, thin en el resto."""
    def s(is_medium):
        return _medium() if is_medium else _thin()
    return Border(
        left=s(left), right=s(right), top=s(top), bottom=s(bottom)
    )

FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")   # azul claro para etiquetas
FILL_DATA   = PatternFill("solid", fgColor="FFFFFF")    # blanco para datos
FONT_BOLD   = Font(bold=True, size=9)
FONT_NORMAL = Font(bold=False, size=9)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
NUM_FMT_2DEC = '0.00'
NUM_FMT_4DEC = '0.0000'


# ---------------------------------------------------------------------------
# Helpers internos
# ---------------------------------------------------------------------------

def _write_cell(ws, row, col, value, font=None, alignment=None,
                border=None, fill=None, number_format=None):
    """Escribe un valor en una celda con formato opcional."""
    cell = ws.cell(row=row, column=col, value=value)
    if font is not None:
        cell.font = font
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
    if fill is not None:
        cell.fill = fill
    if number_format is not None:
        cell.number_format = number_format
    return cell


def _merge_and_write(ws, row, col_start, col_end, value,
                        font=None, alignment=None, border=None, fill=None):
    """Fusiona celdas y escribe el valor en la primera."""
    if col_start < col_end:
        ws.merge_cells(
            start_row=row, start_column=col_start,
            end_row=row,   end_column=col_end
        )
    _write_cell(ws, row, col_start, value, font=font,
                alignment=alignment, border=border, fill=fill)


def _auto_col_width(ws, min_width=8, max_width=30):
    """Ajusta el ancho de columnas automáticamente según contenido."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


# ---------------------------------------------------------------------------
# Escritura del bloque HEADER de una tabla de flechado
# ---------------------------------------------------------------------------

def _escribir_header_tabla(ws, tab_fle_df, start_row=1, start_col=1):
    """
    Escribe la parte superior de la tabla (header) a partir de tab_fle_df.

    tab_fle_df se asume un DataFrame cuyas filas son:
        0: Conductor
        1: Cantón, Poste Inicial
        2: Vano de Regulación, Poste Final
        3: Vano (números de vano)
        4: Longitud (m)
        5: Poste inicial
        6: Poste final
        7: Desnivel

    La primera columna contiene las etiquetas (ej. "Conductor", "Cantón", etc.)
    y las demás columnas contienen los valores.

    Retorna la siguiente fila disponible (start_row + número de filas escritas).
    """
    b_med = _border_medium_all()
    b_thin = _border_thin_all()
    label_fill = FILL_HEADER

    n_rows, n_cols = tab_fle_df.shape
    # col offset: etiqueta en start_col, datos a partir de start_col+1
    label_col = start_col
    data_col_start = start_col + 1

    for rel_row, (idx, row_data) in enumerate(tab_fle_df.iterrows()):
        abs_row = start_row + rel_row
        row_vals = row_data.tolist()

        label = row_vals[0] if len(row_vals) > 0 else ""
        data_vals = row_vals[1:]

        # Decidir si la fila es "de encabezado de vano" (borde medium) o no
        is_vano_row = str(label).strip().lower() in {
            "vano", "longitud (m)", "poste inicial", "poste final", "desnivel"
        }
        border = b_med if is_vano_row else b_thin

        # Etiqueta — fusionar 2 columnas si hay espacio
        _merge_and_write(
            ws, abs_row, label_col, label_col + 1,
            value=label,
            font=FONT_BOLD,
            alignment=ALIGN_LEFT,
            border=border,
            fill=label_fill
        )

        # Valores de datos
        for rel_col, val in enumerate(data_vals):
            c = data_col_start + 1 + rel_col  # +1 porque label ocupa 2 cols
            # Formato numérico para desnivel y longitud
            num_fmt = None
            if isinstance(val, float):
                num_fmt = NUM_FMT_4DEC if abs(val) < 10 else NUM_FMT_2DEC
            _write_cell(
                ws, abs_row, c, val,
                font=FONT_NORMAL,
                alignment=ALIGN_CENTER,
                border=border,
                fill=FILL_DATA,
                number_format=num_fmt
            )

    return start_row + n_rows


# ---------------------------------------------------------------------------
# Escritura del bloque DATOS (tablas_p / tablas_s)
# ---------------------------------------------------------------------------

def _escribir_datos_tabla(ws, tablas_df, start_row=1, start_col=1):
    """
    Escribe la parte inferior de la tabla (datos de temperatura/tense/flechas).

    tablas_df se asume un DataFrame con columnas:
        Temperatura, Tense, f_vano1, f_vano2, ...

    La fila de encabezado del DataFrame se escribe primero (nombres de columna).
    Luego se escriben los datos fila a fila.

    Retorna la siguiente fila disponible.
    """
    b_med = _border_medium_all()
    b_thin = _border_thin_all()

    # Encabezado de columnas
    cols = list(tablas_df.columns)
    for rel_col, col_name in enumerate(cols):
        c = start_col + rel_col
        _write_cell(
            ws, start_row, c, col_name,
            font=FONT_BOLD,
            alignment=ALIGN_CENTER,
            border=b_med,
            fill=FILL_HEADER
        )

    # Subencabezado de unidades (fila extra con unidades si la columna es "Temperatura" o "Tense")
    unit_row = start_row + 1
    for rel_col, col_name in enumerate(cols):
        c = start_col + rel_col
        unit = ""
        col_lower = str(col_name).lower()
        if "temperatura" in col_lower:
            unit = "(ºC)"
        elif "tense" in col_lower:
            unit = "(daN)"
        elif col_name not in ("", None):
            unit = "f (m)"
        _write_cell(
            ws, unit_row, c, unit,
            font=FONT_NORMAL,
            alignment=ALIGN_CENTER,
            border=b_med,
            fill=FILL_HEADER
        )

    # Datos
    data_start_row = start_row + 2
    for rel_row, (_, row_data) in enumerate(tablas_df.iterrows()):
        abs_row = data_start_row + rel_row
        for rel_col, val in enumerate(row_data):
            c = start_col + rel_col
            num_fmt = None
            if isinstance(val, float):
                num_fmt = NUM_FMT_4DEC
            _write_cell(
                ws, abs_row, c, val,
                font=FONT_NORMAL,
                alignment=ALIGN_CENTER,
                border=b_thin,
                fill=FILL_DATA,
                number_format=num_fmt
            )

    return data_start_row + len(tablas_df)


# ---------------------------------------------------------------------------
# Función principal de exportación
# ---------------------------------------------------------------------------

def exportar_tablas_flechado(
    tab_fle,
    tablas_p,
    tab_fle_s,
    tablas_s,
    filepath="Tablas_Flechado.xlsx"
):
    """
    Exporta las tablas de flechado a un archivo Excel con formato.

    Genera n + m hojas:
      - n hojas para cantones normales  → nombres: "Canton_1", "Canton_2", ...
      - m hojas para cantones secundarios → nombres: "Canton_1S", "Canton_2S", ...

    Cada hoja contiene:
      - Parte superior: header del cantón (tab_fle[i] o tab_fle_s[i])
      - Parte inferior: datos de temperatura/flechas (tablas_p[i] o tablas_s[i])

    Parámetros
    ----------
    tab_fle : list[pd.DataFrame]
        Lista de DataFrames con el header de cada cantón normal.
    tablas_p : list[pd.DataFrame]
        Lista de DataFrames con los datos de flechado de cada cantón normal.
        Debe tener la misma longitud que tab_fle.
    tab_fle_s : list[pd.DataFrame]
        Lista de DataFrames con el header de cada cantón secundario.
    tablas_s : list[pd.DataFrame]
        Lista de DataFrames con los datos de flechado de cada cantón secundario.
        Debe tener la misma longitud que tab_fle_s.
    filepath : str
        Ruta del archivo Excel de salida.

    Retorna
    -------
    str
        Ruta del archivo generado.
    """
    if len(tab_fle) != len(tablas_p):
        raise ValueError(
            f"tab_fle tiene {len(tab_fle)} elementos pero tablas_p tiene {len(tablas_p)}. "
            "Deben tener la misma longitud."
        )
    if len(tab_fle_s) != len(tablas_s):
        raise ValueError(
            f"tab_fle_s tiene {len(tab_fle_s)} elementos pero tablas_s tiene {len(tablas_s)}. "
            "Deben tener la misma longitud."
        )

    wb = Workbook()
    # Eliminar hoja vacía por defecto
    wb.remove(wb.active)

    # -- Cantones normales --
    for i, (header_df, datos_df) in enumerate(zip(tab_fle, tablas_p)):
        sheet_name = f"Canton_{i + 1}"
        ws = wb.create_sheet(title=sheet_name)
        next_row = _escribir_header_tabla(ws, header_df, start_row=1, start_col=1)
        # Separador de 1 fila en blanco entre header y datos
        _escribir_datos_tabla(ws, datos_df, start_row=next_row + 1, start_col=1)
        _auto_col_width(ws)

    # -- Cantones secundarios --
    for i, (header_df, datos_df) in enumerate(zip(tab_fle_s, tablas_s)):
        sheet_name = f"Canton_{i + 1}S"
        ws = wb.create_sheet(title=sheet_name)
        next_row = _escribir_header_tabla(ws, header_df, start_row=1, start_col=1)
        _escribir_datos_tabla(ws, datos_df, start_row=next_row + 1, start_col=1)
        _auto_col_width(ws)

    wb.save(filepath)
    print(f"✅ Archivo guardado: {filepath}  ({len(tab_fle)} cantones + {len(tab_fle_s)} secundarios)")
    return filepath


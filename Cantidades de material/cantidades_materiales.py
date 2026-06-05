# -*- coding: utf-8 -*-
"""
cantidades_materiales.py
========================

Módulo para el cálculo de cantidades totales de material de un proyecto a
partir de:

  1. Una *planilla de estructuras* (PlanillaEstTotal*.XLS) que indica, para
     cada poste, los armados primarios y secundarios instalados.
  2. Un *catálogo de cantidades por armado* (Cantidades_de_postes.xlsx) que
     indica, para cada armado, cuántas unidades de cada material lo componen
     (el "multiplicador").

La cantidad total de un material es:

        total(material) = Σ_postes Σ_armados_del_poste  multiplicador(material, armado)

Diseño
------
El módulo está dividido en funciones pequeñas e independientes para que, si
algo falla, el error quede aislado en una sola etapa y sea fácil de reparar.
El flujo completo lo orquesta `generar_cantidades_materiales`, pero cada etapa
puede ejecutarse y depurarse por separado:

    descargar_archivo_drive    ->  trae el catálogo desde Google Drive
    cargar_catalogo            ->  lee el .xlsx y construye el catálogo en memoria
    extraer_armados_planilla   ->  lista los armados de cada poste
    calcular_cantidades        ->  suma usando los multiplicadores
    exportar_cantidades_excel  ->  escribe el .xlsx de salida

El diseño es modular: para añadir una nueva fuente de datos, un nuevo formato
de catálogo o una nueva exportación, basta con escribir una función adicional
y enchufarla en el orquestador, sin tocar el resto.

Autor: (automatización cálculos mecánicos)
"""

from __future__ import annotations

import os
import re
import unicodedata
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd


# =====================================================================
#  0. UTILIDADES DE NORMALIZACIÓN
# =====================================================================

def normalizar_codigo_armado(codigo) -> str:
    """
    Normaliza un código de armado para poder comparar la planilla contra el
    catálogo, aunque difieran en espacios, mayúsculas o sufijos.

    Reglas aplicadas:
      * Pasa a mayúsculas.
      * Elimina cualquier contenido entre paréntesis  ->  "MTF635-1 (S)" => "MTF6351"
      * Elimina espacios, guiones y guiones bajos      ->  "MTF 635-1"    => "MTF6351"

    Devuelve "" para valores vacíos / NaN.

    >>> normalizar_codigo_armado("MTF 635-1")
    'MTF6351'
    >>> normalizar_codigo_armado("MTF635-1 (S)")
    'MTF6351'
    """
    if codigo is None:
        return ""
    if isinstance(codigo, float) and np.isnan(codigo):
        return ""
    s = str(codigo).strip()
    if s == "" or s.lower() == "nan":
        return ""
    s = s.upper()
    s = re.sub(r"\(.*?\)", "", s)        # quita sufijos entre paréntesis: (S), (M)...
    s = re.sub(r"[\s\-_]", "", s)         # quita espacios, guiones, guiones bajos
    return s


def _normalizar_texto(texto) -> str:
    """Normaliza un texto a minúsculas sin acentos (para detectar encabezados)."""
    if texto is None:
        return ""
    s = str(texto)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.strip().lower()


# =====================================================================
#  1. DESCARGA DEL CATÁLOGO DESDE GOOGLE DRIVE
# =====================================================================

def extraer_file_id_drive(enlace_o_id: str) -> str:
    """
    Extrae el file-id de un enlace de Google Drive. Acepta tanto el id directo
    como las formas habituales de URL:

        https://drive.google.com/file/d/<ID>/view?usp=sharing
        https://drive.google.com/open?id=<ID>
        https://drive.google.com/uc?id=<ID>
        <ID>
    """
    if enlace_o_id is None:
        raise ValueError("El enlace/id de Drive está vacío.")
    s = str(enlace_o_id).strip()
    # /file/d/<ID>/
    m = re.search(r"/d/([A-Za-z0-9_-]+)", s)
    if m:
        return m.group(1)
    # ?id=<ID>  o  &id=<ID>
    m = re.search(r"[?&]id=([A-Za-z0-9_-]+)", s)
    if m:
        return m.group(1)
    # ya es un id "pelado"
    if re.fullmatch(r"[A-Za-z0-9_-]{20,}", s):
        return s
    raise ValueError(f"No se pudo extraer un file-id de Drive desde: {enlace_o_id!r}")


def descargar_archivo_drive(enlace_o_id: str,
                            destino: str,
                            force: bool = False,
                            quiet: bool = False) -> str:
    """
    Descarga un archivo público de Google Drive a `destino` usando `gdown`.

    Esta etapa está deliberadamente aislada: si la descarga falla (permisos,
    red, id inválido) el error se origina aquí y no contamina el resto del
    cálculo. Si ya existe el archivo y `force=False`, no lo vuelve a descargar.

    NOTA: en un entorno Colab con Drive ya montado normalmente NO se necesita
    esta función; basta con apuntar `ruta` al archivo dentro de
    /content/drive/MyDrive/... Úsala solo cuando el catálogo se sirva por un
    enlace de Drive en lugar de estar montado.

    Devuelve la ruta local del archivo descargado.
    """
    if (not force) and os.path.exists(destino) and os.path.getsize(destino) > 0:
        if not quiet:
            print(f"[drive] Ya existe, no se re-descarga: {destino}")
        return destino

    try:
        import gdown  # import perezoso: solo se exige si se usa la descarga
    except ImportError as e:
        raise ImportError(
            "Se requiere el paquete 'gdown' para descargar desde Drive. "
            "Instálalo con:  pip install gdown"
        ) from e

    file_id = extraer_file_id_drive(enlace_o_id)
    url = f"https://drive.google.com/uc?id={file_id}"
    os.makedirs(os.path.dirname(os.path.abspath(destino)) or ".", exist_ok=True)
    salida = gdown.download(url, destino, quiet=quiet)
    if not salida or not os.path.exists(destino):
        raise RuntimeError(
            f"gdown no pudo descargar el archivo (id={file_id}). "
            "Verifica que el enlace sea público ('cualquiera con el enlace')."
        )
    if not quiet:
        print(f"[drive] Descargado: {destino}")
    return destino


# =====================================================================
#  2. CARGA DEL CATÁLOGO DE CANTIDADES POR ARMADO
# =====================================================================

@dataclass
class Catalogo:
    """
    Representa el catálogo de cantidades por armado ya procesado.

    Atributos
    ---------
    materiales : dict
        { codigo_armado_normalizado : { clave_material : multiplicador } }
    info_material : dict
        { clave_material : {"nombre":..., "codigo":..., "unidad":...} }
        Guarda la metadata de cada material para la exportación final.
    armados : dict
        { codigo_armado_normalizado : codigo_armado_original }
        Permite mostrar el código tal y como aparece en el catálogo.
    hojas : list
        Hojas del Excel que se fusionaron en este catálogo.
    """
    materiales: Dict[str, Dict[str, float]] = field(default_factory=dict)
    info_material: Dict[str, Dict[str, str]] = field(default_factory=dict)
    armados: Dict[str, str] = field(default_factory=dict)
    hojas: List[str] = field(default_factory=list)

    def armados_disponibles(self) -> List[str]:
        """Códigos de armado (originales) presentes en el catálogo, ordenados."""
        return sorted(self.armados.values())

    def __repr__(self) -> str:
        return (f"<Catalogo hojas={self.hojas} "
                f"armados={len(self.armados)} materiales={len(self.info_material)}>")


def _detectar_layout_hoja(crudo: pd.DataFrame) -> Dict[str, int]:
    """
    Detecta automáticamente la disposición de columnas/filas de una hoja del
    catálogo a partir de las dos primeras filas.

    Soporta los dos formatos observados:
      Formato A (AFINIA): Elementos | Codigo | Unidades | Armado...
      Formato B (ESSA):   Elementos | Codigo JDE |       | Armado...

    Devuelve un dict con:
        fila_codigos      -> índice de fila donde están los códigos de armado
        col_elemento      -> columna del nombre del material
        col_codigo        -> columna del código del material (o None)
        col_unidad        -> columna de la unidad (o None)
        col_inicio_armado -> primera columna con códigos de armado
    """
    fila0 = [_normalizar_texto(x) for x in crudo.iloc[0].tolist()]

    col_elemento = next((i for i, v in enumerate(fila0) if v.startswith("element")), 0)
    col_codigo = next((i for i, v in enumerate(fila0) if v.startswith("codigo")), None)
    col_unidad = next((i for i, v in enumerate(fila0) if v.startswith("unidad")), None)

    # La columna donde comienza la cabecera "Armado" marca el inicio de armados.
    col_armado_marca = next((i for i, v in enumerate(fila0) if v.startswith("armado")), None)
    if col_armado_marca is None:
        # Fallback: empezar tras la última columna de metadata conocida.
        ultima_meta = max([c for c in (col_elemento, col_codigo, col_unidad)
                           if c is not None])
        col_inicio_armado = ultima_meta + 1
    else:
        col_inicio_armado = col_armado_marca

    # Los códigos de armado suelen estar en la fila 1 (segunda fila).
    fila_codigos = 1
    return {
        "fila_codigos": fila_codigos,
        "col_elemento": col_elemento,
        "col_codigo": col_codigo,
        "col_unidad": col_unidad,
        "col_inicio_armado": col_inicio_armado,
    }


def cargar_catalogo(ruta: str,
                    hojas: Optional[Sequence[str]] = None,
                    layout: Optional[Dict[str, int]] = None,
                    verbose: bool = True) -> Catalogo:
    """
    Lee el catálogo de cantidades por armado y lo convierte en un objeto
    `Catalogo` consultable.

    Parámetros
    ----------
    ruta : str
        Ruta al .xlsx del catálogo (Cantidades_de_postes.xlsx).
    hojas : lista de str, opcional
        Nombres de las hojas a usar. Si es None se usan TODAS las hojas que
        contengan la cabecera "Armado". Los armados de distintas hojas se
        fusionan; si un mismo armado aparece en varias hojas, gana la primera.
    layout : dict, opcional
        Fuerza la disposición de columnas (ver `_detectar_layout_hoja`). Si es
        None se detecta automáticamente por hoja.
    verbose : bool
        Imprime un resumen de la carga.

    Devuelve
    --------
    Catalogo
    """
    if not os.path.exists(ruta):
        raise FileNotFoundError(f"No se encontró el catálogo: {ruta}")

    xls = pd.ExcelFile(ruta)
    hojas_objetivo = list(hojas) if hojas else list(xls.sheet_names)

    catalogo = Catalogo()
    for hoja in hojas_objetivo:
        if hoja not in xls.sheet_names:
            if verbose:
                print(f"[catalogo] AVISO: hoja inexistente, se omite: {hoja!r}")
            continue

        crudo = pd.read_excel(xls, sheet_name=hoja, header=None)
        if crudo.shape[0] < 3:
            if verbose:
                print(f"[catalogo] AVISO: hoja sin datos suficientes: {hoja!r}")
            continue

        lay = layout or _detectar_layout_hoja(crudo)
        # Si la hoja no tiene marca "Armado" en fila 0, probablemente no es un
        # catálogo de armados (p.ej. la hoja resumen "Cantidades ESSA"); se omite.
        fila0 = [_normalizar_texto(x) for x in crudo.iloc[0].tolist()]
        if not any(v.startswith("armado") for v in fila0):
            if verbose:
                print(f"[catalogo] Hoja sin cabecera 'Armado', se omite: {hoja!r}")
            continue

        f_cod = lay["fila_codigos"]
        c_ini = lay["col_inicio_armado"]
        c_elem = lay["col_elemento"]
        c_codm = lay["col_codigo"]
        c_uni = lay["col_unidad"]

        # Mapa columna_excel -> codigo_armado para esta hoja.
        codigos_fila = crudo.iloc[f_cod]
        col_a_armado: Dict[int, str] = {}
        for col in range(c_ini, crudo.shape[1]):
            crudo_codigo = codigos_fila.iloc[col]
            if pd.isna(crudo_codigo) or str(crudo_codigo).strip() == "":
                continue
            col_a_armado[col] = str(crudo_codigo).strip()

        n_armados_hoja = 0
        # Filas de materiales: a partir de la fila siguiente a los códigos.
        for r in range(f_cod + 1, crudo.shape[0]):
            nombre = crudo.iloc[r, c_elem]
            if pd.isna(nombre) or str(nombre).strip() == "":
                continue
            nombre = str(nombre).strip()
            codigo_mat = ""
            if c_codm is not None and not pd.isna(crudo.iloc[r, c_codm]):
                codigo_mat = str(crudo.iloc[r, c_codm]).strip()
                # Limpia floats tipo "211274.0"
                codigo_mat = re.sub(r"\.0$", "", codigo_mat)
            unidad = ""
            if c_uni is not None and not pd.isna(crudo.iloc[r, c_uni]):
                unidad = str(crudo.iloc[r, c_uni]).strip()

            # Clave única del material: código si existe y es real, si no el nombre.
            if codigo_mat and codigo_mat.upper() not in ("N/A", "NAN"):
                clave_mat = f"COD::{codigo_mat}"
            else:
                clave_mat = f"NOM::{nombre.upper()}"

            if clave_mat not in catalogo.info_material:
                catalogo.info_material[clave_mat] = {
                    "nombre": nombre,
                    "codigo": codigo_mat,
                    "unidad": unidad,
                }

            for col, armado_orig in col_a_armado.items():
                valor = crudo.iloc[r, col]
                if pd.isna(valor):
                    continue
                try:
                    mult = float(valor)
                except (TypeError, ValueError):
                    continue
                if mult == 0:
                    continue
                armado_norm = normalizar_codigo_armado(armado_orig)
                if armado_norm == "":
                    continue
                catalogo.armados.setdefault(armado_norm, armado_orig)
                fila_mat = catalogo.materiales.setdefault(armado_norm, {})
                # Suma por si el mismo material aparece en varias filas del armado.
                fila_mat[clave_mat] = fila_mat.get(clave_mat, 0.0) + mult

        n_armados_hoja = len(col_a_armado)
        catalogo.hojas.append(hoja)
        if verbose:
            print(f"[catalogo] Hoja {hoja!r}: {n_armados_hoja} armados leídos.")

    if verbose:
        print(f"[catalogo] Total -> {len(catalogo.armados)} armados, "
              f"{len(catalogo.info_material)} materiales distintos.")
    if not catalogo.armados:
        raise ValueError(
            "El catálogo quedó vacío. Revisa el nombre de las hojas o el layout."
        )
    return catalogo


# =====================================================================
#  3. EXTRACCIÓN DE LOS ARMADOS DE CADA POSTE EN LA PLANILLA
# =====================================================================

# Columnas (multi-índice) por defecto en las planillas PlanillaEstTotal*.XLS.
COLUMNAS_ARMADO_DEFAULT = [
    ("Armado Primario",   "Primario1"),
    ("Armado Primario",   "Primario2"),
    ("Armado Secundario", "Secundario1"),
    ("Armado Secundario", "Secundario2"),
]
COL_NOMBRE_DEFAULT = ("Identificación", "Nombre Est.")
COL_NRUTA_DEFAULT = ("Identificación", "N° Est.")
COL_DERIVACION_DEFAULT = ("Identificación", "Derivación")


def cargar_planilla(ruta: str) -> pd.DataFrame:
    """
    Lee una planilla de estructuras PlanillaEstTotal*.XLS con cabecera de dos
    niveles (igual que el resto del notebook). Aislada para depurar la lectura
    del archivo por separado.
    """
    if not os.path.exists(ruta):
        raise FileNotFoundError(f"No se encontró la planilla: {ruta}")
    # engine='xlrd' para .XLS legado; pandas lo autodetecta normalmente, pero
    # se fija explícitamente para evitar sorpresas.
    engine = "xlrd" if ruta.lower().endswith(".xls") else None
    return pd.read_excel(ruta, header=[0, 1], engine=engine)


def extraer_armados_planilla(
    est_df: pd.DataFrame,
    columnas_armado: Sequence[Tuple[str, str]] = COLUMNAS_ARMADO_DEFAULT,
    col_nombre: Tuple[str, str] = COL_NOMBRE_DEFAULT,
    col_nruta: Optional[Tuple[str, str]] = COL_NRUTA_DEFAULT,
    col_derivacion: Optional[Tuple[str, str]] = COL_DERIVACION_DEFAULT,
) -> pd.DataFrame:
    """
    Convierte la planilla (un poste por fila, hasta 4 armados por fila) en una
    tabla "larga" con un armado por fila:

        nombre_poste | derivacion | n_est | tipo_armado | armado

    `tipo_armado` indica de qué columna provino (Primario1, Secundario1, ...),
    útil para auditar. Las celdas vacías se descartan.
    """
    registros: List[dict] = []
    for idx, fila in est_df.iterrows():
        nombre = fila.get(col_nombre, idx)
        derivacion = fila.get(col_derivacion) if col_derivacion else None
        n_est = fila.get(col_nruta) if col_nruta else None
        for col in columnas_armado:
            if col not in est_df.columns:
                continue
            valor = fila.get(col)
            if pd.isna(valor) or str(valor).strip() == "":
                continue
            registros.append({
                "nombre_poste": str(nombre).strip() if pd.notna(nombre) else "",
                "derivacion": str(derivacion).strip() if pd.notna(derivacion) else "",
                "n_est": n_est,
                "tipo_armado": col[1] if isinstance(col, tuple) else str(col),
                "armado": str(valor).strip(),
            })
    df = pd.DataFrame(registros,
                      columns=["nombre_poste", "derivacion", "n_est",
                               "tipo_armado", "armado"])
    df["armado_norm"] = df["armado"].apply(normalizar_codigo_armado)
    return df


# =====================================================================
#  4. CÁLCULO DE CANTIDADES TOTALES
# =====================================================================

def aplicar_alias(armado_norm: str, alias: Optional[Dict[str, str]]) -> str:
    """
    Aplica un mapa de equivalencias manual sobre el código normalizado.

    `alias` se da en términos legibles, p.ej. {"MTF731-1": "MTF 731"}; aquí se
    normalizan ambos lados para comparar. Devuelve el código normalizado
    resultante (al que apunta el alias) o el mismo si no hay alias.
    """
    if not alias:
        return armado_norm
    alias_norm = {normalizar_codigo_armado(k): normalizar_codigo_armado(v)
                  for k, v in alias.items()}
    return alias_norm.get(armado_norm, armado_norm)


def calcular_cantidades(
    armados_planilla: pd.DataFrame,
    catalogo: Catalogo,
    alias: Optional[Dict[str, str]] = None,
    verbose: bool = True,
) -> Dict[str, pd.DataFrame]:
    """
    Suma los materiales de todos los armados de todos los postes.

    Devuelve un diccionario con tres DataFrames:

      'totales'        -> codigo | material | unidad | cantidad_total
      'no_encontrados' -> armados de la planilla que no existen en el catálogo
                          (con cuántas veces aparecen) -> para corregir con alias
      'detalle'        -> aporte de cada armado a cada material (trazabilidad)

    El cálculo nunca falla silenciosamente: los armados sin correspondencia se
    reportan en 'no_encontrados' en lugar de descartarse sin avisar.
    """
    # Acumuladores
    totales: Dict[str, float] = {}            # clave_material -> cantidad
    detalle_rows: List[dict] = []
    faltantes: Dict[str, dict] = {}           # armado_orig -> {conteo, norm}

    for _, fila in armados_planilla.iterrows():
        armado_orig = fila["armado"]
        norm = aplicar_alias(fila["armado_norm"], alias)

        if norm not in catalogo.materiales:
            info = faltantes.setdefault(
                armado_orig, {"armado": armado_orig, "armado_norm": norm, "veces": 0})
            info["veces"] += 1
            continue

        for clave_mat, mult in catalogo.materiales[norm].items():
            totales[clave_mat] = totales.get(clave_mat, 0.0) + mult
            detalle_rows.append({
                "nombre_poste": fila["nombre_poste"],
                "armado": armado_orig,
                "material": catalogo.info_material[clave_mat]["nombre"],
                "codigo": catalogo.info_material[clave_mat]["codigo"],
                "cantidad": mult,
            })

    # --- Tabla de totales ---
    filas_tot = []
    for clave_mat, cant in totales.items():
        info = catalogo.info_material[clave_mat]
        filas_tot.append({
            "codigo": info["codigo"],
            "material": info["nombre"],
            "unidad": info["unidad"],
            "cantidad_total": cant,
        })
    df_totales = (pd.DataFrame(filas_tot,
                               columns=["codigo", "material", "unidad", "cantidad_total"])
                  .sort_values("material", key=lambda s: s.str.lower())
                  .reset_index(drop=True))

    # --- Tabla de no encontrados ---
    df_faltantes = (pd.DataFrame(list(faltantes.values()),
                                 columns=["armado", "armado_norm", "veces"])
                    .sort_values("armado")
                    .reset_index(drop=True))

    df_detalle = pd.DataFrame(
        detalle_rows,
        columns=["nombre_poste", "armado", "material", "codigo", "cantidad"])

    if verbose:
        print(f"[calculo] Materiales totales distintos: {len(df_totales)}")
        if len(df_faltantes):
            print(f"[calculo] ⚠ Armados SIN correspondencia en el catálogo "
                  f"({len(df_faltantes)}):")
            for _, r in df_faltantes.iterrows():
                print(f"          - {r['armado']!r} (aparece {r['veces']} vez/veces)")
            print("          Añádelos al diccionario `alias` o al catálogo.")
        else:
            print("[calculo] ✅ Todos los armados de la planilla se encontraron.")

    return {"totales": df_totales,
            "no_encontrados": df_faltantes,
            "detalle": df_detalle}


# =====================================================================
#  5. EXPORTACIÓN A EXCEL
# =====================================================================

def exportar_cantidades_excel(resultado: Dict[str, pd.DataFrame],
                              ruta_salida: str,
                              incluir_detalle: bool = True,
                              verbose: bool = True) -> str:
    """
    Escribe el resultado en un .xlsx con formato profesional:

      Hoja 'Cantidades'        -> material y cantidad total (entregable principal)
      Hoja 'Armados no hallados'-> trazabilidad de lo que no se pudo mapear
      Hoja 'Detalle' (opcional)-> aporte poste×armado×material

    Devuelve la ruta del archivo escrito.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    os.makedirs(os.path.dirname(os.path.abspath(ruta_salida)) or ".", exist_ok=True)

    df_tot = resultado["totales"].copy()
    df_tot.columns = ["Código", "Material", "Unidad", "Cantidad total"]
    df_falt = resultado["no_encontrados"].copy()
    if len(df_falt):
        df_falt.columns = ["Armado (planilla)", "Código normalizado", "Veces"]

    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        df_tot.to_excel(writer, sheet_name="Cantidades", index=False)
        if len(df_falt):
            df_falt.to_excel(writer, sheet_name="Armados no hallados", index=False)
        else:
            pd.DataFrame({"Estado": ["Todos los armados fueron encontrados ✅"]}) \
                .to_excel(writer, sheet_name="Armados no hallados", index=False)
        if incluir_detalle and len(resultado.get("detalle", [])):
            det = resultado["detalle"].copy()
            det.columns = ["Poste", "Armado", "Material", "Código", "Cantidad"]
            det.to_excel(writer, sheet_name="Detalle", index=False)

        wb = writer.book
        # --- Formato de cabeceras y anchos ---
        encabezado_fill = PatternFill("solid", fgColor="1F4E78")
        encabezado_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        celda_font = Font(name="Arial", size=10)
        centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
        borde = Border(*[Side(style="thin", color="D9D9D9")] * 4)

        for ws in wb.worksheets:
            for col_idx, col_cells in enumerate(ws.iter_cols(), start=1):
                max_len = 0
                for i, celda in enumerate(col_cells):
                    celda.border = borde
                    if i == 0:  # cabecera
                        celda.fill = encabezado_fill
                        celda.font = encabezado_font
                        celda.alignment = centro
                    else:
                        celda.font = celda_font
                    valor = "" if celda.value is None else str(celda.value)
                    max_len = max(max_len, len(valor))
                ws.column_dimensions[get_column_letter(col_idx)].width = \
                    min(max(12, max_len + 2), 60)
            ws.freeze_panes = "A2"

    if verbose:
        print(f"[export] Archivo escrito: {ruta_salida}")
    return ruta_salida


# =====================================================================
#  6. ORQUESTADOR DE ALTO NIVEL
# =====================================================================

def generar_cantidades_materiales(
    ruta_planilla: str,
    ruta_catalogo: Optional[str] = None,
    enlace_catalogo_drive: Optional[str] = None,
    destino_catalogo: str = "Cantidades_de_postes.xlsx",
    hojas_catalogo: Optional[Sequence[str]] = None,
    columnas_armado: Sequence[Tuple[str, str]] = COLUMNAS_ARMADO_DEFAULT,
    alias: Optional[Dict[str, str]] = None,
    ruta_salida: str = "Cantidades_totales_proyecto.xlsx",
    incluir_detalle: bool = True,
    verbose: bool = True,
) -> Dict[str, object]:
    """
    Orquesta el flujo completo, ejecutando cada etapa por separado y atrapando
    su error para que sea fácil saber DÓNDE falló.

    Indica el catálogo de UNA de estas dos formas:
      * `ruta_catalogo`         -> archivo ya disponible localmente / en Drive montado.
      * `enlace_catalogo_drive` -> enlace de Drive; se descarga a `destino_catalogo`.

    Devuelve un dict con las claves:
        'catalogo', 'armados', 'totales', 'no_encontrados', 'detalle', 'ruta_salida'
    """
    etapa = "inicio"
    try:
        # --- Etapa 1: obtener el catálogo ---
        etapa = "obtención del catálogo"
        if ruta_catalogo is None:
            if enlace_catalogo_drive is None:
                raise ValueError(
                    "Debes indicar 'ruta_catalogo' o 'enlace_catalogo_drive'.")
            ruta_catalogo = descargar_archivo_drive(
                enlace_catalogo_drive, destino_catalogo, verbose=verbose)

        # --- Etapa 2: cargar catálogo ---
        etapa = "carga del catálogo"
        catalogo = cargar_catalogo(ruta_catalogo, hojas=hojas_catalogo, verbose=verbose)

        # --- Etapa 3: cargar planilla ---
        etapa = "carga de la planilla"
        est_df = cargar_planilla(ruta_planilla)

        # --- Etapa 4: extraer armados ---
        etapa = "extracción de armados"
        armados = extraer_armados_planilla(est_df, columnas_armado=columnas_armado)
        if verbose:
            print(f"[planilla] {armados['nombre_poste'].nunique()} postes, "
                  f"{len(armados)} armados instalados.")

        # --- Etapa 5: calcular cantidades ---
        etapa = "cálculo de cantidades"
        resultado = calcular_cantidades(armados, catalogo, alias=alias, verbose=verbose)

        # --- Etapa 6: exportar ---
        etapa = "exportación a Excel"
        exportar_cantidades_excel(resultado, ruta_salida,
                                  incluir_detalle=incluir_detalle, verbose=verbose)

    except Exception as e:
        print(f"\n❌ Falló la etapa: {etapa}\n   {type(e).__name__}: {e}")
        raise

    return {
        "catalogo": catalogo,
        "armados": armados,
        "totales": resultado["totales"],
        "no_encontrados": resultado["no_encontrados"],
        "detalle": resultado["detalle"],
        "ruta_salida": ruta_salida,
    }

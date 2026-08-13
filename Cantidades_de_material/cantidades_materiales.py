# -*- coding: utf-8 -*-
"""
cantidades_materiales.py
========================

Módulo para el cálculo de cantidades totales de material de un proyecto a
partir de:

  1. Una *planilla de estructuras* (PlanillaEstTotal*.XLS) que indica, para
     cada poste, los armados primarios y secundarios instalados.
  2. Un *catálogo de cantidades por armado* (Cantidades_de_postes.xlsx) que
     indica, para cada armado, cuántas unidades de cada material lo componen
     (el "multiplicador").

La cantidad total de un material es:

        total(material) = Σ_postes Σ_armados_del_poste  multiplicador(material, armado)

Diseño
------
El módulo está dividido en funciones pequeñas e independientes para que, si
algo falla, el error quede aislado en una sola etapa y sea fácil de reparar.
El flujo completo lo orquesta `generar_cantidades_materiales`, pero cada etapa
puede ejecutarse y depurarse por separado:

    cargar_catalogo            ->  lee el .xlsx desde Drive montado y construye el catálogo
    extraer_armados_planilla   ->  lista los armados de cada poste desde est_v_max
    calcular_cantidades        ->  suma usando los multiplicadores
    exportar_cantidades_excel  ->  escribe el .xlsx de salida

El diseño es modular: para añadir una nueva fuente de datos, un nuevo formato
de catálogo o una nueva exportación, basta con escribir una función adicional
y enchufarla en el orquestador, sin tocar el resto.

Autor: (automatización cálculos mecánicos)
"""

from __future__ import annotations

import os
import re
import sys
import unicodedata
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd

# zona_contaminacion.py debe estar en la misma carpeta que este módulo (o en
# el PYTHONPATH) para poder determinar el nivel de contaminación de la línea
# a partir de las coordenadas de los postes. Se importa de forma perezosa
# (dentro de las funciones que lo usan) para no romper el resto del módulo si
# faltan sus dependencias (Pillow, scipy) o el archivo mapa_contaminacion.png.
_DIR_MODULO = os.path.dirname(os.path.abspath(__file__))
if _DIR_MODULO not in sys.path:
    sys.path.append(_DIR_MODULO)


# =====================================================================
#  0. UTILIDADES DE NORMALIZACIÓN
# =====================================================================

def normalizar_codigo_armado(codigo) -> str:
    """
    Normaliza un código de armado para poder comparar la planilla contra el
    catálogo, aunque difieran en espacios, mayúsculas o sufijos.

    Reglas aplicadas:
      * Pasa a mayúsculas.
      * Elimina cualquier contenido entre paréntesis  ->  "MTF635-1 (S)" => "MTF6351"
      * Elimina espacios, guiones y guiones bajos      ->  "MTF 635-1"    => "MTF6351"

    Devuelve "" para valores vacíos / NaN.

    >>> normalizar_codigo_armado("MTF 635-1")
    'MTF6351'
    >>> normalizar_codigo_armado("MTF635-1 (S)")
    'MTF6351'
    """
    if codigo is None:
        return ""
    if isinstance(codigo, float) and np.isnan(codigo):
        return ""
    s = str(codigo).strip()
    if s == "" or s.lower() == "nan":
        return ""
    s = s.upper()
    s = re.sub(r"\(.*?\)", "", s)        # quita sufijos entre paréntesis: (S), (M)...
    s = re.sub(r"[\s\-_]", "", s)         # quita espacios, guiones, guiones bajos
    return s


def _normalizar_texto(texto) -> str:
    """Normaliza un texto a minúsculas sin acentos (para detectar encabezados)."""
    if texto is None:
        return ""
    s = str(texto)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.strip().lower()


# =====================================================================
#  1. CARGA DEL CATÁLOGO DE CANTIDADES POR ARMADO
# =====================================================================

@dataclass
class Catalogo:
    """
    Representa el catálogo de cantidades por armado ya procesado.

    Atributos
    ---------
    materiales : dict
        { codigo_armado_normalizado : { clave_material : multiplicador } }
    info_material : dict
        { clave_material : {"nombre":..., "codigo":..., "unidad":...} }
        Guarda la metadata de cada material para la exportación final.
    armados : dict
        { codigo_armado_normalizado : codigo_armado_original }
        Permite mostrar el código tal y como aparece en el catálogo.
    hojas : list
        Hojas del Excel que se fusionaron en este catálogo.
    """
    materiales: Dict[str, Dict[str, float]] = field(default_factory=dict)
    info_material: Dict[str, Dict[str, str]] = field(default_factory=dict)
    armados: Dict[str, str] = field(default_factory=dict)
    hojas: List[str] = field(default_factory=list)

    def armados_disponibles(self) -> List[str]:
        """Códigos de armado (originales) presentes en el catálogo, ordenados."""
        return sorted(self.armados.values())

    def __repr__(self) -> str:
        return (f"<Catalogo hojas={self.hojas} "
                f"armados={len(self.armados)} materiales={len(self.info_material)}>")


def _detectar_layout_hoja(crudo: pd.DataFrame) -> Dict[str, int]:
    """
    Detecta automáticamente la disposición de columnas/filas de una hoja del
    catálogo a partir de las dos primeras filas.

    Soporta los dos formatos observados:
      Formato A (AFINIA): Elementos | Codigo | Unidades | Armado...
      Formato B (ESSA):   Elementos | Codigo JDE |       | Armado...

    Devuelve un dict con:
        fila_codigos      -> índice de fila donde están los códigos de armado
        col_elemento      -> columna del nombre del material
        col_codigo        -> columna del código del material (o None)
        col_unidad        -> columna de la unidad (o None)
        col_inicio_armado -> primera columna con códigos de armado
    """
    fila0 = [_normalizar_texto(x) for x in crudo.iloc[0].tolist()]

    col_elemento = next((i for i, v in enumerate(fila0) if v.startswith("element")), 0)
    col_codigo = next((i for i, v in enumerate(fila0) if v.startswith("codigo")), None)
    col_unidad = next((i for i, v in enumerate(fila0) if v.startswith("unidad")), None)

    # Fallback si no se puede determinar el inicio real de armados (ver más abajo):
    # empezar tras la última columna de metadata conocida.
    ultima_meta = max([c for c in (col_elemento, col_codigo, col_unidad)
                       if c is not None])

    # IMPORTANTE: la celda de texto "Armado" en la fila 0 NO siempre está
    # alineada con la primera columna que realmente tiene un código de armado
    # en la fila 1 (en algunas hojas, p.ej. "ESSA (13.2 kV - con viento)",
    # Excel deja la celda "Armado" varias columnas más a la derecha del
    # primer código real). Usar esa posición como `col_inicio_armado` hace
    # que se pierdan las primeras columnas de armados y que el resto quede
    # desalineado (los multiplicadores terminan asociados a un código de
    # armado que no les corresponde).
    #
    # Por eso el inicio real de armados se determina a partir de la fila de
    # códigos (fila 1): es la primera columna, después de la metadata
    # conocida (elemento/codigo/unidad), que tiene un valor no vacío en la
    # fila 1.
    fila1 = crudo.iloc[1].tolist() if crudo.shape[0] > 1 else []
    col_inicio_armado = None
    for i in range(ultima_meta + 1, len(fila1)):
        val = fila1[i]
        if pd.notna(val) and str(val).strip() != "":
            col_inicio_armado = i
            break
    if col_inicio_armado is None:
        # Fallback: no se encontró ningún código en la fila 1 tras la
        # metadata; usar el criterio anterior basado en la celda "Armado".
        col_armado_marca = next((i for i, v in enumerate(fila0) if v.startswith("armado")), None)
        col_inicio_armado = col_armado_marca if col_armado_marca is not None else ultima_meta + 1

    # Los códigos de armado suelen estar en la fila 1 (segunda fila).
    fila_codigos = 1
    return {
        "fila_codigos": fila_codigos,
        "col_elemento": col_elemento,
        "col_codigo": col_codigo,
        "col_unidad": col_unidad,
        "col_inicio_armado": col_inicio_armado,
    }


def cargar_catalogo(ruta: str,
                    hojas: Optional[Sequence[str]] = None,
                    layout: Optional[Dict[str, int]] = None,
                    verbose: bool = True) -> Catalogo:
    """
    Lee el catálogo de cantidades por armado y lo convierte en un objeto
    `Catalogo` consultable.

    Parámetros
    ----------
    ruta : str
        Ruta al .xlsx del catálogo (Cantidades_de_postes.xlsx).
    hojas : lista de str, opcional
        Nombres de las hojas a usar. Si es None se usan TODAS las hojas que
        contengan la cabecera "Armado". Los armados de distintas hojas se
        fusionan; si un mismo armado aparece en varias hojas, gana la primera.
    layout : dict, opcional
        Fuerza la disposición de columnas (ver `_detectar_layout_hoja`). Si es
        None se detecta automáticamente por hoja.
    verbose : bool
        Imprime un resumen de la carga.

    Devuelve
    --------
    Catalogo
    """
    if not os.path.exists(ruta):
        raise FileNotFoundError(f"No se encontró el catálogo: {ruta}")

    xls = pd.ExcelFile(ruta)
    hojas_objetivo = list(hojas) if hojas else list(xls.sheet_names)

    catalogo = Catalogo()
    for hoja in hojas_objetivo:
        if hoja not in xls.sheet_names:
            if verbose:
                print(f"[catalogo] AVISO: hoja inexistente, se omite: {hoja!r}")
            continue

        crudo = pd.read_excel(xls, sheet_name=hoja, header=None)
        if crudo.shape[0] < 3:
            if verbose:
                print(f"[catalogo] AVISO: hoja sin datos suficientes: {hoja!r}")
            continue

        lay = layout or _detectar_layout_hoja(crudo)
        # Si la hoja no tiene marca "Armado" en fila 0, probablemente no es un
        # catálogo de armados (p.ej. la hoja resumen "Cantidades ESSA"); se omite.
        fila0 = [_normalizar_texto(x) for x in crudo.iloc[0].tolist()]
        if not any(v.startswith("armado") for v in fila0):
            if verbose:
                print(f"[catalogo] Hoja sin cabecera 'Armado', se omite: {hoja!r}")
            continue

        f_cod = lay["fila_codigos"]
        c_ini = lay["col_inicio_armado"]
        c_elem = lay["col_elemento"]
        c_codm = lay["col_codigo"]
        c_uni = lay["col_unidad"]

        # Mapa columna_excel -> codigo_armado para esta hoja.
        codigos_fila = crudo.iloc[f_cod]
        col_a_armado: Dict[int, str] = {}
        for col in range(c_ini, crudo.shape[1]):
            crudo_codigo = codigos_fila.iloc[col]
            if pd.isna(crudo_codigo) or str(crudo_codigo).strip() == "":
                continue
            col_a_armado[col] = str(crudo_codigo).strip()

        n_armados_hoja = 0
        # Filas de materiales: a partir de la fila siguiente a los códigos.
        for r in range(f_cod + 1, crudo.shape[0]):
            nombre = crudo.iloc[r, c_elem]
            if pd.isna(nombre) or str(nombre).strip() == "":
                continue
            nombre = str(nombre).strip()
            codigo_mat = ""
            if c_codm is not None and not pd.isna(crudo.iloc[r, c_codm]):
                codigo_mat = str(crudo.iloc[r, c_codm]).strip()
                # Limpia floats tipo "211274.0"
                codigo_mat = re.sub(r"\.0$", "", codigo_mat)
            unidad = ""
            if c_uni is not None and not pd.isna(crudo.iloc[r, c_uni]):
                unidad = str(crudo.iloc[r, c_uni]).strip()

            # Clave única del material: código si existe y es real, si no el nombre.
            #
            # IMPORTANTE: varios materiales del catálogo NO tienen código JDE y
            # esa ausencia se representa con el texto literal "-" (además de
            # "N/A"/"NaN"). Si se tratara "-" como un código real, decenas de
            # materiales DISTINTOS colisionarían bajo la misma clave "COD::-",
            # y el catálogo terminaría fusionándolos en un solo renglón: se
            # quedaría con el nombre del primero que se haya leído (de
            # cualquier hoja) pero sumando los multiplicadores de todos los
            # demás, atribuyendo cantidades de un material a armados que en
            # realidad correspondían a otro material completamente distinto.
            codigo_normalizado = codigo_mat.strip().upper()
            codigo_es_real = (
                codigo_normalizado != ""
                and codigo_normalizado not in ("N/A", "NAN", "-", "--", "S/C", "SC")
            )
            if codigo_es_real:
                clave_mat = f"COD::{codigo_mat}"
            else:
                clave_mat = f"NOM::{nombre.upper()}"

            if clave_mat not in catalogo.info_material:
                catalogo.info_material[clave_mat] = {
                    "nombre": nombre,
                    "codigo": codigo_mat,
                    "unidad": unidad,
                }

            for col, armado_orig in col_a_armado.items():
                valor = crudo.iloc[r, col]
                if pd.isna(valor):
                    continue
                try:
                    mult = float(valor)
                except (TypeError, ValueError):
                    continue
                if mult == 0:
                    continue
                armado_norm = normalizar_codigo_armado(armado_orig)
                if armado_norm == "":
                    continue
                catalogo.armados.setdefault(armado_norm, armado_orig)
                fila_mat = catalogo.materiales.setdefault(armado_norm, {})
                # Suma por si el mismo material aparece en varias filas del armado.
                fila_mat[clave_mat] = fila_mat.get(clave_mat, 0.0) + mult

        n_armados_hoja = len(col_a_armado)
        catalogo.hojas.append(hoja)
        if verbose:
            print(f"[catalogo] Hoja {hoja!r}: {n_armados_hoja} armados leídos.")

    if verbose:
        print(f"[catalogo] Total -> {len(catalogo.armados)} armados, "
              f"{len(catalogo.info_material)} materiales distintos.")
    if not catalogo.armados:
        raise ValueError(
            "El catálogo quedó vacío. Revisa el nombre de las hojas o el layout."
        )
    return catalogo


# =====================================================================
#  2-bis. NIVEL DE CONTAMINACIÓN DE LA LÍNEA (a partir de la 1ª coordenada)
# =====================================================================
#
# El nivel de contaminación ("Alto" = altamente contaminado / "Normal" = baja
# contaminación) se determina UNA sola vez para todo el proyecto/línea, a
# partir de la coordenada del primer poste válido de la planilla (columnas
# 'Topografía' -> 'X'/'Y'), usando el mapa de zona_contaminacion.py (región
# Caribe de Colombia). Ese único nivel se aplica luego a todos los postes del
# proyecto para decidir el aislador (ver sección 3-quater).
#
# Las coordenadas de la planilla ('Topografía' X/Y) están en un sistema de
# referencia PROYECTADO (no son lat/lon), por lo que hay que transformarlas
# antes de poder consultar el mapa de contaminación. Por defecto se asume
# MAGNA-SIRGAS / UTM zona 18N (EPSG:32618), que es razonable para la costa
# Caribe (Atlántico, Magdalena, Cesar, Guajira, Bolívar, Sucre, Córdoba), pero
# **debe verificarse contra el CRS real del proyecto** y ajustarse con el
# parámetro `epsg_planilla` si no corresponde (p.ej. otra franja UTM o un
# sistema MAGNA-SIRGAS de origen específico).

COL_TOPO_X_DEFAULT = ("Topografía", "X")
COL_TOPO_Y_DEFAULT = ("Topografía", "Y")

# CRS de las coordenadas de la planilla. AJUSTAR si el proyecto usa otro
# sistema (ver advertencia arriba).
EPSG_PLANILLA_DEFAULT = "EPSG:32618"  # MAGNA-SIRGAS / UTM zona 18N

_TRANSFORMER_CACHE: Dict[str, object] = {}


def _obtener_transformador(epsg_origen: str):
    """Crea (una sola vez, con caché) el transformador epsg_origen -> lat/lon."""
    if epsg_origen not in _TRANSFORMER_CACHE:
        try:
            from pyproj import Transformer
        except ImportError as e:
            raise ImportError(
                "Se requiere el paquete 'pyproj' para convertir las coordenadas "
                "de la planilla (proyectadas) a lat/lon. Instálalo con "
                "`pip install pyproj`."
            ) from e
        _TRANSFORMER_CACHE[epsg_origen] = Transformer.from_crs(
            epsg_origen, "EPSG:4326", always_xy=True)
    return _TRANSFORMER_CACHE[epsg_origen]


def coordenadas_planilla_a_latlon(
    x: float, y: float, epsg_origen: str = EPSG_PLANILLA_DEFAULT
) -> Tuple[float, float]:
    """
    Convierte una coordenada (x, y) de la planilla (CRS proyectado) a
    (lat, lon) en WGS84, para poder consultarla en zona_contaminacion.py.
    """
    transformador = _obtener_transformador(epsg_origen)
    lon, lat = transformador.transform(x, y)
    return float(lat), float(lon)


def determinar_nivel_contaminacion_linea(
    est_df: pd.DataFrame,
    col_x: Tuple[str, str] = COL_TOPO_X_DEFAULT,
    col_y: Tuple[str, str] = COL_TOPO_Y_DEFAULT,
    epsg_origen: str = EPSG_PLANILLA_DEFAULT,
    verbose: bool = True,
) -> dict:
    """
    Determina el nivel de contaminación de TODA la línea/proyecto, a partir de
    la coordenada del primer poste de la planilla con X/Y válidos.

    Devuelve un dict con:
        nivel            : 'Alto', 'Normal' o None si no se pudo determinar
        lat, lon          : coordenada usada (ya convertida a WGS84)
        x, y              : coordenada original de la planilla
        mensaje           : explicación cuando nivel es None
    """
    import zona_contaminacion as zc

    resultado = {"nivel": None, "lat": None, "lon": None, "x": None, "y": None,
                 "mensaje": None}

    if col_x not in est_df.columns or col_y not in est_df.columns:
        resultado["mensaje"] = (
            f"No se encontraron las columnas de coordenadas {col_x!r}/{col_y!r} "
            "en la planilla."
        )
        if verbose:
            print(f"[contaminacion] AVISO: {resultado['mensaje']}")
        return resultado

    fila_valida = None
    for _, fila in est_df.iterrows():
        x, y = fila.get(col_x), fila.get(col_y)
        if pd.notna(x) and pd.notna(y):
            fila_valida = (float(x), float(y))
            break

    if fila_valida is None:
        resultado["mensaje"] = "Ningún poste de la planilla tiene coordenadas X/Y válidas."
        if verbose:
            print(f"[contaminacion] AVISO: {resultado['mensaje']}")
        return resultado

    x, y = fila_valida
    lat, lon = coordenadas_planilla_a_latlon(x, y, epsg_origen=epsg_origen)
    r = zc.obtener_nivel_contaminacion(lat, lon)

    resultado.update({"nivel": r["nivel"], "lat": lat, "lon": lon, "x": x, "y": y,
                       "mensaje": r.get("mensaje")})

    if verbose:
        if r["nivel"] is None:
            print(f"[contaminacion] No se pudo determinar el nivel: {r.get('mensaje')}")
        else:
            print(f"[contaminacion] Nivel de contaminación de la línea: {r['nivel']} "
                  f"(1er poste -> lat={lat:.4f}, lon={lon:.4f}, "
                  f"confianza={r.get('confianza')})")

    return resultado


# =====================================================================
#  3. EXTRACCIÓN DE LOS ARMADOS DE CADA POSTE EN LA PLANILLA
# =====================================================================

# Columnas (multi-índice) por defecto en las planillas PlanillaEstTotal*.XLS.
COLUMNAS_ARMADO_DEFAULT = [
    ("Armado Primario",   "Primario1"),
    ("Armado Primario",   "Primario2"),
    ("Armado Secundario", "Secundario1"),
    ("Armado Secundario", "Secundario2"),
]
COL_NOMBRE_DEFAULT = ("Identificación", "Nombre Est.")
COL_NRUTA_DEFAULT = ("Identificación", "N° Est.")
COL_DERIVACION_DEFAULT = ("Identificación", "Derivación")
COL_TIPO_SOPORTE_DEFAULT = ("Estructura", "Tipo Soporte")
COL_TIPO_PAT_DEFAULT = ("Estructura", "Tipo PAT")

# Columnas de retenidas en el grupo "Estructura": el nombre de columna ES el
# código de armado a buscar en el catálogo (p.ej. "RT003"), y el VALOR de la
# celda es la cantidad de esa retenida instalada en ese poste (0, 1, 2, ...).
COLUMNAS_RETENIDA_DEFAULT = [
    ("Estructura", "RT001"),
    ("Estructura", "RT002"),
    ("Estructura", "RT003"),
    ("Estructura", "RT004"),
    ("Estructura", "RT005"),
    ("Estructura", "RT006"),
    ("Estructura", "RT002-RS"),
    ("Estructura", "RT003-RS"),
    ("Estructura", "RT Existente-BT"),
    ("Estructura", "RT Existente-MT"),
]


COL_CONDUCTOR_PRINCIPAL1_DEFAULT = ("Conductor Principal1", "Tipo Conductor")
COL_CONDUCTOR_PRINCIPAL2_DEFAULT = ("Conductor Principal2", "Tipo Conductor")

# Valores que en la columna 'Tipo Conductor' significan "no aplica".
_VALORES_VACIOS_CONDUCTOR = {"", "nan", "none", "-", "n/a", "na"}


# =====================================================================
#  3-ter. AJUSTE DE FASE (reemplazo del "" por el calibre del conductor)
# =====================================================================
#
# Algunos materiales del catálogo (p.ej. "GRAPA DE RETENCION RECTA "" ")
# tienen el texto literal "" en su nombre en lugar de un calibre concreto,
# porque ese material depende del calibre del conductor de fase instalado
# en el poste. Esta sección resuelve ese calibre a partir de la columna
# 'Tipo Conductor' de la planilla y lo usa para reemplazar el "" cuando se
# calculan las cantidades de material de cada armado.

# Materiales de conductor reconocidos dentro del texto de 'Tipo Conductor'.
MATERIALES_CONDUCTOR = ("AAAC", "ACSR")

# Marcador literal que se reemplaza por el calibre dentro del nombre del
# material, p.ej. 'GRAPA DE RETENCION RECTA "" ' -> 'GRAPA DE RETENCION RECTA 1/0 AWG'.
MARCADOR_CALIBRE_FASE = '""'

# Regex que reconoce el patrón "#xSM..." (p.ej. "1xSM34.5-3x1/0ACSR / Al7N8"
# o "3xSM..."), usado para distinguir, en un 'Tipo Conductor' con dos cables
# unidos por "+", cuál de los dos lados corresponde al cable de guarda/mensajero
# tipo "SM..." y cuál al conductor de fase simple (ACSR/AAAC "suelto").
_RE_LADO_SM = re.compile(r"[13]\s*X\s*SM", re.IGNORECASE)

# Armados cuyo conductor de fase asociado, cuando el 'Tipo Conductor' tiene
# dos cables unidos por "+", es el lado "#xSM..." (y no el lado "suelto").
# Ver `_seleccionar_lado_conductor`.
_RE_ARMADO_LADO_SM = re.compile(r"^MTF[67]\d\d-\d", re.IGNORECASE)


def _seleccionar_lado_conductor(tipo_conductor: str, armado: Optional[str]) -> str:
    """
    Cuando 'Tipo Conductor' describe DOS cables distintos unidos por "+"
    (p.ej. "1xACSR 1/0 AWG+1xSM34.5-3x1/0ACSR / Al7N8"), decide cuál de los
    dos lados usar para extraer el calibre, según el código de `armado`:

      * Si `armado` es de la forma "MTF6XX-X" o "MTF7XX-X" (ver
        `_RE_ARMADO_LADO_SM`): se toma el lado que contiene el patrón
        "#xSM..." (# = 1 o 3). Si por algún motivo NINGÚN lado tiene ese
        patrón, se toma el lado izquierdo por defecto.
      * En cualquier otro caso (otro tipo de armado, o `armado` es None):
        se toma el lado que NO contiene "#xSM...". Si ambos o ninguno lo
        tienen, se deja el string sin dividir (se procesa completo, igual
        que antes de este ajuste).

    Si `tipo_conductor` no contiene "+", se devuelve sin cambios.
    """
    if "+" not in tipo_conductor:
        return tipo_conductor

    lados = tipo_conductor.split("+")
    if len(lados) != 2:
        # Más de un "+": caso no contemplado, se deja el string completo.
        return tipo_conductor

    izq, der = lados[0], lados[1]
    izq_es_sm = bool(_RE_LADO_SM.search(izq))
    der_es_sm = bool(_RE_LADO_SM.search(der))

    quiere_lado_sm = bool(armado) and bool(_RE_ARMADO_LADO_SM.match(str(armado).strip()))

    if quiere_lado_sm:
        if izq_es_sm and not der_es_sm:
            return izq
        if der_es_sm and not izq_es_sm:
            return der
        # Ninguno (o ambos) tienen el patrón #xSM: por defecto, lado izquierdo.
        return izq

    # Se quiere el lado que NO es "#xSM..."
    if izq_es_sm and not der_es_sm:
        return der
    if der_es_sm and not izq_es_sm:
        return izq
    # Ambiguo (ambos o ninguno son "#xSM..."): se deja el string completo.
    return tipo_conductor


def extraer_calibre_conductor(tipo_conductor, armado: Optional[str] = None) -> Optional[str]:
    """
    Extrae el calibre de un texto de 'Tipo Conductor' (columnas
    'Conductor Principal1'/'Conductor Principal2' de la planilla).

    Si el texto describe dos cables unidos por "+" (p.ej. un conductor de
    fase simple junto con un cable tipo "SM..."), primero se selecciona el
    lado correcto según el código de `armado` (ver
    `_seleccionar_lado_conductor`) antes de aplicar las reglas de extracción.

    Sobre el texto ya reducido a un solo cable, se buscan los materiales
    conocidos (AAAC, ACSR):

      1. Si el material aparece justo al inicio del string, el calibre son
         las siguientes 2 palabras que le siguen al material.
             "AAAC 123,3 kcmil"  ->  "123,3 kcmil"
             "ACSR 1/0 AWG"      ->  "1/0 AWG"

      2. Si el material NO aparece al inicio, el calibre es el texto escrito
         entre la última "x" antes del material y el material mismo. A ese
         texto se le agrega la unidad: "AWG" si contiene "/", o "kcmil" si es
         solo un número.
             "SM34.5-3x1/0ACSR / Al7N8"        ->  "1/0 AWG"
             'SM34.5-3x1/0ACSR / EHS 3/8"'     ->  "1/0 AWG"

    Devuelve None si el valor está vacío/"-"/NaN, o si no se reconoce ningún
    material dentro del texto (no se puede determinar el calibre).
    """
    if tipo_conductor is None:
        return None
    if isinstance(tipo_conductor, float) and np.isnan(tipo_conductor):
        return None
    s = str(tipo_conductor).strip()
    if s == "" or s.lower() in _VALORES_VACIOS_CONDUCTOR:
        return None

    s = _seleccionar_lado_conductor(s, armado).strip()
    if s == "":
        return None

    s_upper = s.upper()
    idx_material = None
    material = None
    for mat in MATERIALES_CONDUCTOR:
        idx = s_upper.find(mat)
        if idx != -1 and (idx_material is None or idx < idx_material):
            idx_material = idx
            material = mat
    if idx_material is None:
        return None

    antes_material = s[:idx_material]

    # Caso 1: el material está al inicio, o lo único que lo precede es un
    # simple contador de cables tipo "1x"/"3x" (p.ej. "1xACSR 1/0 AWG"), que
    # no es un calibre sino la cantidad de conductores de ese tipo -> el
    # calibre son las 2 palabras que siguen al material.
    if idx_material == 0 or re.fullmatch(r"[13]\s*x\s*", antes_material, re.IGNORECASE):
        resto = s[idx_material + len(material):].strip()
        palabras = resto.split()
        if not palabras:
            return None
        calibre = " ".join(palabras[:2]).strip()
        return calibre or None

    # Caso 2: material no está al inicio (y lo que lo precede no es un simple
    # contador) -> calibre = texto entre la última "x" antes del material y
    # el material, más la unidad correspondiente.
    pos_x = antes_material.lower().rfind("x")
    if pos_x == -1:
        return None
    crudo = antes_material[pos_x + 1:].strip()
    if not crudo:
        return None
    if "/" in crudo:
        return f"{crudo} AWG"
    return f"{crudo} kcmil"


# =====================================================================
#  3-quater. SELECCIÓN DE AISLADOR (contaminación + nivel de aislamiento +
#            tipo de conductor)
# =====================================================================
#
# Reglas de selección (confirmadas con el usuario):
#
#   1) Tipo de conductor (forrado / desnudo), por código de armado:
#        - "MTF..." (tiene F después de MT)  -> conductor FORRADO
#        - "MT..."  (sin F después de MT)    -> conductor DESNUDO
#      Esta verificación se hace de forma independiente para el conductor
#      Principal1 (armados Primario1/Primario2) y Principal2 (armados
#      Secundario1/Secundario2): es posible tener, por ejemplo, Principal1
#      forrado (MTF) y Principal2 desnudo (MT) en el mismo poste.
#
#   2) Nivel de aislamiento (13.2 kV o 34.5 kV): lo indica el último dígito
#      del código de armado (después del guion), ya diferenciado en la
#      planilla:
#        - termina en 1  ->  13.2 kV
#        - termina en 2  ->  34.5 kV
#
#   3) Aislador a usar:
#        - Conductor DESNUDO -> siempre "AISLADOR PORCELANA TIPO POSTE X kV"
#          (X = 13.2 o 34.5 según el nivel de aislamiento), sin importar la
#          contaminación.
#        - Conductor FORRADO -> depende de la contaminación de la línea
#          (determinada una sola vez para todo el proyecto, ver
#          `determinar_nivel_contaminacion_linea`):
#            * Contaminación NORMAL (baja)  -> AISLADOR PORCELANA TIPO POSTE X kV
#            * Contaminación ALTA           -> AISLADOR COMPUESTO HIBRIDO 13,2 kV
#                                               (si nivel = 13.2 kV), o
#                                               AISLADOR LINEPOST 66KV 1143mm
#                                               (ANSI 57-5) (si nivel = 34.5 kV)

# Regex para extraer el sufijo numérico final de un código de armado, p.ej.
# "MTF635-1 (S)" -> "1", "MT331-2" -> "2".
_RE_SUFIJO_ARMADO = re.compile(r"-(\d+)\s*(?:\(.*\))?\s*$")


def es_conductor_forrado(codigo_armado) -> Optional[bool]:
    """
    Indica si el conductor asociado a un código de armado es forrado o
    desnudo, según el prefijo del código:

        "MTF..."  ->  True   (forrado)
        "MT..."   ->  False  (desnudo, sin F después de MT)

    Devuelve None si el código no tiene ese prefijo (p.ej. una retenida
    "RT00X", donde el concepto de forrado/desnudo no aplica).

    >>> es_conductor_forrado("MTF635-1 (S)")
    True
    >>> es_conductor_forrado("MT331-2")
    False
    >>> es_conductor_forrado("RT003")
    """
    if codigo_armado is None:
        return None
    s = str(codigo_armado).strip().upper()
    if s.startswith("MTF"):
        return True
    if s.startswith("MT"):
        return False
    return None


def nivel_aislamiento_armado(codigo_armado) -> Optional[str]:
    """
    Devuelve el nivel de aislamiento ('13.2' o '34.5', en kV) según el último
    dígito del código de armado (el que sigue al guion):

        termina en 1  ->  '13.2'
        termina en 2  ->  '34.5'

    Devuelve None si no se reconoce un sufijo numérico terminado en 1 o 2.

    >>> nivel_aislamiento_armado("MTF731-1")
    '13.2'
    >>> nivel_aislamiento_armado("MT331-2")
    '34.5'
    """
    if codigo_armado is None:
        return None
    s = str(codigo_armado).strip()
    m = _RE_SUFIJO_ARMADO.search(s)
    if not m:
        return None
    ultimo_digito = m.group(1)[-1]
    if ultimo_digito == "1":
        return "13.2"
    if ultimo_digito == "2":
        return "34.5"
    return None


def determinar_aislador(
    forrado: Optional[bool],
    nivel_kv: Optional[str],
    nivel_contaminacion: Optional[str],
) -> Optional[str]:
    """
    Determina el aislador a usar para un armado, según:

        forrado             : True (forrado) / False (desnudo) / None (n/a)
        nivel_kv            : '13.2' o '34.5' (ver `nivel_aislamiento_armado`)
        nivel_contaminacion : 'Alto' o 'Normal' (ver
                               `determinar_nivel_contaminacion_linea`), solo
                               relevante si el conductor es forrado.

    Devuelve None si falta información suficiente para decidir (p.ej. no se
    reconoció el nivel de aislamiento, o -siendo forrado- no se pudo
    determinar la contaminación de la línea).
    """
    if forrado is None or nivel_kv is None:
        return None

    if forrado is False:
        # Conductor desnudo: siempre porcelana tipo poste, del nivel que
        # corresponda, sin importar la contaminación.
        return f"AISLADOR PORCELANA TIPO POSTE {nivel_kv} kV"

    # Conductor forrado: depende de la contaminación de la línea.
    if nivel_contaminacion == "Normal":
        return f"AISLADOR PORCELANA TIPO POSTE {nivel_kv} kV"
    if nivel_contaminacion == "Alto":
        if nivel_kv == "13.2":
            return "AISLADOR COMPUESTO HIBRIDO 13,2 kV"
        if nivel_kv == "34.5":
            return "AISLADOR LINEPOST 66KV 1143mm (ANSI 57-5)"
        return None
    return None


# Familias de aislador "alternativas" que aparecen en el catálogo para un
# mismo armado (p.ej. la hoja "AFINIAAIR-E (Forradas - 13,2 kV" trae, para el
# mismo armado, tanto "AISLADOR PORCELANA TIPO POSTE 13,2 kV (ANSI-57-1)"
# como "AISLADOR COMPUESTO HIBRIDO 13,2 kV.", cada una con su propia
# cantidad). El catálogo NO decide cuál de las dos corresponde: eso lo decide
# `determinar_aislador` según la contaminación. `_familia_aislador` permite
# reconocer, tanto en el nombre de un material del catálogo como en el
# resultado de `determinar_aislador`, a cuál de esas familias pertenece, para
# poder quedarnos únicamente con la fila que aplica y descartar la(s) otra(s).
#
# IMPORTANTE: al revisar Cantidades_de_postes.xlsx, el catálogo NO tiene
# ningún renglón "AISLADOR LINEPOST 66KV..." (ni en las hojas de 34,5 kV
# forradas ni en las desnudas): solo existe la alternativa PORCELANA para
# 34,5 kV. Es decir, hoy no hay forma de cuantificar el caso "forrado + alta
# contaminación + 34,5 kV" con este catálogo. El código lo señala en el
# reporte 'aislador_sin_correspondencia' en vez de inventar una cantidad;
# hay que agregar esa fila al catálogo (con su cantidad por armado) para que
# se pueda calcular.
_FAMILIAS_AISLADOR = {
    "PORCELANA": "aislador porcelana tipo poste",
    "COMPUESTO_HIBRIDO": "aislador compuesto hibrido",
    "LINEPOST": "aislador linepost",
}


def _familia_aislador(nombre: Optional[str]) -> Optional[str]:
    """
    Clasifica un nombre de material (del catálogo) o un aislador ya
    determinado (ver `determinar_aislador`) en una de las familias
    'PORCELANA' / 'COMPUESTO_HIBRIDO' / 'LINEPOST'.

    Devuelve None si el nombre no corresponde a ninguna de estas familias
    (p.ej. "AISLADOR PIN PARA RED FORRADA" o "AISLADOR COMPUESTO TIPO
    SUSPENSION ANSI DS 15 70 kN", que son materiales fijos que no dependen de
    la contaminación y por lo tanto no se filtran).
    """
    if not nombre:
        return None
    nombre_norm = _normalizar_texto(nombre)
    for familia, patron in _FAMILIAS_AISLADOR.items():
        if patron in nombre_norm:
            return familia
    return None


def _principal_para_tipo_armado(tipo_armado: str) -> Optional[int]:
    """
    Indica a qué conductor principal (1 o 2) se asocia un armado según su
    columna de origen: los armados 'Primario1'/'Primario2' se asocian al
    conductor de 'Conductor Principal1' y los 'Secundario1'/'Secundario2' al
    de 'Conductor Principal2'. Devuelve None para cualquier otro tipo (p.ej.
    retenidas), que no tienen conductor de fase asociado.
    """
    if not tipo_armado:
        return None
    t = str(tipo_armado).strip().lower()
    if t.startswith("primario"):
        return 1
    if t.startswith("secundario"):
        return 2
    return None


def ajustar_nombre_material_fase(nombre: str, calibre: Optional[str]) -> str:
    """
    Reemplaza el marcador "" dentro de `nombre` por `calibre`.

    Si `nombre` no contiene el marcador, se devuelve sin cambios. Si lo
    contiene pero `calibre` es None (no se pudo determinar el conductor de
    fase), también se devuelve sin cambios, dejando el "" visible para que
    quede en evidencia en la exportación.
    """
    if nombre is None or MARCADOR_CALIBRE_FASE not in nombre:
        return nombre
    if not isinstance(calibre, str) or not calibre.strip():
        return nombre
    ajustado = nombre.replace(MARCADOR_CALIBRE_FASE, calibre)
    ajustado = re.sub(r"\s+", " ", ajustado).strip()
    return ajustado


# =====================================================================
#  3-quinquies. CANTIDADES TOTALES DE CABLE (conductores) POR VANO
# =====================================================================
#
# Esta sección determina, a partir de 'Tipo Conductor' (Conductor Principal1
# / Conductor Principal2), 'Vano Adelante' y los armados Primario1/Primario2
# / Secundario1/Secundario2, la cantidad TOTAL de cada cable (fase,
# mensajero, o cable "normal") que hay en toda la línea.
#
# Ver docstring de `extraer_longitudes_cable_planilla` para el detalle
# completo de las reglas (identificación de red compacta/normal,
# multiplicadores por armado, definición de vano, etc.).

# Columna con la cantidad en metros del vano ENTRE un poste y el siguiente
# (en la misma ruta/derivación).
COL_VANO_ADELANTE_DEFAULT = ("Topografía", "Vano Adelante")

# Prefijo que identifica una red tipo "compacta" (ver
# `identificar_cables_conductor`): "SM" + nivel de tensión + "-" + cantidad
# de fases + "x" + calibre+tipo de fase + "/" + cable mensajero.
#
# IMPORTANTE: el separador "/" entre el bloque de fase y el mensajero debe
# tener espacios a ambos lados (" / "). Esto es necesario porque el propio
# calibre de la fase puede contener una "/" SIN espacios (p.ej. "1/0ACSR"
# en "SM34.5-3x1/0ACSR / Al7N8"); si se aceptara cualquier "/" como
# separador, la primera "/" (la del calibre) se confundiría con la del
# mensajero y el calibre quedaría truncado.
_RE_RED_COMPACTA = re.compile(
    r"^SM\s*([\d.,]+)\s*-\s*(\d+)\s*x\s*(.+?)\s+/\s+(.+)$", re.IGNORECASE
)

# Dentro del bloque "calibre+tipo" de una red compacta (p.ej. "63AAAC" o
# "1/0ACSR"), separa el calibre (números, comas, puntos, "/") del tipo de
# cable (letras) que le sigue sin espacio.
_RE_CALIBRE_TIPO_FASE = re.compile(r"^([\d.,/]+)\s*([A-Za-zÀ-ÿ]+)\s*$")

# Multiplicador base (sin duplicar) de una red compacta: 1 mensajero y 3
# fases por circuito.
_MENSAJEROS_BASE_COMPACTA = 1
_FASES_BASE_COMPACTA = 3

# Patrón de armado que indica que el circuito está duplicado (2 circuitos
# autosoportados, ver `identificar_poste`/`numero_fases` en
# funciones_mecanicas.py: primer dígito numérico = 7): "MTF7XX-X". En ese
# caso los multiplicadores de mensajero/fase de la red compacta se
# multiplican x2 (2 mensajeros, 6 fases).
_RE_ARMADO_DOBLE_CIRCUITO = re.compile(r"^MTF7\d\d-\d", re.IGNORECASE)


def es_armado_doble_circuito(codigo_armado) -> bool:
    """
    Indica si un código de armado corresponde a un poste autosoportado de
    2 circuitos (patrón "MTF7XX-X"), caso en el que los multiplicadores de
    mensajero/fase de una red compacta se duplican (2 mensajeros, 6 fases
    en vez de 1 mensajero y 3 fases).

    >>> es_armado_doble_circuito("MTF731-1")
    True
    >>> es_armado_doble_circuito("MTF631-1")
    False
    >>> es_armado_doble_circuito("MT331-2")
    False
    """
    if codigo_armado is None:
        return False
    s = str(codigo_armado).strip().upper()
    s = re.sub(r"\s+", "", s)
    return bool(_RE_ARMADO_DOBLE_CIRCUITO.match(s))


def identificar_cables_conductor(tipo_conductor) -> Optional[List[dict]]:
    """
    Identifica el/los cable(s) que describe un texto de 'Tipo Conductor'
    (columnas 'Conductor Principal1'/'Conductor Principal2' de la planilla),
    y determina si la red es COMPACTA o NORMAL:

    1. Red COMPACTA: el texto empieza con el prefijo "SM", seguido del nivel
       de tensión, un "-", la cantidad de conductores de fase, una "x", el
       calibre y tipo de la fase, una "/" y finalmente el cable mensajero.
       Ejemplo:
           "SM13.2-3x63AAAC / Aluminium Clad Steel 7 Nº 8"
       produce DOS cables:
         - Fase     = "63 AAAC 13.2 kV"       (calibre + tipo + nivel kV)
         - Mensajero = "Aluminium Clad Steel 7 Nº 8"  (texto completo tras "/")

    2. Red NORMAL: no tiene el prefijo "SM" de red compacta. El cable a
       incluir es el nombre completo tal cual aparece, p.ej. "ACSR 1/0 AWG".
       Produce UN solo cable.

    Devuelve una lista de dicts, cada uno con:
        {"nombre": <nombre del cable>, "rol": "fase" | "mensajero" | "normal"}

    El orden de los multiplicadores (fases=3, mensajero=1) se aplica después,
    en `calcular_longitudes_cable` / `extraer_longitudes_cable_planilla`;
    esta función solo IDENTIFICA los cables, no calcula cantidades.

    Devuelve None si el texto está vacío/"-"/NaN (no hay conductor).
    """
    if tipo_conductor is None:
        return None
    if isinstance(tipo_conductor, float) and np.isnan(tipo_conductor):
        return None
    s = str(tipo_conductor).strip()
    if s == "" or s.lower() in _VALORES_VACIOS_CONDUCTOR:
        return None

    m = _RE_RED_COMPACTA.match(s)
    if m:
        nivel_kv, _cant_fases, calibre_tipo, mensajero = m.groups()
        calibre_tipo = calibre_tipo.strip()
        mensajero = mensajero.strip()
        nivel_kv = nivel_kv.strip()

        m_ct = _RE_CALIBRE_TIPO_FASE.match(calibre_tipo)
        if m_ct:
            calibre, tipo = m_ct.group(1).strip(), m_ct.group(2).strip()
            nombre_fase = f"{calibre} {tipo} {nivel_kv} kV"
        else:
            # No se pudo separar calibre/tipo: se deja el bloque completo
            # seguido del nivel de tensión, en vez de fallar.
            nombre_fase = f"{calibre_tipo} {nivel_kv} kV"

        return [
            {"nombre": nombre_fase, "rol": "fase"},
            {"nombre": mensajero, "rol": "mensajero"},
        ]

    # Red normal: el nombre completo tal cual aparece es el cable.
    return [{"nombre": s, "rol": "normal"}]


def _armado_para_conductor_principal(
    valor_primero: Optional[str], valor_segundo: Optional[str]
) -> Optional[str]:
    """
    Determina qué código de armado usar para calcular el multiplicador de
    fase/mensajero de un conductor principal, a partir de sus dos posibles
    columnas de armado (p.ej. Primario1/Primario2 para Conductor Principal1,
    o Secundario1/Secundario2 para Conductor Principal2):

        * Si solo uno de los dos tiene valor, se usa ese.
        * Si ambos tienen valor, se toma el PRIMERO por defecto (ver
          especificación: "en caso de que estén ambos se toma el 1 por
          defecto").
        * Si ninguno tiene valor, devuelve None.
    """
    def _valido(v) -> bool:
        return v is not None and pd.notna(v) and str(v).strip() != ""

    if _valido(valor_primero):
        return str(valor_primero).strip()
    if _valido(valor_segundo):
        return str(valor_segundo).strip()
    return None


def calcular_longitudes_cable_poste(
    tipo_conductor_principal1,
    tipo_conductor_principal2,
    armado_primario1,
    armado_primario2,
    armado_secundario1,
    armado_secundario2,
    vano_adelante: Optional[float],
) -> List[dict]:
    """
    Calcula el aporte de cable (en metros) de UN poste hacia su vano
    adelante, tanto para el conductor primario (Conductor Principal1) como
    para el secundario (Conductor Principal2).

    Reglas (ver especificación funcional completa):

      1. El armado a usar para el multiplicador de Conductor Principal1 es
         el de Primario1 (o Primario2 si Primario1 está vacío; si ambos
         están, se usa Primario1 por defecto). Análogamente, el armado para
         Conductor Principal2 es el de Secundario1 (o Secundario2, con el
         mismo criterio de "1 por defecto").

      2. Si 'Vano Adelante' es None/NaN/<=0, o el 'Tipo Conductor'
         correspondiente está vacío, ese conductor no aporta cable (lista
         vacía para ese lado).

      3. Red COMPACTA (ver `identificar_cables_conductor`): el vano aporta
         siempre 1 mensajero y 3 fases; si el armado usado en ese lado es de
         la forma "MTF7XX-X" (ver `es_armado_doble_circuito`), el aporte se
         duplica (2 mensajeros, 6 fases).

      4. Red NORMAL: el vano aporta tantas veces el cable como fases indique
         `funciones_mecanicas.numero_fases` (importado de forma perezosa)
         aplicado al armado usado en ese lado.

    Devuelve una lista de dicts:
        [{"nombre": <cable>, "metros": <float>, "lado": "principal1"|"principal2"}, ...]

    (uno por cada cable distinto que aporta ese poste; puede tener 0, 1 o 2
    entradas por lado, según si hay fase+mensajero o un solo cable normal).
    """
    aportes: List[dict] = []

    if vano_adelante is None or pd.isna(vano_adelante) or float(vano_adelante) <= 0:
        return aportes
    vano = float(vano_adelante)

    def _procesar_lado(tipo_conductor, armado_a, armado_b, etiqueta_lado):
        cables = identificar_cables_conductor(tipo_conductor)
        if not cables:
            return
        armado_usado = _armado_para_conductor_principal(armado_a, armado_b)

        for cable in cables:
            nombre, rol = cable["nombre"], cable["rol"]
            if rol == "normal":
                # Red normal: se multiplica por el número de fases del
                # armado usado en este lado.
                n_fases = _numero_fases_armado(armado_usado)
                if n_fases is None:
                    # No se pudo determinar el número de fases: se reporta
                    # tal cual (factor 1) para no perder el cable, pero
                    # queda visible en el detalle para poder auditar.
                    n_fases = 1
                metros = vano * n_fases
            else:
                # Red compacta: multiplicadores base (fase=3, mensajero=1),
                # duplicados si el armado usado es "MTF7XX-X".
                factor = 2 if es_armado_doble_circuito(armado_usado) else 1
                base = _FASES_BASE_COMPACTA if rol == "fase" else _MENSAJEROS_BASE_COMPACTA
                metros = vano * base * factor

            aportes.append({
                "nombre": nombre,
                "rol": rol,
                "metros": metros,
                "lado": etiqueta_lado,
                "armado_usado": armado_usado,
            })

    _procesar_lado(tipo_conductor_principal1, armado_primario1, armado_primario2, "principal1")
    _procesar_lado(tipo_conductor_principal2, armado_secundario1, armado_secundario2, "principal2")

    return aportes


_FUNCIONES_MECANICAS_MODULO = None


def _numero_fases_armado(codigo_armado: Optional[str]) -> Optional[int]:
    """
    Determina el número de fases de un código de armado de red NORMAL
    (p.ej. "MT331-1" -> 3 fases, "MT321-1" -> 2 fases), reutilizando la
    misma lógica que `funciones_mecanicas.numero_fases` (segundo dígito
    numérico del código de armado).

    Se importa `funciones_mecanicas` de forma perezosa (y se cachea el
    módulo) para no romper este archivo si esa dependencia no está
    disponible en algún entorno; en ese caso se recalcula la misma regex
    localmente como respaldo.

    Devuelve None si no se reconoce el patrón "MT(F?)###-#" en el código.
    """
    if codigo_armado is None:
        return None
    codigo = str(codigo_armado).strip()
    if codigo == "" or codigo.lower() in ("nan", "-", "none"):
        return None

    global _FUNCIONES_MECANICAS_MODULO
    if _FUNCIONES_MECANICAS_MODULO is None:
        try:
            import funciones_mecanicas as _fm
            _FUNCIONES_MECANICAS_MODULO = _fm
        except ImportError:
            _FUNCIONES_MECANICAS_MODULO = False  # marca "no disponible"

    codigo_norm = codigo.replace(" ", "")
    if not re.search(r"MT\s*F?\s*\d{3}-\d", codigo_norm, re.IGNORECASE) and \
       re.search(r"MT\s*F?\s*\d{3}$", codigo_norm, re.IGNORECASE):
        codigo_norm = codigo_norm + "-1"
    match = re.search(r"MT(F?)(\d{3})-(\d)", codigo_norm, re.IGNORECASE)
    if not match:
        return None
    return int(match.group(2)[1])


COL_VANO_ADELANTE_DEFAULT_KEY = COL_VANO_ADELANTE_DEFAULT


def extraer_longitudes_cable_planilla(
    est_df: pd.DataFrame,
    col_nombre: Tuple[str, str] = COL_NOMBRE_DEFAULT,
    col_nruta: Optional[Tuple[str, str]] = COL_NRUTA_DEFAULT,
    col_derivacion: Optional[Tuple[str, str]] = COL_DERIVACION_DEFAULT,
    col_n_general: Tuple[str, str] = ("Identificación", "N°"),
    col_conductor_principal1: Tuple[str, str] = COL_CONDUCTOR_PRINCIPAL1_DEFAULT,
    col_conductor_principal2: Tuple[str, str] = COL_CONDUCTOR_PRINCIPAL2_DEFAULT,
    col_armado_primario1: Tuple[str, str] = COLUMNAS_ARMADO_DEFAULT[0],
    col_armado_primario2: Tuple[str, str] = COLUMNAS_ARMADO_DEFAULT[1],
    col_armado_secundario1: Tuple[str, str] = COLUMNAS_ARMADO_DEFAULT[2],
    col_armado_secundario2: Tuple[str, str] = COLUMNAS_ARMADO_DEFAULT[3],
    col_vano_adelante: Tuple[str, str] = COL_VANO_ADELANTE_DEFAULT,
) -> pd.DataFrame:
    """
    Calcula la longitud total de cable (por tipo de cable) de TODA la línea,
    a partir de 'Tipo Conductor' (Conductor Principal1/2), los armados
    (Primario1/2, Secundario1/2) y 'Vano Adelante' de cada poste.

    Un VANO es la distancia entre un poste y el siguiente EN LA MISMA RUTA
    (misma combinación de 'N°'+'Derivación'); por eso la planilla se agrupa
    por esa combinación y, dentro de cada grupo, se procesa en el orden de
    'N° Est.' -- 'Vano Adelante' de un poste es el vano hacia el poste
    siguiente dentro de ese mismo grupo.

    Ver `calcular_longitudes_cable_poste` para el detalle de las reglas de
    identificación de cables (compacta/normal) y sus multiplicadores.

    Devuelve un DataFrame "largo" con una fila por cada aporte de cable de
    cada poste (para trazabilidad), con columnas:

        nombre_poste | derivacion | n_ruta | n_est | lado | rol |
        cable | armado_usado | vano_adelante | metros

    donde `cable` es el nombre del cable (fase/mensajero/normal), `lado` es
    "principal1" o "principal2", y `rol` es "fase"/"mensajero"/"normal".

    La suma total por cable (para las cantidades finales exportadas) se
    obtiene agregando esta tabla por `cable` (ver `sumar_totales_cable`).
    """
    columnas_salida = ["nombre_poste", "derivacion", "n_ruta", "n_est", "lado",
                        "rol", "cable", "armado_usado", "vano_adelante", "metros"]

    faltantes_col = [c for c in (col_conductor_principal1, col_conductor_principal2,
                                  col_vano_adelante) if c not in est_df.columns]
    if faltantes_col:
        raise KeyError(f"No se encontraron las columnas {faltantes_col!r} en la planilla.")

    # --- Determinar la clave de agrupación (ruta) ---
    tiene_n_general = col_n_general in est_df.columns
    tiene_derivacion = col_derivacion is not None and col_derivacion in est_df.columns
    tiene_n_est = col_nruta is not None and col_nruta in est_df.columns

    df = est_df.copy()
    df["_orden_original"] = range(len(df))

    if tiene_n_general and tiene_derivacion:
        clave_ruta = list(zip(df[col_n_general], df[col_derivacion]))
    elif tiene_derivacion:
        clave_ruta = df[col_derivacion]
    else:
        # Sin columna de derivación/ruta: se asume que TODA la planilla es
        # una sola ruta (se procesa en el orden en que aparece).
        clave_ruta = [0] * len(df)
    df["_clave_ruta"] = clave_ruta

    if tiene_n_est:
        df["_orden_en_ruta"] = df[col_nruta]
    else:
        df["_orden_en_ruta"] = df["_orden_original"]

    registros: List[dict] = []

    for _clave, grupo in df.groupby("_clave_ruta", sort=False):
        grupo_ordenado = grupo.sort_values("_orden_en_ruta", kind="stable")

        for _, fila in grupo_ordenado.iterrows():
            nombre = fila.get(col_nombre)
            nombre = str(nombre).strip() if pd.notna(nombre) else ""
            derivacion = (str(fila.get(col_derivacion)).strip()
                          if tiene_derivacion and pd.notna(fila.get(col_derivacion)) else "")
            n_ruta = fila.get(col_n_general) if tiene_n_general else None
            n_est = fila.get(col_nruta) if tiene_n_est else None
            vano_adelante = fila.get(col_vano_adelante)

            aportes = calcular_longitudes_cable_poste(
                tipo_conductor_principal1=fila.get(col_conductor_principal1),
                tipo_conductor_principal2=fila.get(col_conductor_principal2),
                armado_primario1=(fila.get(col_armado_primario1)
                                  if col_armado_primario1 in est_df.columns else None),
                armado_primario2=(fila.get(col_armado_primario2)
                                  if col_armado_primario2 in est_df.columns else None),
                armado_secundario1=(fila.get(col_armado_secundario1)
                                    if col_armado_secundario1 in est_df.columns else None),
                armado_secundario2=(fila.get(col_armado_secundario2)
                                    if col_armado_secundario2 in est_df.columns else None),
                vano_adelante=vano_adelante,
            )

            for aporte in aportes:
                registros.append({
                    "nombre_poste": nombre,
                    "derivacion": derivacion,
                    "n_ruta": n_ruta,
                    "n_est": n_est,
                    "lado": aporte["lado"],
                    "rol": aporte["rol"],
                    "cable": aporte["nombre"],
                    "armado_usado": aporte["armado_usado"],
                    "vano_adelante": vano_adelante,
                    "metros": aporte["metros"],
                })

    return pd.DataFrame(registros, columns=columnas_salida)


def sumar_totales_cable(detalle_cable: pd.DataFrame) -> pd.DataFrame:
    """
    Suma el detalle de `extraer_longitudes_cable_planilla` para obtener la
    cantidad TOTAL de cada cable distinto en toda la línea.

    Devuelve un DataFrame con columnas:
        cable | metros_total
    ordenado alfabéticamente por `cable`.
    """
    columnas = ["cable", "metros_total"]
    if detalle_cable is None or len(detalle_cable) == 0:
        return pd.DataFrame(columns=columnas)

    agregado = (detalle_cable.groupby("cable", as_index=False)["metros"]
                .sum()
                .rename(columns={"metros": "metros_total"}))
    agregado = agregado.sort_values("cable", key=lambda s: s.str.lower()).reset_index(drop=True)
    return agregado[columnas]


def cargar_planilla(ruta: str) -> pd.DataFrame:
    """
    Lee una planilla de estructuras PlanillaEstTotal*.XLS con cabecera de dos
    niveles (igual que el resto del notebook). Aislada para depurar la lectura
    del archivo por separado.
    """
    if not os.path.exists(ruta):
        raise FileNotFoundError(f"No se encontró la planilla: {ruta}")
    # engine='xlrd' para .XLS legado; pandas lo autodetecta normalmente, pero
    # se fija explícitamente para evitar sorpresas.
    engine = "xlrd" if ruta.lower().endswith(".xls") else None
    return pd.read_excel(ruta, header=[0, 1], engine=engine)


def extraer_armados_planilla(
    est_df: pd.DataFrame,
    columnas_armado: Sequence[Tuple[str, str]] = COLUMNAS_ARMADO_DEFAULT,
    col_nombre: Tuple[str, str] = COL_NOMBRE_DEFAULT,
    col_nruta: Optional[Tuple[str, str]] = COL_NRUTA_DEFAULT,
    col_derivacion: Optional[Tuple[str, str]] = COL_DERIVACION_DEFAULT,
    col_conductor_principal1: Optional[Tuple[str, str]] = COL_CONDUCTOR_PRINCIPAL1_DEFAULT,
    col_conductor_principal2: Optional[Tuple[str, str]] = COL_CONDUCTOR_PRINCIPAL2_DEFAULT,
    nivel_contaminacion: Optional[str] = None,
) -> pd.DataFrame:
    """
    Convierte la planilla (un poste por fila, hasta 4 armados por fila) en una
    tabla "larga" con un armado por fila:

        nombre_poste | derivacion | n_est | tipo_armado | armado | calibre | aislador

    `tipo_armado` indica de qué columna provino (Primario1, Secundario1, ...),
    útil para auditar. Las celdas vacías se descartan.

    `calibre` es el calibre del conductor de fase asociado a ese armado (ver
    `extraer_calibre_conductor` y `ajustar_fase`), usado más adelante para
    reemplazar el "" que aparece en algunos nombres de material del catálogo.
    Los armados Primario1/Primario2 toman el calibre de 'Conductor Principal1'
    y los armados Secundario1/Secundario2 el de 'Conductor Principal2'. Si el
    conductor respectivo no tiene un valor reconocible, `calibre` queda en None.

    Cuando el texto de 'Tipo Conductor' describe DOS cables unidos por "+"
    (p.ej. "1xACSR 1/0 AWG+1xSM34.5-3x1/0ACSR / Al7N8"), el lado del "+" que
    se usa depende del propio código de armado de esa fila (no es fijo por
    poste): ver `_seleccionar_lado_conductor`. Por eso el calibre se calcula
    aquí dentro del loop de armados, con el código de armado de cada uno,
    en vez de una sola vez por poste.

    `aislador` es el aislador que corresponde a ese armado según el tipo de
    conductor (forrado/desnudo, deducido del propio código de armado), el
    nivel de aislamiento (13.2/34.5 kV, también deducido del código de
    armado) y `nivel_contaminacion` ('Alto'/'Normal', el mismo para toda la
    línea, ver `determinar_nivel_contaminacion_linea`). Ver `determinar_aislador`.
    Queda en None para armados donde no aplica o no se pudo determinar
    (p.ej. retenidas, o códigos sin el prefijo MT/MTF reconocido).
    """
    registros: List[dict] = []
    for idx, fila in est_df.iterrows():
        nombre = fila.get(col_nombre, idx)
        derivacion = fila.get(col_derivacion) if col_derivacion else None
        n_est = fila.get(col_nruta) if col_nruta else None

        texto_conductor_p1 = (fila.get(col_conductor_principal1)
                               if col_conductor_principal1 and col_conductor_principal1 in est_df.columns
                               else None)
        texto_conductor_p2 = (fila.get(col_conductor_principal2)
                               if col_conductor_principal2 and col_conductor_principal2 in est_df.columns
                               else None)

        for col in columnas_armado:
            if col not in est_df.columns:
                continue
            valor = fila.get(col)
            if pd.isna(valor) or str(valor).strip() == "":
                continue
            tipo_armado = col[1] if isinstance(col, tuple) else str(col)
            principal = _principal_para_tipo_armado(tipo_armado)

            codigo_armado = str(valor).strip()
            texto_conductor = (texto_conductor_p1 if principal == 1
                                else (texto_conductor_p2 if principal == 2 else None))
            calibre = extraer_calibre_conductor(texto_conductor, armado=codigo_armado)
            forrado = es_conductor_forrado(codigo_armado)
            nivel_kv = nivel_aislamiento_armado(codigo_armado)
            aislador = determinar_aislador(forrado, nivel_kv, nivel_contaminacion)

            registros.append({
                "nombre_poste": str(nombre).strip() if pd.notna(nombre) else "",
                "derivacion": str(derivacion).strip() if pd.notna(derivacion) else "",
                "n_est": n_est,
                "tipo_armado": tipo_armado,
                "armado": codigo_armado,
                "calibre": calibre,
                "aislador": aislador,
            })
    df = pd.DataFrame(registros,
                      columns=["nombre_poste", "derivacion", "n_est",
                               "tipo_armado", "armado", "calibre", "aislador"])
    df["armado_norm"] = df["armado"].apply(normalizar_codigo_armado)
    return df


def extraer_retenidas_planilla(
    est_df: pd.DataFrame,
    columnas_retenida: Sequence[Tuple[str, str]] = COLUMNAS_RETENIDA_DEFAULT,
    col_nombre: Tuple[str, str] = COL_NOMBRE_DEFAULT,
    col_nruta: Optional[Tuple[str, str]] = COL_NRUTA_DEFAULT,
    col_derivacion: Optional[Tuple[str, str]] = COL_DERIVACION_DEFAULT,
) -> pd.DataFrame:
    """
    Extrae los armados de RETENIDA de cada poste, a partir de las columnas
    tipo RT001, RT002, RT003, ... donde:

        * el NOMBRE de la columna es el código de armado a buscar en el
          catálogo (p.ej. "RT003"), y
        * el VALOR de la celda es la cantidad de esa retenida en ese poste
          (0 = no tiene, 1 = una, 2 = dos, ...).

    Genera el mismo formato "largo" que `extraer_armados_planilla`
    (nombre_poste | derivacion | n_est | tipo_armado | armado | armado_norm),
    repitiendo el armado tantas veces como indique la cantidad, para poder
    reutilizar `calcular_cantidades` sin modificarla: cada repetición aporta
    una vez los materiales del armado, así que N retenidas del mismo tipo
    aportan N veces esos materiales.

    Celdas vacías, NaN, None o con valor 0 se ignoran (no hay retenida).
    """
    registros: List[dict] = []
    for idx, fila in est_df.iterrows():
        nombre = fila.get(col_nombre, idx)
        derivacion = fila.get(col_derivacion) if col_derivacion else None
        n_est = fila.get(col_nruta) if col_nruta else None

        for col in columnas_retenida:
            if col not in est_df.columns:
                continue
            valor = fila.get(col)
            if pd.isna(valor):
                continue
            try:
                cantidad = int(float(valor))
            except (TypeError, ValueError):
                continue
            if cantidad <= 0:
                continue

            codigo_armado = col[1] if isinstance(col, tuple) else str(col)
            for _ in range(cantidad):
                registros.append({
                    "nombre_poste": str(nombre).strip() if pd.notna(nombre) else "",
                    "derivacion": str(derivacion).strip() if pd.notna(derivacion) else "",
                    "n_est": n_est,
                    "tipo_armado": codigo_armado,
                    "armado": codigo_armado,
                    # Las retenidas no tienen un conductor de fase asociado, por
                    # lo que no participan del ajuste de fase (ver `ajustar_fase`)
                    # ni de la selección de aislador (ver `determinar_aislador`).
                    "calibre": None,
                    "aislador": None,
                })

    df = pd.DataFrame(registros,
                      columns=["nombre_poste", "derivacion", "n_est",
                               "tipo_armado", "armado", "calibre", "aislador"])
    df["armado_norm"] = df["armado"].apply(normalizar_codigo_armado)
    return df


# Valores que se consideran "vacíos" en 'Tipo Soporte' / 'Tipo PAT' y por lo
# tanto no cuentan como un tipo válido.
_VALORES_VACIOS_TIPO_SOPORTE = {"", "nan", "none", "0", "-", "n/a", "na"}


def _es_tipo_soporte_valido(valor) -> bool:
    """
    Indica si un valor de 'Tipo Soporte' / 'Tipo PAT' debe considerarse
    válido, es decir que no es None, NaN, 0, ni equivalentes textuales
    vacíos como "-" o "".
    """
    if valor is None:
        return False
    if isinstance(valor, float) and np.isnan(valor):
        return False
    if isinstance(valor, (int, float)) and valor == 0:
        return False
    s = str(valor).strip().lower()
    return s not in _VALORES_VACIOS_TIPO_SOPORTE


def extraer_pat_planilla(
    est_df: pd.DataFrame,
    col_tipo_pat: Tuple[str, str] = COL_TIPO_PAT_DEFAULT,
    col_nombre: Tuple[str, str] = COL_NOMBRE_DEFAULT,
    col_nruta: Optional[Tuple[str, str]] = COL_NRUTA_DEFAULT,
    col_derivacion: Optional[Tuple[str, str]] = COL_DERIVACION_DEFAULT,
) -> pd.DataFrame:
    """
    Extrae el SPT (sistema de puesta a tierra) de cada poste a partir de la
    columna 'Tipo PAT', donde el VALOR de la celda ES el código de armado a
    buscar en el catálogo (p.ej. "SPT001").

    A diferencia de las retenidas (RT00X), aquí no hay una columna por cada
    código posible con una cantidad: hay una única columna 'Tipo PAT' cuyo
    contenido es el código del SPT instalado en ese poste (o vacío/"-" si no
    tiene). Se asume una unidad de SPT por poste cuando el valor es válido.

    Genera el mismo formato "largo" que `extraer_armados_planilla` /
    `extraer_retenidas_planilla` (nombre_poste | derivacion | n_est |
    tipo_armado | armado | armado_norm), para poder reutilizar
    `calcular_cantidades` sin modificarla.

    Celdas vacías, NaN, None, 0 o equivalentes textuales vacíos ("-", "n/a",
    etc., ver `_es_tipo_soporte_valido`) se ignoran (el poste no tiene SPT).
    """
    if col_tipo_pat not in est_df.columns:
        raise KeyError(f"No se encontró la columna {col_tipo_pat!r} en la planilla.")

    registros: List[dict] = []
    for idx, fila in est_df.iterrows():
        nombre = fila.get(col_nombre, idx)
        derivacion = fila.get(col_derivacion) if col_derivacion else None
        n_est = fila.get(col_nruta) if col_nruta else None

        valor = fila.get(col_tipo_pat)
        if not _es_tipo_soporte_valido(valor):
            continue

        codigo_armado = str(valor).strip()
        registros.append({
            "nombre_poste": str(nombre).strip() if pd.notna(nombre) else "",
            "derivacion": str(derivacion).strip() if pd.notna(derivacion) else "",
            "n_est": n_est,
            "tipo_armado": "SPT",
            "armado": codigo_armado,
            # El SPT no tiene conductor de fase asociado ni participa de la
            # selección de aislador (igual que las retenidas).
            "calibre": None,
            "aislador": None,
        })

    df = pd.DataFrame(registros,
                      columns=["nombre_poste", "derivacion", "n_est",
                               "tipo_armado", "armado", "calibre", "aislador"])
    df["armado_norm"] = df["armado"].apply(normalizar_codigo_armado)
    return df


# =====================================================================
#  3-bis. CONTEO DE POSTES POR TIPO DE SOPORTE
# =====================================================================

def contar_tipos_soporte(
    est_df: pd.DataFrame,
    col_tipo_soporte: Tuple[str, str] = COL_TIPO_SOPORTE_DEFAULT,
    col_nombre: Tuple[str, str] = COL_NOMBRE_DEFAULT,
    verbose: bool = True,
) -> pd.DataFrame:
    """
    Cuenta la cantidad total de postes de cada 'Tipo Soporte' distinto.

    Un mismo poste ('Nombre Est.') puede aparecer repetido en varias filas de
    la planilla (p. ej. una fila de llegada y otra de salida). Para no
    duplicar el conteo, por cada nombre de poste se toma únicamente la
    PRIMERA fila (en el orden en que aparece en la planilla) cuyo
    'Tipo Soporte' sea válido:

      * Si la primera aparición del poste ya trae un 'Tipo Soporte' válido,
        esa es la que se cuenta.
      * Si la primera aparición trae un valor vacío/None/NaN/0/"-", el poste
        se descarta por completo para el conteo (no se buscan apariciones
        posteriores), siguiendo el criterio acordado con el usuario.

    Parámetros
    ----------
    est_df : pd.DataFrame
        DataFrame de la planilla de estructuras ya cargado (p.ej. est_df).
    col_tipo_soporte : tupla
        Columna (multi-índice) donde está el tipo de soporte/poste.
    col_nombre : tupla
        Columna (multi-índice) donde está el nombre de la estructura/poste.
    verbose : bool
        Imprime un resumen del conteo.

    Devuelve
    --------
    pd.DataFrame con columnas:
        tipo_soporte | cantidad
    ordenado alfabéticamente por 'tipo_soporte'.
    """
    if col_tipo_soporte not in est_df.columns:
        raise KeyError(f"No se encontró la columna {col_tipo_soporte!r} en la planilla.")
    if col_nombre not in est_df.columns:
        raise KeyError(f"No se encontró la columna {col_nombre!r} en la planilla.")

    vistos: set = set()          # nombres de poste ya resueltos (contados o descartados)
    conteo: Dict[str, int] = {}

    for _, fila in est_df.iterrows():
        nombre = fila.get(col_nombre)
        nombre = str(nombre).strip() if pd.notna(nombre) else ""

        if nombre in vistos:
            # Ya se resolvió este poste con su primera aparición: se ignora.
            continue
        vistos.add(nombre)

        valor_tipo = fila.get(col_tipo_soporte)
        if not _es_tipo_soporte_valido(valor_tipo):
            # Primera aparición sin tipo de soporte válido -> se descarta el poste.
            continue

        tipo = str(valor_tipo).strip()
        conteo[tipo] = conteo.get(tipo, 0) + 1

    df = (pd.DataFrame(
                [{"tipo_soporte": k, "cantidad": v} for k, v in conteo.items()],
                columns=["tipo_soporte", "cantidad"])
          .sort_values("tipo_soporte", key=lambda s: s.str.lower())
          .reset_index(drop=True))

    if verbose:
        total_postes = len(df) and df["cantidad"].sum()
        print(f"[postes] Tipos de soporte distintos: {len(df)}  "
              f"(postes contados: {total_postes})")

    return df


# =====================================================================
#  4. CÁLCULO DE CANTIDADES TOTALES
# =====================================================================

def _construir_indice_catalogo(catalogo) -> dict:
    """
    Construye un índice  { nucleo : clave_interna_catalogo }  donde nucleo es
    el resultado de normalizar_codigo_armado sobre el código original del
    catálogo, y clave_interna es la clave con la que catalogo.materiales indexa
    sus multiplicadores.

    Permite resolver en O(1) cualquier código de la planilla contra el catálogo
    comparando únicamente núcleos alfanuméricos, sin aliases manuales.
    """
    return {normalizar_codigo_armado(orig): clave
            for clave, orig in catalogo.armados.items()}


def calcular_cantidades(
    armados_planilla: pd.DataFrame,
    catalogo: Catalogo,
    totales_cable: Optional[pd.DataFrame] = None,
    verbose: bool = True,
) -> Dict[str, pd.DataFrame]:
    """
    Suma los materiales de todos los armados de todos los postes.

    El match entre planilla y catálogo se hace por núcleo alfanumérico: dos
    códigos son equivalentes si, al eliminar cualquier separador o sufijo,
    sus letras y números coinciden. Así "MTF331-1", "MTF 331-1", "MTF_331_1"
    y "MTF331-1 (2)" apuntan todos al mismo armado del catálogo.

    Devuelve un diccionario con DataFrames:

      'totales'                    -> codigo | material | unidad | cantidad_total
      'no_encontrados'             -> armados cuyo núcleo no existe en el catálogo
      'detalle'                    -> aporte de cada armado a cada material (trazabilidad)
      'fase_sin_calibre'           -> materiales con "" que no se pudieron ajustar por
                                       no haberse podido determinar el calibre del
                                       conductor del poste/armado correspondiente
      'aislador_sin_determinar'    -> armados MT/MTF donde no se pudo decidir qué
                                       aislador corresponde (falta nivel de
                                       aislamiento reconocible o nivel de
                                       contaminación de la línea)
      'aislador_sin_correspondencia' -> armados donde SÍ se determinó el aislador
                                       a usar, pero el catálogo no tiene ese
                                       renglón para ese armado (ver más abajo)

    Ajuste de fase
    --------------
    Si el nombre de un material contiene el marcador "" (ver `ajustar_fase`),
    se reemplaza por el calibre del conductor de fase asociado a ese armado
    (columna `calibre` de `armados_planilla`, ver `extraer_armados_planilla`).
    Como el mismo código de material puede terminar representando nombres
    distintos según el calibre de cada poste, los totales se acumulan por el
    nombre YA ajustado (no por la clave interna del catálogo) para no mezclar
    cantidades de calibres distintos bajo un mismo renglón.

    Selección de aislador (filtrado, NO se suma aparte)
    ----------------------------------------------------
    El catálogo (Cantidades_de_postes.xlsx) trae, para un mismo armado, más
    de un renglón de aislador "alternativo" con su propia cantidad (p.ej. las
    hojas "AFINIAAIR-E (Forradas/Desnudas - 13,2 kV" traen tanto "AISLADOR
    PORCELANA TIPO POSTE 13,2 kV (ANSI-57-1)" como "AISLADOR COMPUESTO
    HIBRIDO 13,2 kV.", cada una con su propia cantidad para ese armado). El
    aislador NO se suma aparte: se recorre el catálogo igual que cualquier
    otro material, pero cuando un renglón pertenece a una de las familias de
    aislador "alternativas" (ver `_familia_aislador`), solo se incluye si su
    familia coincide con el aislador determinado para ese armado (columna
    `aislador` de `armados_planilla`, ver `determinar_aislador`); si no
    coincide, se descarta esa fila para ese armado. Los demás materiales del
    armado (crucetas, grapas, pines, etc.) se incluyen siempre, sin filtrar.

    Si el aislador determinado no tiene NINGÚN renglón correspondiente en el
    catálogo para ese armado (p.ej. hoy el catálogo no trae ningún "AISLADOR
    LINEPOST 66KV..." para la combinación forrado + alta contaminación +
    34,5 kV), el armado queda sin ese material y se reporta en
    'aislador_sin_correspondencia' en vez de inventar una cantidad.

    Cantidades de cable (conductores)
    ----------------------------------
    Si se pasa `totales_cable` (ver `sumar_totales_cable`, con columnas
    'cable'/'metros_total'), esas cantidades se agregan a la tabla
    'totales' devuelta, con unidad "m" y sin código (no provienen del
    catálogo de armados sino del cálculo de vanos). Así quedan incluidas en
    las cantidades finales exportadas a Excel junto con el resto de
    materiales.
    """
    # Índice nucleo -> clave_interna del catálogo (construido una sola vez)
    indice = _construir_indice_catalogo(catalogo)

    # Acumuladores
    totales: Dict[str, float] = {}
    info_efectiva: Dict[str, dict] = {}
    detalle_rows: List[dict] = []
    faltantes: Dict[str, dict] = {}
    sin_calibre: Dict[Tuple[str, str], dict] = {}
    aislador_sin_determinar: Dict[str, dict] = {}
    aislador_sin_correspondencia: Dict[Tuple[str, str], dict] = {}

    for _, fila in armados_planilla.iterrows():
        armado_orig = fila["armado"]
        nucleo = fila["armado_norm"]          # ya es el núcleo alfanumérico
        calibre = fila.get("calibre")
        # Según el dtype de la columna (puede variar entre versiones de
        # pandas), un calibre ausente puede llegar como None, NaN o <NA>.
        if calibre is None or (not isinstance(calibre, str) and pd.isna(calibre)):
            calibre = None

        # --- Aislador determinado para este armado (ver determinar_aislador) ---
        aislador_val = fila.get("aislador")
        if aislador_val is None or (not isinstance(aislador_val, str) and pd.isna(aislador_val)):
            aislador_val = None
        familia_aislador_esperada = _familia_aislador(aislador_val) if aislador_val else None
        if aislador_val is None and es_conductor_forrado(armado_orig) is not None:
            # Es un armado tipo MT/MTF (aplica selección de aislador) pero no
            # se pudo determinar (falta nivel de aislamiento reconocible o
            # falta el nivel de contaminación de la línea).
            info_ad = aislador_sin_determinar.setdefault(
                armado_orig, {"armado": armado_orig, "veces": 0})
            info_ad["veces"] += 1

        clave_cat = indice.get(nucleo)

        if clave_cat is None:
            info = faltantes.setdefault(
                armado_orig, {"armado": armado_orig, "nucleo": nucleo, "veces": 0})
            info["veces"] += 1
            continue

        familia_incluida = False
        familias_presentes_armado = set()
        for clave_mat, mult in catalogo.materiales[clave_cat].items():
            info_original = catalogo.info_material[clave_mat]
            nombre_original = info_original["nombre"]

            # --- Filtro de aislador: descartar la(s) alternativa(s) que no
            # correspondan al aislador determinado para este armado ---
            familia_material = _familia_aislador(nombre_original)
            if familia_material is not None:
                # Se registra que el catálogo SÍ modela una alternativa de
                # aislador para este armado (con cantidad > 0, ya que
                # `cargar_catalogo` descarta los multiplicadores en 0), sin
                # importar si es la que corresponde o no. Sirve para no
                # advertir "sin correspondencia" cuando el armado
                # sencillamente no lleva ninguna de estas alternativas (p.ej.
                # porque usa un aislador tipo pin fijo en su lugar).
                familias_presentes_armado.add(familia_material)
                if familia_material != familia_aislador_esperada:
                    continue
                familia_incluida = True

            nombre_ajustado = ajustar_nombre_material_fase(nombre_original, calibre)

            if nombre_ajustado != nombre_original:
                clave_efectiva = f"{clave_mat}::FASE::{nombre_ajustado}"
            elif MARCADOR_CALIBRE_FASE in nombre_original and not calibre:
                # No se pudo determinar el calibre: se deja el "" visible y se
                # registra para el reporte de advertencias.
                clave_efectiva = clave_mat
                info_sc = sin_calibre.setdefault(
                    (fila["nombre_poste"], armado_orig),
                    {"nombre_poste": fila["nombre_poste"], "armado": armado_orig,
                     "material": nombre_original, "veces": 0})
                info_sc["veces"] += 1
            else:
                clave_efectiva = clave_mat

            if clave_efectiva not in info_efectiva:
                info_efectiva[clave_efectiva] = {
                    "codigo": info_original["codigo"],
                    "nombre": nombre_ajustado,
                    "unidad": info_original["unidad"],
                }

            totales[clave_efectiva] = totales.get(clave_efectiva, 0.0) + mult
            detalle_rows.append({
                "nombre_poste": fila["nombre_poste"],
                "armado": armado_orig,
                "material": nombre_ajustado,
                "codigo": info_original["codigo"],
                "cantidad": mult,
            })

        if (familia_aislador_esperada is not None and not familia_incluida
                and familias_presentes_armado):
            # El catálogo SÍ modela alguna alternativa de aislador para este
            # armado (con cantidad > 0), pero no la que corresponde según la
            # contaminación/nivel de aislamiento determinados (ver docstring:
            # hoy pasa con "forrado + alta contaminación + 34,5 kV" ->
            # LINEPOST, que no existe en Cantidades_de_postes.xlsx). Si el
            # catálogo no modela NINGUNA alternativa para este armado (p.ej.
            # porque usa un aislador tipo pin fijo en su lugar), no se
            # reporta nada: es un caso legítimo, no un dato faltante.
            info_sc2 = aislador_sin_correspondencia.setdefault(
                (armado_orig, aislador_val),
                {"armado": armado_orig, "aislador_esperado": aislador_val, "veces": 0})
            info_sc2["veces"] += 1

    # --- Tabla de totales ---
    filas_tot = []
    for clave_efectiva, cant in totales.items():
        info = info_efectiva[clave_efectiva]
        filas_tot.append({
            "codigo": info["codigo"],
            "material": info["nombre"],
            "unidad": info["unidad"],
            "cantidad_total": cant,
        })
    # --- Agregar cantidades de cable (conductores), si se proporcionaron ---
    if totales_cable is not None and len(totales_cable):
        for _, r in totales_cable.iterrows():
            filas_tot.append({
                "codigo": "",
                "material": r["cable"],
                "unidad": "m",
                "cantidad_total": float(r["metros_total"]),
            })

    df_totales = (pd.DataFrame(filas_tot,
                               columns=["codigo", "material", "unidad", "cantidad_total"])
                  .sort_values("material", key=lambda s: s.str.lower())
                  .reset_index(drop=True))

    # --- Tabla de no encontrados ---
    df_faltantes = (pd.DataFrame(list(faltantes.values()),
                                 columns=["armado", "nucleo", "veces"])
                    .sort_values("armado")
                    .reset_index(drop=True))

    df_detalle = pd.DataFrame(
        detalle_rows,
        columns=["nombre_poste", "armado", "material", "codigo", "cantidad"])

    # --- Tabla de ajustes de fase sin calibre resuelto ---
    df_sin_calibre = pd.DataFrame(
        list(sin_calibre.values()),
        columns=["nombre_poste", "armado", "material", "veces"])
    if len(df_sin_calibre):
        df_sin_calibre = df_sin_calibre.sort_values(
            ["nombre_poste", "armado"]).reset_index(drop=True)

    # --- Tabla de aisladores sin determinar (falta info para decidir) ---
    df_aislador_sin_det = pd.DataFrame(
        list(aislador_sin_determinar.values()),
        columns=["armado", "veces"])
    if len(df_aislador_sin_det):
        df_aislador_sin_det = df_aislador_sin_det.sort_values("armado").reset_index(drop=True)

    # --- Tabla de aisladores sin correspondencia en el catálogo ---
    df_aislador_sin_corr = pd.DataFrame(
        list(aislador_sin_correspondencia.values()),
        columns=["armado", "aislador_esperado", "veces"])
    if len(df_aislador_sin_corr):
        df_aislador_sin_corr = df_aislador_sin_corr.sort_values(
            ["armado", "aislador_esperado"]).reset_index(drop=True)

    if verbose:
        print(f"[calculo] Materiales totales distintos: {len(df_totales)}")
        if len(df_aislador_sin_det):
            print(f"[calculo] ⚠ Armados MT/MTF sin aislador determinado "
                  f"({len(df_aislador_sin_det)}):")
            for _, r in df_aislador_sin_det.iterrows():
                print(f"          - {r['armado']!r}  (aparece {r['veces']} vez/veces)")
            print("          Revisa que el código de armado termine en '-1'/'-2' y que "
                  "se haya podido determinar el nivel de contaminación de la línea.")
        if len(df_aislador_sin_corr):
            print(f"[calculo] ⚠ Aislador determinado SIN correspondencia en el catálogo "
                  f"({len(df_aislador_sin_corr)}):")
            for _, r in df_aislador_sin_corr.iterrows():
                print(f"          - armado={r['armado']!r}  se esperaba {r['aislador_esperado']!r}  "
                      f"(aparece {r['veces']} vez/veces)")
            print("          Falta ese renglón (con su cantidad) en Cantidades_de_postes.xlsx "
                  "para ese armado.")
        if len(df_sin_calibre):
            print(f"[calculo] ⚠ Materiales con \"\" sin calibre resuelto "
                  f"({len(df_sin_calibre)}):")
            for _, r in df_sin_calibre.iterrows():
                print(f"          - poste={r['nombre_poste']!r} armado={r['armado']!r} "
                      f"material={r['material']!r} (aparece {r['veces']} vez/veces)")
            print("          Revisa el 'Tipo Conductor' del poste: no se reconoció "
                  "ningún material (AAAC/ACSR).")
        if len(df_faltantes):
            print(f"[calculo] ⚠ Armados SIN correspondencia en el catálogo "
                  f"({len(df_faltantes)}):")
            for _, r in df_faltantes.iterrows():
                print(f"          - {r['armado']!r}  núcleo={r['nucleo']!r}  (aparece {r['veces']} vez/veces)")
            print("          Añádelos al diccionario `alias` o al catálogo.")
        else:
            print("[calculo] ✅ Todos los armados de la planilla se encontraron.")

    return {"totales": df_totales,
            "no_encontrados": df_faltantes,
            "detalle": df_detalle,
            "fase_sin_calibre": df_sin_calibre,
            "aislador_sin_determinar": df_aislador_sin_det,
            "aislador_sin_correspondencia": df_aislador_sin_corr}




# =====================================================================
#  5. EXPORTACIÓN A EXCEL
# =====================================================================

def exportar_cantidades_excel(resultado: Dict[str, pd.DataFrame],
                              ruta_salida: str,
                              incluir_detalle: bool = True,
                              verbose: bool = True) -> str:
    """
    Escribe el resultado en un .xlsx con formato profesional:

      Hoja 'Cantidades'          -> material y cantidad total (entregable principal;
                                     el aislador de cada armado ya viene filtrado
                                     según contaminación/nivel de aislamiento/tipo
                                     de conductor, con la cantidad tomada del propio
                                     catálogo, ver `calcular_cantidades`)
      Hoja 'Tipos de Soporte'    -> cantidad total de postes por tipo de soporte
      Hoja 'Contaminación'       -> nivel de contaminación de la línea y coordenada usada
      Hoja 'Armados no hallados' -> trazabilidad de lo que no se pudo mapear
      Hoja 'Fase sin calibre' (si aplica) -> materiales con "" sin calibre resuelto
      Hoja 'Aislador sin determinar' (si aplica) -> armados MT/MTF sin aislador resuelto
      Hoja 'Aislador sin catálogo' (si aplica) -> aislador determinado pero sin
                                     ese renglón en Cantidades_de_postes.xlsx
      Hoja 'Detalle' (opcional)  -> aporte poste×armado×material
      Hoja 'Detalle Cable' (opcional, si aplica) -> aporte poste×vano×cable
                                     (trazabilidad del cálculo de longitudes
                                     de cable; ver
                                     `extraer_longitudes_cable_planilla`). Las
                                     cantidades totales de cable YA están
                                     incluidas en la hoja 'Cantidades'
                                     (ver `calcular_cantidades`).

    Devuelve la ruta del archivo escrito.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    os.makedirs(os.path.dirname(os.path.abspath(ruta_salida)) or ".", exist_ok=True)

    df_tot = resultado["totales"].copy()
    df_tot.columns = ["Código", "Material", "Unidad", "Cantidad total"]
    df_falt = resultado["no_encontrados"].copy()
    if len(df_falt):
        df_falt.columns = ["Armado (planilla)", "Núcleo buscado", "Veces"]
    df_tipos = resultado.get("tipos_soporte")
    if df_tipos is not None and len(df_tipos):
        df_tipos = df_tipos.copy()
        df_tipos.columns = ["Tipo Soporte", "Cantidad"]
    df_fase_sc = resultado.get("fase_sin_calibre")
    if df_fase_sc is not None and len(df_fase_sc):
        df_fase_sc = df_fase_sc.copy()
        df_fase_sc.columns = ["Poste", "Armado", "Material", "Veces"]
    df_aislador_sd = resultado.get("aislador_sin_determinar")
    if df_aislador_sd is not None and len(df_aislador_sd):
        df_aislador_sd = df_aislador_sd.copy()
        df_aislador_sd.columns = ["Armado", "Veces"]
    df_aislador_sc = resultado.get("aislador_sin_correspondencia")
    if df_aislador_sc is not None and len(df_aislador_sc):
        df_aislador_sc = df_aislador_sc.copy()
        df_aislador_sc.columns = ["Armado", "Aislador esperado", "Veces"]
    contaminacion = resultado.get("contaminacion")

    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        df_tot.to_excel(writer, sheet_name="Cantidades", index=False)
        if df_tipos is not None and len(df_tipos):
            df_tipos.to_excel(writer, sheet_name="Tipos de Soporte", index=False)
        if contaminacion is not None:
            filas_contam = [
                {"Campo": "Nivel de contaminación de la línea", "Valor": contaminacion.get("nivel")},
                {"Campo": "Latitud (1er poste)", "Valor": contaminacion.get("lat")},
                {"Campo": "Longitud (1er poste)", "Valor": contaminacion.get("lon")},
                {"Campo": "X planilla (1er poste)", "Valor": contaminacion.get("x")},
                {"Campo": "Y planilla (1er poste)", "Valor": contaminacion.get("y")},
                {"Campo": "Observación", "Valor": contaminacion.get("mensaje") or ""},
            ]
            pd.DataFrame(filas_contam).to_excel(writer, sheet_name="Contaminación", index=False)
        if len(df_falt):
            df_falt.to_excel(writer, sheet_name="Armados no hallados", index=False)
        else:
            pd.DataFrame({"Estado": ["Todos los armados fueron encontrados ✅"]}) \
                .to_excel(writer, sheet_name="Armados no hallados", index=False)
        if df_fase_sc is not None and len(df_fase_sc):
            df_fase_sc.to_excel(writer, sheet_name="Fase sin calibre", index=False)
        if df_aislador_sd is not None and len(df_aislador_sd):
            df_aislador_sd.to_excel(writer, sheet_name="Aislador sin determinar", index=False)
        if df_aislador_sc is not None and len(df_aislador_sc):
            df_aislador_sc.to_excel(writer, sheet_name="Aislador sin catálogo", index=False)
        if incluir_detalle and len(resultado.get("detalle", [])):
            det = resultado["detalle"].copy()
            det.columns = ["Poste", "Armado", "Material", "Código", "Cantidad"]
            det.to_excel(writer, sheet_name="Detalle", index=False)
        if incluir_detalle and len(resultado.get("detalle_cable", [])):
            det_cable = resultado["detalle_cable"].copy()
            det_cable.columns = ["Poste", "Derivación", "N° Ruta", "N° Est.",
                                  "Lado", "Rol", "Cable", "Armado usado",
                                  "Vano Adelante (m)", "Metros"]
            det_cable.to_excel(writer, sheet_name="Detalle Cable", index=False)

        wb = writer.book
        # --- Formato de cabeceras y anchos ---
        encabezado_fill = PatternFill("solid", fgColor="1F4E78")
        encabezado_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        celda_font = Font(name="Arial", size=10)
        centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
        borde = Border(*[Side(style="thin", color="D9D9D9")] * 4)

        for ws in wb.worksheets:
            for col_idx, col_cells in enumerate(ws.iter_cols(), start=1):
                max_len = 0
                for i, celda in enumerate(col_cells):
                    celda.border = borde
                    if i == 0:  # cabecera
                        celda.fill = encabezado_fill
                        celda.font = encabezado_font
                        celda.alignment = centro
                    else:
                        celda.font = celda_font
                    valor = "" if celda.value is None else str(celda.value)
                    max_len = max(max_len, len(valor))
                ws.column_dimensions[get_column_letter(col_idx)].width = \
                    min(max(12, max_len + 2), 60)
            ws.freeze_panes = "A2"

    if verbose:
        print(f"[export] Archivo escrito: {ruta_salida}")
    return ruta_salida


# =====================================================================
#  6. ORQUESTADOR DE ALTO NIVEL
# =====================================================================

def generar_cantidades_materiales(
    est_df: pd.DataFrame,
    ruta_catalogo: str,
    hojas_catalogo: Optional[Sequence[str]] = None,
    columnas_armado: Sequence[Tuple[str, str]] = COLUMNAS_ARMADO_DEFAULT,
    columnas_retenida: Sequence[Tuple[str, str]] = COLUMNAS_RETENIDA_DEFAULT,
    col_tipo_soporte: Tuple[str, str] = COL_TIPO_SOPORTE_DEFAULT,
    col_tipo_pat: Tuple[str, str] = COL_TIPO_PAT_DEFAULT,
    col_conductor_principal1: Optional[Tuple[str, str]] = COL_CONDUCTOR_PRINCIPAL1_DEFAULT,
    col_conductor_principal2: Optional[Tuple[str, str]] = COL_CONDUCTOR_PRINCIPAL2_DEFAULT,
    col_topografia_x: Tuple[str, str] = COL_TOPO_X_DEFAULT,
    col_topografia_y: Tuple[str, str] = COL_TOPO_Y_DEFAULT,
    epsg_planilla: str = EPSG_PLANILLA_DEFAULT,
    nivel_contaminacion_forzado: Optional[str] = None,
    incluir_retenidas: bool = True,
    incluir_pat: bool = True,
    incluir_cable: bool = True,
    col_vano_adelante: Tuple[str, str] = COL_VANO_ADELANTE_DEFAULT,
    ruta_salida: str = "Cantidades_totales_proyecto.xlsx",
    incluir_detalle: bool = True,
    verbose: bool = True,
) -> Dict[str, object]:
    """
    Orquesta el flujo completo, ejecutando cada etapa por separado y atrapando
    su error para que sea fácil saber DÓNDE falló.

    Parámetros
    ----------
    est_df : pd.DataFrame
        DataFrame de la planilla de estructuras ya cargado (p.ej. est_v_max).
        Se usa directamente sin volver a leer el archivo desde disco.
    ruta_catalogo : str
        Ruta al archivo Cantidades_de_postes.xlsx en Drive montado.
    col_tipo_soporte : tupla
        Columna (multi-índice) con el tipo de soporte/poste, usada para el
        conteo de postes por tipo (hoja 'Tipos de Soporte' en la salida).
    col_conductor_principal1 / col_conductor_principal2 : tuplas
        Columnas 'Tipo Conductor' de 'Conductor Principal1'/'Conductor
        Principal2', usadas para el ajuste de fase (reemplazo del "" en
        nombres de material por el calibre real del conductor de cada
        poste/armado). Ver `extraer_calibre_conductor` y `ajustar_fase`.
    col_topografia_x / col_topografia_y : tuplas
        Columnas con la coordenada (X, Y) de cada poste ('Topografía'),
        usadas para determinar el nivel de contaminación de la línea a
        partir del primer poste con coordenada válida. Ver
        `determinar_nivel_contaminacion_linea`.
    epsg_planilla : str
        CRS de las coordenadas X/Y de la planilla (ver advertencia en
        `EPSG_PLANILLA_DEFAULT`: por defecto se asume MAGNA-SIRGAS/UTM 18N,
        pero debe verificarse contra el proyecto real).
    nivel_contaminacion_forzado : str, opcional
        Si se indica ('Alto' o 'Normal'), se usa este nivel de contaminación
        directamente y NO se intenta determinar a partir de las coordenadas
        (útil si ya se conoce el nivel, o para pruebas).
    columnas_retenida : lista de tuplas
        Columnas RT00X (retenidas) a incluir en el cálculo de materiales.
    incluir_retenidas : bool
        Si es False, omite por completo el aporte de retenidas (equivalente
        al comportamiento anterior a esta función).
    col_tipo_pat : tupla
        Columna 'Tipo PAT' con el código de SPT (sistema de puesta a tierra)
        instalado en cada poste (p.ej. "SPT001"). Ver `extraer_pat_planilla`.
    incluir_pat : bool
        Si es False, omite por completo el aporte de SPT/PAT.
    incluir_cable : bool
        Si es False, omite por completo el cálculo de cantidades de cable
        (conductores) por vano. Si es True (por defecto), calcula la
        longitud total de cada cable distinto (fase, mensajero o cable
        normal) de toda la línea a partir de 'Tipo Conductor', los armados
        y 'Vano Adelante' de cada poste (ver
        `extraer_longitudes_cable_planilla`) y la incluye en 'totales'.
    col_vano_adelante : tupla
        Columna 'Vano Adelante' (grupo 'Topografía') con la distancia en
        metros entre cada poste y el siguiente de su misma ruta.

    Aislador
    --------
    A partir del nivel de contaminación de la línea (determinado una sola vez
    a partir del primer poste) y, por cada armado, del tipo de conductor
    (forrado/desnudo) y nivel de aislamiento (13.2/34.5 kV) deducidos de su
    propio código, se determina el aislador a instalar (ver
    `determinar_aislador`), el aislador correcto ya queda filtrado dentro de
    'totales' (mismo mecanismo de `calcular_cantidades`: se recorren los
    materiales del catálogo para ese armado y solo se conserva la alternativa
    de aislador cuya familia coincide con la determinada).

    Devuelve un dict con las claves:
        'catalogo', 'armados', 'retenidas', 'pat', 'totales', 'no_encontrados',
        'detalle', 'fase_sin_calibre', 'aislador_sin_determinar',
        'aislador_sin_correspondencia', 'contaminacion', 'tipos_soporte',
        'detalle_cable', 'totales_cable', 'ruta_salida'
    """
    etapa = "inicio"
    try:
        # --- Etapa 1: cargar catálogo desde Drive montado ---
        etapa = "carga del catálogo"
        catalogo = cargar_catalogo(ruta_catalogo, hojas=hojas_catalogo, verbose=verbose)

        # --- Etapa 1-bis: nivel de contaminación de la línea (1er poste) ---
        etapa = "determinación del nivel de contaminación"
        if nivel_contaminacion_forzado is not None:
            contaminacion = {"nivel": nivel_contaminacion_forzado, "lat": None, "lon": None,
                              "x": None, "y": None, "mensaje": "Nivel forzado por parámetro."}
            if verbose:
                print(f"[contaminacion] Nivel forzado por parámetro: {nivel_contaminacion_forzado}")
        else:
            contaminacion = determinar_nivel_contaminacion_linea(
                est_df, col_x=col_topografia_x, col_y=col_topografia_y,
                epsg_origen=epsg_planilla, verbose=verbose)
        nivel_contaminacion = contaminacion.get("nivel")

        # --- Etapa 2: extraer armados del DataFrame ya en memoria ---
        etapa = "extracción de armados"
        armados = extraer_armados_planilla(
            est_df, columnas_armado=columnas_armado,
            col_conductor_principal1=col_conductor_principal1,
            col_conductor_principal2=col_conductor_principal2,
            nivel_contaminacion=nivel_contaminacion)
        if verbose:
            print(f"[planilla] {armados['nombre_poste'].nunique()} postes, "
                  f"{len(armados)} armados instalados en total.")

        # --- Etapa 2-bis: extraer retenidas (RT00X) y unirlas a los armados ---
        retenidas = pd.DataFrame(
            columns=["nombre_poste", "derivacion", "n_est",
                     "tipo_armado", "armado", "calibre", "aislador", "armado_norm"])
        if incluir_retenidas:
            etapa = "extracción de retenidas"
            retenidas = extraer_retenidas_planilla(est_df, columnas_retenida=columnas_retenida)
            if verbose:
                print(f"[planilla] {retenidas['nombre_poste'].nunique()} postes con retenida, "
                      f"{len(retenidas)} retenidas instaladas en total.")
            armados = pd.concat([armados, retenidas], ignore_index=True)

        # --- Etapa 2-ter: extraer SPT/PAT ('Tipo PAT') y unirlas a los armados ---
        pat = pd.DataFrame(
            columns=["nombre_poste", "derivacion", "n_est",
                     "tipo_armado", "armado", "calibre", "aislador", "armado_norm"])
        if incluir_pat:
            etapa = "extracción de SPT (Tipo PAT)"
            pat = extraer_pat_planilla(est_df, col_tipo_pat=col_tipo_pat)
            if verbose:
                print(f"[planilla] {pat['nombre_poste'].nunique()} postes con SPT, "
                      f"{len(pat)} SPT instalados en total.")
            armados = pd.concat([armados, pat], ignore_index=True)

        # --- Etapa 2-quater: calcular cantidades de cable (conductores) por vano ---
        detalle_cable = pd.DataFrame(
            columns=["nombre_poste", "derivacion", "n_ruta", "n_est", "lado",
                     "rol", "cable", "armado_usado", "vano_adelante", "metros"])
        totales_cable = pd.DataFrame(columns=["cable", "metros_total"])
        if incluir_cable:
            etapa = "cálculo de longitudes de cable"
            detalle_cable = extraer_longitudes_cable_planilla(
                est_df,
                col_conductor_principal1=col_conductor_principal1,
                col_conductor_principal2=col_conductor_principal2,
                col_armado_primario1=columnas_armado[0],
                col_armado_primario2=columnas_armado[1],
                col_armado_secundario1=columnas_armado[2],
                col_armado_secundario2=columnas_armado[3],
                col_vano_adelante=col_vano_adelante,
            )
            totales_cable = sumar_totales_cable(detalle_cable)
            if verbose:
                print(f"[cable] {len(totales_cable)} tipos de cable distintos, "
                      f"{detalle_cable['metros'].sum():.1f} m en total.")

        # --- Etapa 3: calcular cantidades (armados + retenidas + SPT juntos) ---
        etapa = "cálculo de cantidades"
        resultado = calcular_cantidades(armados, catalogo, totales_cable=totales_cable,
                                        verbose=verbose)
        resultado["contaminacion"] = contaminacion
        resultado["detalle_cable"] = detalle_cable
        resultado["totales_cable"] = totales_cable

        # --- Etapa 3-ter: contar postes por tipo de soporte ---
        etapa = "conteo de tipos de soporte"
        resultado["tipos_soporte"] = contar_tipos_soporte(
            est_df, col_tipo_soporte=col_tipo_soporte, verbose=verbose)

        # --- Etapa 4: exportar ---
        etapa = "exportación a Excel"
        exportar_cantidades_excel(resultado, ruta_salida,
                                  incluir_detalle=incluir_detalle, verbose=verbose)

    except Exception as e:
        print(f"\n❌ Falló la etapa: {etapa}\n   {type(e).__name__}: {e}")
        raise

    return {
        "catalogo": catalogo,
        "armados": armados,
        "retenidas": retenidas,
        "pat": pat,
        "detalle_cable": resultado["detalle_cable"],
        "totales_cable": resultado["totales_cable"],
        "totales": resultado["totales"],
        "no_encontrados": resultado["no_encontrados"],
        "detalle": resultado["detalle"],
        "fase_sin_calibre": resultado["fase_sin_calibre"],
        "aislador_sin_determinar": resultado["aislador_sin_determinar"],
        "aislador_sin_correspondencia": resultado["aislador_sin_correspondencia"],
        "contaminacion": resultado["contaminacion"],
        "tipos_soporte": resultado["tipos_soporte"],
        "ruta_salida": ruta_salida,
    }
